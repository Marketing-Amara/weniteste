#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Funil Comercial Weni x Contatos x Pedidos -- Amara NZero
========================================================
Le tres fontes e gera o painel HTML + a planilha XLSX:
  - conversas  (message_export*.xlsx)  -- exportacao da Weni / API
  - contatos   (contact_export*.xlsx)  -- base com vendedor, CNPJ, UF
  - pedidos    (Relatorio*.xlsx / pedidos*.xlsx) -- relatorio de pedidos (cruzado por CNPJ)

USO:
    python funil_weni.py --input entradas --output saidas

Procura na pasta --input:
  message_export*.xlsx, contact_export*.xlsx, e um arquivo de pedidos
  (nome contendo "relatorio", "pedido" ou "order"). O de pedidos e opcional:
  sem ele, "Fecharam pedido" fica vazio (nao da pra confirmar por CNPJ).
"""
import argparse, collections, glob, json, os, re, sys, unicodedata
from datetime import datetime, timedelta

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    sys.exit("Dependencia faltando: %s\nRode:  pip install -r requirements.txt" % e)

PRODUCTION_CHANNEL = "Amara NZero"

# ---------------------------------------------------------------------------
# Contatos a EXCLUIR da analise (equipe interna, vendedores como atendentes,
# contas de teste). Edite esta lista conforme necessario.
# ---------------------------------------------------------------------------
EXCLUDE_NAMES = [
    "Maria Dantas V. Ferreira", "Maria Dantas", "Camila Dias", "M\u00f4nica Silva",
    "Amanda Barbosa", "Bruno barreto", "Kalila Caetano", "Andressa Silva",
    "Marcelo Souza", "Rafaela Menezes", "Rainei Trindade", "Vanessa Vieira",
    "Virginia Vieira", "Virg\u00ednia Vieira", "Luana Castilho",
    "Luana Castilho Comercial Amara", "Gabriel Borges", "Elen Cruz",
    "Celen\u00ea Carmo", "Artur - Suporte T\u00e9cnico", "Augusto Batista",
    "Produtos Financeiros", "Marketing", "SAC - Amara", "- Amara NZero",
]


def _norm_name(s):
    s = str(s).lower().strip()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s)


_EXCLUDE_SET = {_norm_name(x) for x in EXCLUDE_NAMES}


def is_excluded(name):
    if not isinstance(name, str):
        return False
    n = _norm_name(name)
    if n in _EXCLUDE_SET:
        return True
    if "amara nzero |" in n or "| amara nzero" in n:   # contas de atendente
        return True
    if n.endswith("- amara") or n.endswith("-amara") or n.endswith("| amara"):
        return True
    if "@amaranzero.com" in n or "@vtex.com" in n or "suporte weni" in n:
        return True
    return False

AUTO_PAT = re.compile(
    r"(?:agradecemos (?:o |seu )?contato|fora do (?:nosso )?hor\u00e1rio|hor\u00e1rio de atendimento|"
    r"no momento[, ]+n\u00e3o estamos|resposta autom\u00e1tica|atendimento autom\u00e1tico|chave pix|"
    r"deixe sua mensagem|retornaremos (?:o |seu )?contato|assim que poss\u00edvel retorn|"
    r"seja bem[- ]vindo\(?a?\)? (?:a|\u00e0|ao)|voc\u00ea est\u00e1 falando com|menu principal|digite o n\u00famero|"
    r"voc\u00ea entrou em contato com|nosso (?:time|in[i\u00ed]cio)|estarei respons\u00e1vel pelo seu atendimento|"
    r"aguarde um momento que j\u00e1 iremos|para agilizar seu atendimento|"
    r"desenvolvemos projetos personalizados|consultor de energia solar|central de atendimento)",
    re.I)
AGENT_PREFIX = re.compile(r"^([A-Z\u00c0-\u00dd][a-z\u00e0-\u00fd\u00e7]+):\s*\n")
BOT_PHRASE = re.compile(r"como voc\u00ea avalia|deixe um coment\u00e1rio|atendimento foi finalizado|assistente virtual|enviei um c\u00f3digo|percebi que talvez|agradecemos a sua colabora|sou a assistente", re.I)
FLUIG_CLIENT = re.compile(r"\b(fluig)\b|repasse (?:da nota|do pedido|de nota|fiscal)|fazer (?:o )?repasse|preciso (?:do |de )?repasse|sobre (?:o )?repasse", re.I)
FLOW_ANSWERS = re.compile(r"^(sim|n\u00e3o|ok|1|2|3|4|5|topo|topa|claro|pode ser|oi|ol\u00e1|bom dia|boa tarde|boa noite|obrigad[oa]|valeu|blz|beleza|ola)[\s!.,]*$", re.I)
ELSEWHERE_NEG = re.compile(r"n\u00e3o comprei|nunca comprei", re.I)
AGENT_BAD = re.compile(r"programa de pontos|repasse na amara|demonstra\u00e7\u00e3o conta e ordem|voc\u00ea (?:quer|j\u00e1 tem)|vai emitir depois|podemos tentar", re.I)

CLIENT_CLOSE = [
 (r"pedido\s*(?:n[\u00ba\u00b0o\.]?|numero|n\u00famero)?\s*#?\d{4,}", "Citou n\u00famero de pedido"),
 (r"\b(?:fechei|finalizei|conclu\u00ed|fechamos|finalizamos)\b[^.!?\n]{0,40}\b(?:o |meu |a |minha )?(?:pedido|compra)\b", "Confirmou ter fechado o pedido"),
 (r"\b(?:fiz|realizei|efetuei|acabei de fazer)\b[^.!?\n]{0,30}\b(?:o |um |uma |minha |meu )?(?:pedido|compra)\b", "Confirmou ter feito o pedido"),
 (r"\bcomprovante\b[^.!?\n]{0,40}(?:pagamento|pgto|compra|pedido|boleto|pix)|segue (?:o )?comprovante|comprovante (?:de|do) (?:pgto|pagamento)", "Enviou comprovante de pagamento"),
 (r"\b(?:j\u00e1 )?paguei\b|pagamento (?:foi )?(?:feito|realizado|efetuado|aprovado)|fiz o pagamento|fiz o pix", "Confirmou pagamento"),
 (r"\b(?:meu |o |do )pedido\b[^.!?\n]{0,40}\b(?:chegou|foi faturado|foi entregue|saiu para entrega)\b|\brastreio (?:do|de) (?:meu )?pedido\b", "Pedido j\u00e1 faturado/entregue"),
 (r"\bvou subir o pedido\b|\bsubi o pedido\b|\bpode faturar\b|pedido que acabamos de fechar|pedido (?:j\u00e1 )?(?:foi )?confirmado pela amara", "Autorizou faturar / pedido confirmado"),
]
AGENT_CLOSE = [
 (r"pagamento confirmado com sucesso", "Atendente: pagamento confirmado"),
 (r"seu pedido (?:est\u00e1|foi) (?:confirmado|faturado|em separa\u00e7\u00e3o|em processamento)", "Atendente: pedido confirmado/em separa\u00e7\u00e3o"),
 (r"pedido \*?confirmado\*?", "Atendente: pedido confirmado"),
 (r"retirada est\u00e1 prevista", "Atendente: retirada agendada"),
 (r"nota (?:fiscal )?(?:foi )?emitida com sucesso", "Atendente: nota emitida"),
]
AGENT_PROGRESS = [
 (r"envie?\s*(?:o\s*)?(?:n\u00famero\s*d[eo]\s*)?pedido|n\u00famero do (?:seu )?pedido", "Atendente pediu n\u00ba do pedido"),
 (r"(?:envie?|anexe?|manda?r?)\s*(?:o\s*)?comprovante|ao finalizar (?:pode )?envi", "Atendente pediu comprovante"),
 (r"(?:pode|consegue) finalizar (?:seu |o )?pedido|finalizar (?:a )?compra (?:na plataforma|no site)", "Atendente orientou finalizar"),
]
# Categoria A: intencao explicita / Categoria B: interesse comercial (palavras do usuario)
CONTATAR = [
 (r"\bcomprar\b", "Categoria A: comprar", 5),
 (r"\bquero\b", "Categoria A: quero", 4),
 (r"\bpreciso\b", "Categoria A: preciso", 4),
 (r"\bfechar\b", "Categoria A: fechar", 5),
 (r"\bfinalizar\b", "Categoria A: finalizar", 5),
 (r"\bpedido\b", "Categoria A: pedido", 4),
 (r"\bor\u00e7amento\b", "Categoria A: or\u00e7amento", 4),
 (r"\bcontratar\b", "Categoria A: contratar", 4),
 (r"\badquirir\b", "Categoria A: adquirir", 4),
 (r"\bpre\u00e7o\b", "Categoria B: pre\u00e7o", 3),
 (r"\bvalor\b", "Categoria B: valor", 3),
 (r"\bfrete\b", "Categoria B: frete", 3),
 (r"\bentrega\b", "Categoria B: entrega", 3),
 (r"\bdesconto\b", "Categoria B: desconto", 3),
 (r"\bpromo\u00e7\u00e3o\b", "Categoria B: promo\u00e7\u00e3o", 3),
 (r"\bparcelamento\b", "Categoria B: parcelamento", 3),
 (r"\bpix\b", "Categoria B: pix", 3),
 (r"\bboleto\b", "Categoria B: boleto", 3),
 (r"\bcart\u00e3o\b", "Categoria B: cart\u00e3o", 3),
 (r"\bestoque\b", "Categoria B: estoque", 3),
]

# Extrai valor em R$ mencionado pelo cliente em texto (ex: "fechei por R$ 5.000,00",
# "paguei 3500 reais"). Usado como fallback quando nao ha comprovante de imagem com OCR.
PAT_VALOR_RS = re.compile(r"R\$\s*(\d[\d\.]*(?:,\d{2})?)", re.I)
PAT_VALOR_REAIS = re.compile(r"(\d[\d\.,]*)\s*reais", re.I)

def extrai_valor_texto(texto):
    """Procura um valor monetario em R$ mencionado no texto do cliente.
    Retorna float ou None. Usado apenas como complemento/fallback ao
    valor de comprovante (OCR), nunca o sobrescreve quando ja existe."""
    m = PAT_VALOR_RS.search(texto)
    if m:
        bruto = m.group(1)
    else:
        m = PAT_VALOR_REAIS.search(texto)
        if not m:
            return None
        bruto = m.group(1)
    # normaliza formato brasileiro: virgula e decimal (ultimos 2 digitos),
    # pontos antes da virgula sao separador de milhar
    if "," in bruto:
        partes = bruto.rsplit(",", 1)
        limpo = partes[0].replace(".", "") + "." + partes[1]
    else:
        limpo = bruto.replace(".", "")
    try:
        val = float(limpo)
        if 1 <= val <= 10_000_000:
            return val
    except ValueError:
        pass
    return None

# Frases que contem palavras da lista mas sao suporte/rastreio, nao intencao de compra
SUPORTE_PATS = [
    r"\bpreciso (?:falar|conversar)\s+(?:com|por)\b",
    r"\bquero (?:falar|conversar)\s+(?:com|por)\b",
    r"\bquero atendimento humano\b",
    r"\b(?:consultar|verificar|saber)\s+(?:o\s+)?(?:meu\s+|seu\s+)?pedido\b.*(?:nota fiscal|chegada|rastre|entrega)",
    r"\brastre(?:ar|io)\b.*pedido",
    r"\bnota fiscal\b.*(?:pedido|chegada)",
    r"\bpreciso (?:de\s+)?suporte\b",
    r"\bd\u00favida t\u00e9cnica\b",
    r"\bproblema (?:com|no)\b",
    r"\bn\u00e3o (?:funciona|chegou|recebi)\b",
    r"\bdigite a op\u00e7\u00e3o\b",
    r"\b\d\s*-\s*(?:or\u00e7amento|pedido)\b",
    r"\bbem vindo\(a\) ao atendimento\b",
]
SUP = [re.compile(p, re.I) for p in SUPORTE_PATS]

def is_suporte(text):
    return any(p.search(text) for p in SUP)

# Padroes que indicam que o ATENDENTE avancou a demanda de fato (nao resposta generica)
AGENT_ADVANCE = [
 (r"envie?\s*(?:o\s*)?(?:n\u00famero\s*d[eo]\s*)?pedido|n\u00famero do (?:seu )?pedido", "Atendente pediu n\u00ba do pedido"),
 (r"(?:envie?|anexe?|manda?r?)\s*(?:o\s*)?comprovante|ao finalizar (?:pode )?envi", "Atendente pediu comprovante"),
 (r"(?:pode|consegue) finalizar (?:seu |o )?pedido|finalizar (?:a )?compra (?:na plataforma|no site)", "Atendente orientou finalizar"),
 (r"segue (?:o )?(?:or\u00e7amento|proposta|valor)|valor (?:fica|ficou|\u00e9|seria)|(?:o )?pre\u00e7o (?:fica|\u00e9|seria)", "Atendente informou valor/or\u00e7amento"),
 (r"(?:o )?frete (?:fica|\u00e9|seria|custa)|prazo de entrega (?:\u00e9|fica|seria)", "Atendente informou frete/prazo"),
 (r"(?:temos|tem) (?:em )?estoque|dispon\u00edvel (?:sim|para entrega)|produto dispon\u00edvel", "Atendente confirmou disponibilidade"),
 (r"pagamento confirmado com sucesso", "Atendente confirmou pagamento"),
 (r"seu pedido (?:est\u00e1|foi) (?:confirmado|faturado|em separa\u00e7\u00e3o|em processamento)", "Atendente confirmou pedido"),
 (r"pedido \*?confirmado\*?", "Atendente confirmou pedido"),
 (r"pode (?:gerar|fazer|enviar) (?:a )?proposta|proposta comercial (?:enviada|segue)", "Atendente enviou proposta"),
 (r"desconto (?:de|aplicado|fica|seria)", "Atendente informou desconto"),
]
ADV = [(re.compile(p, re.I), d) for p, d in AGENT_ADVANCE]
ORCAMENTO = [(r"\b(?:or\u00e7amento|or\u00e7ar|or\u00e7a)\b", "Mencionou or\u00e7amento"), (r"\b(?:cota\u00e7\u00e3o|cotar|me cota|cotando)\b", "Pediu cota\u00e7\u00e3o")]
PERDIDO = [
 (r"comprei\s[^.!?\n]{0,30}(?:na weg|outra empresa|concorrente|outro fornecedor|outro lugar|com a aldo|na aldo|na edeltec)|fechei com (?:outra|outro)", "Comprou no concorrente"),
 (r"\b(?:muito caro|t\u00e1 caro|est\u00e1 caro|ficou caro|pre\u00e7o (?:muito )?(?:alto|elevado)|acima do mercado|mais caro que|valor (?:muito )?alto)\b", "Achou caro / condi\u00e7\u00e3o comercial"),
 (r"\b(?:sem estoque|n\u00e3o tem (?:em )?estoque|fora de estoque|produto indispon\u00edvel)\b", "Sem estoque"),
 (r"\b(?:n\u00e3o tenho (?:mais )?interesse|sem interesse|n\u00e3o quero mais|desisti|deixa pra l\u00e1|j\u00e1 resolvi|n\u00e3o vou (?:comprar|fechar)|cancelar? (?:o )?pedido)\b", "Recusou / desistiu"),
 (r"\b(?:prazo (?:muito )?(?:longo|grande|alto)|demora (?:muito|demais)|entrega (?:muito )?demorada)\b", "Prazo de entrega"),
 (r"\b(?:burocr\u00e1tic|muito complicado|dif\u00edcil (?:de )?comprar)\b", "Burocracia / fric\u00e7\u00e3o"),
]
INTENT = [
 (r"\bquero (?:fechar|seguir com|finalizar)\b|\bpodemos fechar\b|fechar (?:o )?(?:pedido|neg\u00f3cio)|seguir com (?:a )?compra|pr\u00f3ximo passo (?:para|pra) fechar", 5, "Quer fechar neg\u00f3cio"),
 (r"\bquero comprar agora\b|comprar agora|\bj\u00e1 quero (?:fazer|fechar|comprar)\b|fazer o pagamento|pode (?:gerar|emitir) (?:a )?(?:nota|nf|nota fiscal)|emiss\u00e3o (?:de|da) nota", 5, "Compra imediata / pagamento / NF"),
 (r"\b(?:tem|teria|tem algum)\b[^.!?\n]{0,15}desconto|consegue (?:melhorar|fazer melhor)|melhor(?:ar)? (?:o )?pre\u00e7o|\bdesconto\b|condi\u00e7\u00e3o (?:especial|melhor|de pagamento)", 4, "Desconto / negocia\u00e7\u00e3o"),
 (r"or\u00e7amento (?:para|pra) (?:fechar|comprar|efetivar|pedido)|quero (?:um )?or\u00e7amento", 4, "Or\u00e7amento para efetivar"),
 (r"pode (?:gerar|fazer|enviar) (?:a )?proposta|proposta comercial", 4, "Solicitou proposta"),
 (r"como (?:fa\u00e7o|faz|posso) (?:para |pra )?(?:realizar|fazer|fechar) (?:o )?(?:pedido|compra)|qual (?:o )?pr\u00f3ximo passo", 3, "Como proceder"),
]
TIME_PHRASES = [
 (r"\b(?:daqui|em)\s*(\d{1,2})\s*dias?\b", lambda m: ("days", int(m.group(1)))),
 (r"\bdepois de amanh\u00e3\b", lambda m: ("days", 2)),
 (r"\bamanh\u00e3\b", lambda m: ("days", 1)),
 (r"\b(?:semana que vem|pr\u00f3xima semana)\b", lambda m: ("days", 7)),
 (r"\b(?:segunda|segunda-feira)\b", lambda m: ("weekday", 0)),
 (r"\b(?:ter\u00e7a|ter\u00e7a-feira)\b", lambda m: ("weekday", 1)),
 (r"\b(?:quarta|quarta-feira)\b", lambda m: ("weekday", 2)),
 (r"\b(?:quinta|quinta-feira)\b", lambda m: ("weekday", 3)),
 (r"\b(?:sexta|sexta-feira)\b", lambda m: ("weekday", 4)),
]
CC=[(re.compile(p,re.I),d) for p,d in CLIENT_CLOSE]
ACR=[(re.compile(p,re.I),d) for p,d in AGENT_CLOSE]
AP=[(re.compile(p,re.I),d) for p,d in AGENT_PROGRESS]

# ---- como o fechamento aconteceu: sozinho, com apoio do comercial (so revisou/direcionou)
# ou com apoio ativo (refez orcamento, negociou, mandou proposta nova) ----
AGENT_REWORK = [
    (r"proposta[s]?\s+atualizada[s]?|or[çc]amento[s]?\s+(?:atualizado[s]?|novo[s]?)|novo[s]?\s+or[çc]amento", "Enviou proposta/orçamento atualizado"),
    (r"quer\s+que\s+(?:eu\s+)?altere|posso\s+alterar\s+(?:o|seu)\s+pedido|vou\s+alterar\s+(?:o|seu)\s+pedido", "Ofereceu alterar item do pedido"),
    (r"consegui\s+(?:um\s+)?desconto|condi[çc][ãa]o\s+especial|abaixei\s+o\s+valor|ajustei\s+o\s+valor", "Negociou desconto/condição especial"),
    (r"n[ãa]o\s+estamos\s+com.{0,30}em\s+estoque|temos.{0,30}em\s+estoque", "Verificou estoque para o pedido"),
]
AGENT_SELF_SERVICE = [
    (r"exclusivamente\s+pelo\s+portal|realizad[oa]s?\s+(?:exclusivamente\s+)?pelo\s+portal", "Direcionou para o portal"),
    (r"consegu(?:e|iu)\s+acessar\s+(?:o\s+)?(?:portal|site|app)|[ée]\s+s[óo]\s+(?:acessar|fazer)\s+(?:pelo|no)\s+(?:portal|site|app)", "Direcionou para autoatendimento"),
]
REW=[(re.compile(p,re.I),d) for p,d in AGENT_REWORK]
SSV=[(re.compile(p,re.I),d) for p,d in AGENT_SELF_SERVICE]
# mensagem que e so um anexo (o Weni exporta o nome do arquivo como texto, ex.: "042dde3f-...-....pdf")
ATTACH_PAT=re.compile(r"^[\w\-]+\.(pdf|jpg|jpeg|png)$", re.I)
# atendente humano = mensagem OUT que comeca com "Nome:" (nao usado como sinal de apoio -
# TODO cliente atendido por humano tem esse prefixo -, so para achar o texto do que ele disse)
HUMAN_PREFIX=re.compile(r"^\s*[A-ZÀ-Ÿ][a-zà-ÿ]+(?:\s[A-ZÀ-Ÿ][a-zà-ÿ]+)?\s*:\s*")
CT=[(re.compile(p,re.I),d,w) for p,d,w in CONTATAR]
ORG=[(re.compile(p,re.I),d) for p,d in ORCAMENTO]
PE=[(re.compile(p,re.I),d) for p,d in PERDIDO]
INT=[(re.compile(p,re.I),s,d) for p,s,d in INTENT]
TP=[(re.compile(p,re.I),f) for p,f in TIME_PHRASES]

def norm_cnpj(v):
    if v is None or (isinstance(v,float)): 
        try:
            import math
            if isinstance(v,float) and math.isnan(v): return ""
        except: pass
    s=re.sub(r"\D","",str(v))
    return s.zfill(14) if s else ""

def clean(v):
    if v is None: return ""
    s=str(v).strip()
    return "" if s in ("","nan","none","NaN",".","S/ VENDEDOR","SEM EXECUTIVO") else s

def add_bd(s,d):
    if d<=0: return s
    x=s; a=0
    while a<d:
        x+=timedelta(days=1)
        if x.weekday()<5: a+=1
    return x

def next_wd(s,wd):
    x=s+timedelta(days=1)
    while x.weekday()!=wd: x+=timedelta(days=1)
    return x

def find_file(d,*prefixes):
    for pre in prefixes:
        hits=sorted(glob.glob(os.path.join(d,pre)),key=os.path.getmtime,reverse=True)
        if hits: return hits[0]
    return None

def load_orders(path):
    if not path: return {}
    ped=pd.read_excel(path)
    if "Status" in ped.columns:
        ped=ped[~ped["Status"].isin(["Cancelada"])]
    if "Dt. Cria\u00e7\u00e3o" in ped.columns:
        ped=ped.copy(); ped["_dt"]=pd.to_datetime(ped["Dt. Cria\u00e7\u00e3o"],errors="coerce")
        ped=ped.sort_values("_dt")
    out={}
    for _,r in ped.iterrows():
        cn=norm_cnpj(r.get("CNPJ"))
        if not cn: continue
        dt=r.get("_dt")
        out[cn]=dict(pedido=str(r.get("Pedido","")),status=str(r.get("Status","")),
                     data=dt.strftime("%d/%m/%Y") if pd.notna(dt) else "",
                     data_iso=dt.strftime("%Y-%m-%d") if pd.notna(dt) else "")
    return out

def load_contacts(path):
    ct=pd.read_excel(path,dtype=str)
    m={}
    for _,row in ct.iterrows():
        razao=clean(row.get("Field:Razao Social")) or clean(row.get("Field:Nome Fantasia"))
        cn=norm_cnpj(row.get("Field:CNPJ"))
        cidade=clean(row.get("Field:Cidade"))
        if cidade and cidade.isupper(): cidade=cidade.title()
        is_raw=clean(row.get("Field:Score de intencao de compra") or "")
        ip_raw=clean(row.get("Field:Prioridade de intencao de compra") or "")
        try: is_val=int(float(is_raw)) if is_raw else 0
        except: is_val=0
        m[row["Contact UUID"]]=dict(
            vendedor=clean(row.get("Field:Analista")) or "Sem vendedor",
            cnpj=("%s.%s.%s/%s-%s"%(cn[:2],cn[2:5],cn[5:8],cn[8:12],cn[12:14])) if len(cn)==14 else "",
            cnpj_raw=cn, empresa=razao, regional=clean(row.get("Field:Regional")),
            uf=clean(row.get("Field:UF")) or clean(row.get("Field:Estado")), cidade=cidade,
            weni_intent_score=is_val, weni_intent_priority=ip_raw)
    return m


# Mapa Vendedor -> Time, usando o nome EXATO como aparece em Field:Analista
# (truncado em 25 caracteres pela Weni). Conferido contra a base real de contatos.
TEAM_MAP = {
    "RAFAELA PEREIRA DE MENEZE": "Digital",
    "GABRIEL BORGES DOS SANTOS": "Digital",
    "VANESSA VIEIRA DE CARVALH": "Digital",
    "BRUNO RICARDO CONCEICAO B": "Digital",
    "MARCELO SOUZA FERREIRA DI": "Digital",
    "ANDRESSA CLICIA DE JESUS":  "Digital",
    "KALILA DOS SANTOS CAETANO": "Digital",
    "LUANA CASTILHO RIBEIRO":    "Digital",
    "RAINEI TRINDADE DE SOUZA":  "Digital",
    "CELENE CARMO DOS SANTOS":   "Digital",
    "ELEN DA CRUZ SANTOS":       "Digital",
    "VICTORIA STEPHANIE SILVA SANTOS": "Digital",
    "KAREN GOMES SOUZA":         "Digital",
    "MAIANE MATOS MELO":         "Consultivo",
    "RAFAELA FERNANDES FERREIR": "Consultivo",
    "LIANDRA LOPES DE SOUSA":    "Consultivo",
    "LUIS HENRIQUE FERREIRA SA": "Consultivo",
    "LUANA ELIZABETT DE SANTAN": "Consultivo",
    "ALINE MIEKO MIYASHIRO HIG": "Consultivo",
    "BRUNA APARECIDA DOS SANTO": "Consultivo",
    "JORDANA BAIAO":             "Consultivo",
    "THAMIRES FERNANDA SOARES":  "Consultivo",
    "TIAGO DE OLIVEIRA JACINTO": "Consultivo",
    "ELMA EDLLA":                "Consultivo",
    "ARIELSON SANTANA":          "Consultivo",
    "NARA RIBEIRO COSTA ALBUQU": "Consultivo",
    "BRUNO FELIPE OLIVEIRA DE":  "Consultivo",
    "LUCAS MASSIAS FREITAS":     "Consultivo",
    "JOYCE SANTOS OLIVEIRA":     "Grandes Usinas/BESS",
    "NARA ELZIRA OLIVEIRA DE S": "Grandes Usinas/BESS",
    "AMARA BRASIL SSA":          "Outros",
    "VITOR BERTIN TEIXEIRA":     "Outros",
}

def team_of(vendedor):
    if not vendedor or vendedor == "Sem vendedor":
        return "S/ Vendedor"
    return TEAM_MAP.get(vendedor.strip(), "Outros")


def classify(conv_path, ct_map, orders):
    df=pd.read_excel(conv_path); df["Date"]=pd.to_datetime(df["Date"])
    df=df[df["Channel"].str.contains(PRODUCTION_CHANNEL,na=False)]
    period_start=df["Date"].min().strftime("%Y-%m-%d") if len(df) else "0000-00-00"
    first_dir=df.sort_values("Date").groupby("Contact UUID")["Direction"].first()
    last_info={}
    for uuid,g in df.groupby("Contact UUID"):
        g=g.sort_values("Date"); txt=g["Text"].iloc[-1]
        last_info[uuid]=dict(last_dir=g["Direction"].iloc[-1],last_date=g["Date"].iloc[-1],
                             last_text=str(txt)[:160] if pd.notna(txt) else "")
    real_in=df[(df["Direction"]=="IN") & df["Text"].notna()].copy()
    real_in=real_in[~real_in["Text"].str.contains(AUTO_PAT)]
    ins_uuids=set(real_in["Contact UUID"].unique())
    contacts=[]
    for uuid in ins_uuids:
        g=df[df["Contact UUID"]==uuid].sort_values("Date")
        _nm=g["Name"].iloc[0]
        if is_excluded(_nm):
            continue
        in_msgs=[(r["Date"],str(r["Text"])) for _,r in g.iterrows() if r["Direction"]=="IN" and pd.notna(r["Text"]) and not AUTO_PAT.search(str(r["Text"]))]
        out_msgs=[(r["Date"],str(r["Text"])) for _,r in g.iterrows() if r["Direction"]=="OUT" and pd.notna(r["Text"])]
        if not in_msgs:
            continue
        # ---- valor do negocio: prioriza comprovante (OCR) sobre valor mencionado em texto ----
        valor_origem = ""  # "comprovante" ou "texto" ou ""
        valor_negocio = None
        ocr_texto_evid = ""
        if "valor_comprovante" in g.columns:
            comprov = g[(g["Direction"]=="IN") & g["valor_comprovante"].notna()]
            if len(comprov):
                row_c = comprov.sort_values("Date").iloc[-1]
                valor_negocio = float(row_c["valor_comprovante"])
                valor_origem = "comprovante"
                ocr_texto_evid = str(row_c.get("ocr_texto",""))[:160]
        if valor_negocio is None:
            for dt,t in in_msgs:
                v = extrai_valor_texto(t)
                if v is not None:
                    valor_negocio = v
                    valor_origem = "texto"
        hit={"CLOSE":[],"ACLOSE":[],"CONTATAR":[],"ORCAMENTO":[],"PERDIDO":[]}
        ctw=0; deadline=None; iscore=0; idesc=""
        for dt,t in in_msgs:
            for pat,desc in CC:
                if pat.search(t): hit["CLOSE"].append((desc,dt,t[:220].strip())); break
            for pat,desc in PE:
                if pat.search(t):
                    if desc=="Comprou no concorrente" and ELSEWHERE_NEG.search(t): continue
                    hit["PERDIDO"].append((desc,dt,t[:220].strip())); break
            # Categoria A/B -- mas filtra suporte/rastreio antes
            if not is_suporte(t):
                for pat,desc,w in CT:
                    if pat.search(t): hit["CONTATAR"].append((desc,dt,t[:220].strip())); ctw=max(ctw,w); break
            for pat,desc in ORG:
                if pat.search(t): hit["ORCAMENTO"].append((desc,dt,t[:220].strip())); break
            for pat,fn in TP:
                mm=pat.search(t)
                if mm: deadline=(fn(mm),dt,t[:120].strip()); break
            for pat,s,desc in INT:
                if pat.search(t) and s>iscore: iscore=s; idesc=desc
        for dt,t in out_msgs:
            if "?" not in t.strip()[-60:] and not AGENT_BAD.search(t):
                for pat,desc in ACR:
                    if pat.search(t): hit["ACLOSE"].append((desc,dt,t[:220].strip())); break
        # contato so e marcado SUPORTE se TODAS as mencoes de categoria foram filtradas como suporte
        is_suporte_only = False
        if not hit["CONTATAR"]:
            for dt,t in in_msgs:
                if is_suporte(t):
                    for pat,desc,w in CT:
                        if pat.search(t):
                            is_suporte_only = True
                            hit["CONTATAR"].append(("SUPORTE:"+desc,dt,t[:220].strip()))
                            break
                    if is_suporte_only: break
        name=g["Name"].iloc[0]
        name=name if isinstance(name,str) and name.strip() not in ("",".","$","&") else "(sem nome)"
        meta=ct_map.get(uuid,{}); li=last_info[uuid]
        client_waiting=li["last_dir"]=="IN" and not FLOW_ANSWERS.match(li["last_text"].strip())
        convo_is_fluig=any(FLUIG_CLIENT.search(t) for _,t in in_msgs)
        cn=meta.get("cnpj_raw",""); oi=orders.get(cn,{}); has_order=bool(oi)
        convo_closed=bool(hit["CLOSE"]) or bool(hit["ACLOSE"])
        if convo_is_fluig and not hit["CLOSE"]: convo_closed=False

        # ---- fechou sozinho / com apoio do comercial (revisao ou reelaboracao do orcamento)? ----
        comercial_suporte=""; comercial_evid=""
        if convo_closed:
            rework_hit=None; selfserv_hit=None; human_out=False
            for dt,t in out_msgs:
                if HUMAN_PREFIX.search(t): human_out=True
                if rework_hit is None and ATTACH_PAT.match(t.strip()):
                    rework_hit=("Enviou proposta em anexo",dt,t[:160].strip())
                if rework_hit is None:
                    for pat,desc in REW:
                        if pat.search(t): rework_hit=(desc,dt,t[:220].strip()); break
                if selfserv_hit is None:
                    for pat,desc in SSV:
                        if pat.search(t): selfserv_hit=(desc,dt,t[:220].strip()); break
            if rework_hit:
                comercial_suporte="Comercial refez orçamento"; comercial_evid=rework_hit[2]
            elif not human_out:
                comercial_suporte="Fechou sozinho (sem atendente)"
            elif selfserv_hit:
                comercial_suporte="Fechou sozinho (via portal)"; comercial_evid=selfserv_hit[2]
            else:
                comercial_suporte="Comercial revisou orçamento"

        # atendente avancou a demanda DEPOIS do ultimo sinal de categoria A/B?
        last_signal_dt = max((e[1] for e in hit["CONTATAR"]), default=None) if hit["CONTATAR"] and not is_suporte_only else None
        agent_advanced=False; advance_ev=None
        if last_signal_dt is not None:
            for dt,t in out_msgs:
                if dt>last_signal_dt:
                    for pat,desc in ADV:
                        if pat.search(t): agent_advanced=True; advance_ev=(desc,dt,t[:220].strip()); break
                    if agent_advanced: break

        if convo_closed:
            stage="FECHADO"; ev=hit["CLOSE"][0] if hit["CLOSE"] else hit["ACLOSE"][0]
            ck="Cliente confirmou" if hit["CLOSE"] else "Atendente confirmou"
        elif hit["PERDIDO"]:
            stage,ev,ck="PERDIDO",hit["PERDIDO"][0],""
        elif is_suporte_only:
            stage,ck="SUPORTE",""; ev=hit["CONTATAR"][0]
        elif hit["CONTATAR"] and agent_advanced:
            stage,ck="ENTROU",""; ev=advance_ev
        elif hit["CONTATAR"]:
            stage,ck="CONTATAR",""; ev=max(hit["CONTATAR"],key=lambda e:e[1])
        elif hit["ORCAMENTO"]:
            stage,ev,ck="ORCAMENTO",hit["ORCAMENTO"][-1],""
        else:
            stage,ev,ck="SEM_SINAL",None,""

        contact_date=None; basis=""
        if stage=="CONTATAR":
            last_d=li["last_date"].normalize()
            if deadline:
                (kind,val),bdt,btxt=deadline; base=bdt.normalize()
                contact_date=add_bd(base,val) if kind=="days" else next_wd(base,val); basis='Cliente disse: "%s"'%btxt
            else:
                if ctw>=5 or client_waiting: contact_date=add_bd(last_d,0 if client_waiting else 1); tier="quente"
                elif ctw>=4: contact_date=add_bd(last_d,2); tier="morno"
                else: contact_date=add_bd(last_d,5); tier="frio"
                basis="Cad\u00eancia %s (sem prazo dito)"%tier
        contacts.append(dict(uuid=uuid,name=name,urn=str(g["URN"].iloc[0]),
            vendedor=meta.get("vendedor","Sem vendedor"),cnpj=meta.get("cnpj",""),empresa=meta.get("empresa",""),
            regional=meta.get("regional",""),uf=meta.get("uf",""),cidade=meta.get("cidade",""),
            weni_intent_score=meta.get("weni_intent_score",0),weni_intent_priority=meta.get("weni_intent_priority",""),
            team=team_of(meta.get("vendedor","Sem vendedor")),
            valor_negocio=valor_negocio,valor_origem=valor_origem,ocr_texto_evid=ocr_texto_evid,
            stage=stage,close_kind=ck,ev_desc=ev[0] if ev else "",ev_date=ev[1].strftime("%d/%m/%Y") if ev else "",ev_text=ev[2] if ev else "",
            mencionou_orcamento=bool(hit["ORCAMENTO"]),intent_score=iscore,intent_desc=idesc,
            comercial_suporte=comercial_suporte,comercial_evid=comercial_evid,
            erp_match=has_order,erp_recent=bool(oi.get("data_iso","") and oi.get("data_iso","")>=period_start),
            order_pedido=oi.get("pedido",""),order_status=oi.get("status",""),order_data=oi.get("data",""),
            last=li["last_date"].strftime("%d/%m/%Y"),last_iso=li["last_date"].strftime("%Y-%m-%d"),last_month=li["last_date"].strftime("%Y-%m"),
            client_waiting=bool(client_waiting),seller_replied=bool(agent_advanced),contacted_at=(advance_ev[1].strftime("%d/%m/%Y") if advance_ev else ""),
            origem="Inbound" if first_dir.get(uuid)=="IN" else "Disparo",
            contact_date=contact_date.strftime("%d/%m/%Y") if contact_date else "",
            contact_date_iso=contact_date.strftime("%Y-%m-%d") if contact_date else "",date_basis=basis))
    return contacts


def build_html(contacts, out_path):
    from datetime import datetime, timezone, timedelta
    import re as _re
    BRT = timezone(timedelta(hours=-3))
    now = datetime.now(tz=BRT).strftime("%d/%m/%Y %H:%M")
    # sanitize text fields to avoid unescaped newlines in JS
    _text_fields = ["ev_text","ev_desc","date_basis","name","empresa","cidade","vendedor","intent_desc","weni_intent_priority","comercial_evid"]
    clean_contacts = []
    for c in contacts:
        cc = dict(c)
        for f in _text_fields:
            if isinstance(cc.get(f), str):
                cc[f] = cc[f].replace("\n"," ").replace("\r"," ").replace("\t"," ").replace("\x00","")
        clean_contacts.append(cc)
    html=HTML_TEMPLATE.replace("__DATA__", json.dumps(clean_contacts,ensure_ascii=False))
    html=html.replace("__UPDATED__", now)
    with open(out_path,"w",encoding="utf-8") as f: f.write(html)


def build_xlsx(contacts, out_path):
    F="Arial"; C=dict(navy="1F3864",erp="0B6B57",done="0E7C66",grey="6B7280")
    thin=Border(*[Side(style="thin",color="D9D9D9")]*4)
    PT={"CONTATAR":"Entrar em contato","ATRASADO":"Atrasado","ORCAMENTO":"Mencionou or\u00e7amento","FECHADO":"Fecharam pedido","PERDIDO":"Perdido","ENTROU":"Entrou em contato","SUPORTE":"Pedido suporte/atendimento","SEM_SINAL":"Sem sinal","STALE":"Pendente (m\u00eas ant.)"}
    COLOR={"CONTATAR":"FBE2CB","ATRASADO":"F8CBCB","ORCAMENTO":"D6E8F5","FECHADO":"C6EFCE","PERDIDO":"E4DCF0","ENTROU":"D6EFE8","SUPORTE":"E6E0F0","SEM_SINAL":"F2F2F2","STALE":"EFEFEF"}
    order={"CONTATAR":0,"ATRASADO":1,"ORCAMENTO":2,"FECHADO":3,"PERDIDO":4,"ENTROU":5,"SUPORTE":6,"STALE":7,"SEM_SINAL":8}
    icolor={5:"1E7A4F",4:"C75B12",3:"A8860B"}
    now=datetime.now(); ws_=now.replace(hour=0,minute=0,second=0,microsecond=0)
    week_start=ws_-timedelta(days=ws_.weekday()); month_start=now.replace(day=1,hour=0,minute=0,second=0,microsecond=0)
    def iso(c): return datetime.strptime(c["contact_date_iso"],"%Y-%m-%d") if c["contact_date_iso"] else None
    def fstage(c):
        if c["stage"]=="SEM_SINAL": return "SEM_SINAL"
        if c["stage"]=="CONTATAR":
            d=iso(c)
            if d and d<month_start: return "STALE"
            if d and d<week_start: return "ATRASADO"
            return "CONTATAR"
        return c["stage"]
    wb=Workbook()
    def hrow(ws,r,n,fill):
        for c in range(1,n+1):
            cell=ws.cell(row=r,column=c);cell.font=Font(name=F,bold=True,color="FFFFFF",size=10)
            cell.fill=PatternFill("solid",start_color=fill);cell.alignment=Alignment(vertical="center",wrap_text=True);cell.border=thin
    ws=wb.create_sheet("Funil (conversa)")
    hdr=["Etapa","Inten\u00e7\u00e3o","Contatar at\u00e9","Atendido em","Tem pedido ERP","Cliente","Telefone","UF","Cidade","CNPJ","Empresa","Vendedor","Origem","Or\u00e7amento","\u00daltima Msg","Evid\u00eancia","Contact UUID"]
    ws.append(hdr); hrow(ws,1,len(hdr),C["navy"])
    def sk(c): return (order[fstage(c)],-c["intent_score"],c["contact_date_iso"] or "9999",c["last_iso"])
    for c in sorted(contacts,key=sk):
        es=fstage(c)
        ws.append([PT[es],c["intent_score"] or "",c["contact_date"],c.get("contacted_at",""),
                   ("SIM "+c["order_pedido"]) if c["erp_match"] else "",c["name"],c["urn"],c["uf"],c["cidade"],c["cnpj"],c["empresa"],c["vendedor"],c["origem"],
                   "Sim" if c["mencionou_orcamento"] else "",c["last"],
                   ('%s: "%s"'%(c["ev_desc"],c["ev_text"][:90].strip())) if c["ev_desc"] else "",c["uuid"]])
    for r in ws.iter_rows(min_row=2,max_row=ws.max_row):
        st=[k for k,v in PT.items() if v==r[0].value][0]
        for cell in r: cell.font=Font(name=F,size=9);cell.border=thin
        r[0].fill=PatternFill("solid",start_color=COLOR[st]);r[0].font=Font(name=F,size=9,bold=True)
        iv=r[1].value
        if iv in (3,4,5): r[1].fill=PatternFill("solid",start_color=icolor[iv]);r[1].font=Font(name=F,size=9,bold=True,color="FFFFFF");r[1].alignment=Alignment(horizontal="center")
        if r[4].value: r[4].font=Font(name=F,size=9,bold=True,color="0B6B57")
    for i,w in enumerate([18,9,12,12,16,22,14,6,15,18,22,22,9,9,11,46,38],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref="A1:Q%d"%ws.max_row
    we=wb.create_sheet("Confirmado no ERP (CNPJ)")
    h2=["Pedido","Status","Data Pedido","No per\u00edodo","Etapa na conversa","Cliente","CNPJ","Empresa","UF","Vendedor","Telefone","Contact UUID"]
    we.append(h2); hrow(we,1,len(h2),C["erp"])
    erp=[c for c in contacts if c["erp_match"]]
    erp.sort(key=lambda c:(c["order_data"].split("/")[::-1] if c["order_data"] else [""]),reverse=True)
    for c in erp:
        we.append([c["order_pedido"],c["order_status"],c["order_data"],"Sim" if c["erp_recent"] else "",PT.get(c["stage"],c["stage"]),
                   c["name"],c["cnpj"],c["empresa"],c["uf"],c["vendedor"],c["urn"],c["uuid"]])
    for r in we.iter_rows(min_row=2,max_row=we.max_row):
        for cell in r: cell.font=Font(name=F,size=9);cell.border=thin
        r[0].font=Font(name=F,size=9,bold=True,color="0B6B57")
    for i,w in enumerate([20,18,12,10,18,22,18,24,6,22,14,38],1): we.column_dimensions[get_column_letter(i)].width=w
    we.freeze_panes="A2"; we.auto_filter.ref="A1:L%d"%we.max_row
    rs=wb.active;rs.title="Resumo";rs.sheet_view.showGridLines=False
    rs["B2"]="FUNIL COMERCIAL \u2014 AMARA NZERO (Weni x Contatos x Pedidos)"
    rs["B2"].font=Font(name=F,bold=True,size=14,color=C["navy"])
    cnt=collections.Counter(fstage(c) for c in contacts); intent=collections.Counter(c["intent_score"] for c in contacts if c["intent_score"])
    erp_total=sum(1 for c in contacts if c["erp_match"]); erp_rec=sum(1 for c in contacts if c["erp_recent"])
    rows=[("FUNIL PELA CONVERSA",None,C["navy"]),("Entrar em contato",cnt["CONTATAR"]),("Atrasados",cnt["ATRASADO"]),
          ("Mencionou or\u00e7amento",cnt["ORCAMENTO"]),("Fecharam pedido (conversa)",cnt["FECHADO"]),("Perdidos",cnt["PERDIDO"]),
          ("Entrou em contato",cnt["ENTROU"]),("Pendente m\u00eas anterior (oculto)",cnt.get("STALE",0)),("Sem sinal (oculto)",cnt.get("SEM_SINAL",0)),
          ("CONFRONTO COM PEDIDOS (CNPJ)",None,C["erp"]),("Confirmado no ERP (todos)",erp_total),("Confirmado no ERP (no per\u00edodo)",erp_rec),
          ("INTEN\u00c7\u00c3O",None,C["done"]),("Inten\u00e7\u00e3o 5",intent.get(5,0)),("Inten\u00e7\u00e3o 4",intent.get(4,0)),("Inten\u00e7\u00e3o 3",intent.get(3,0))]
    r=5
    for item in rows:
        if len(item)==3:
            rs.cell(row=r,column=2,value=item[0]).font=Font(name=F,bold=True,size=11,color="FFFFFF")
            for c in range(2,5): rs.cell(row=r,column=c).fill=PatternFill("solid",start_color=item[2])
        else:
            rs.cell(row=r,column=2,value=item[0]).font=Font(name=F,size=10)
            vc=rs.cell(row=r,column=4,value=item[1]);vc.font=Font(name=F,bold=True,size=10);vc.alignment=Alignment(horizontal="right")
        r+=1
    rs.column_dimensions["B"].width=42;rs.column_dimensions["C"].width=3;rs.column_dimensions["D"].width=10
    piv=collections.defaultdict(lambda:collections.Counter()); erpby=collections.Counter()
    for c in contacts:
        piv[c["vendedor"]][fstage(c)]+=1
        if c["erp_match"]: erpby[c["vendedor"]]+=1
    wv=wb.create_sheet("Por Vendedor")
    h=["Vendedor","Entrar em contato","Atrasados","Or\u00e7amento","Fecharam (conversa)","Perdidos","Entrou","Confirmado ERP"]
    wv.append(h); hrow(wv,1,len(h),C["done"])
    for vend in sorted(piv,key=lambda v:-(piv[v]["CONTATAR"]+piv[v]["ATRASADO"])):
        p=piv[vend]
        wv.append([vend,p["CONTATAR"],p["ATRASADO"],p["ORCAMENTO"],p["FECHADO"],p["PERDIDO"],p["ENTROU"],erpby.get(vend,0)])
    for rr in wv.iter_rows(min_row=2,max_row=wv.max_row):
        for cell in rr: cell.font=Font(name=F,size=9);cell.border=thin
        rr[0].font=Font(name=F,size=9,bold=True)
        if rr[2].value and rr[2].value>0: rr[2].font=Font(name=F,size=9,bold=True,color="B0322E")
        if rr[7].value and rr[7].value>0: rr[7].font=Font(name=F,size=9,bold=True,color="0B6B57")
    for i,w in enumerate([28,16,11,11,16,10,9,14],1): wv.column_dimensions[get_column_letter(i)].width=w
    wv.freeze_panes="A2"; wv.auto_filter.ref="A1:H%d"%wv.max_row
    wb.save(out_path)

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--input",default="./entradas"); ap.add_argument("--output",default="./saidas")
    ap.add_argument("--conversas"); ap.add_argument("--contatos"); ap.add_argument("--pedidos")
    a=ap.parse_args()
    conv=a.conversas or find_file(a.input,"message_export*.xlsx")
    cont=a.contatos or find_file(a.input,"contact_export*.xlsx")
    ped=a.pedidos or find_file(a.input,"*elat*rio*.xlsx","*edido*.xlsx","*rder*.xlsx")
    if not conv: sys.exit("Nao encontrei message_export*.xlsx em %s"%a.input)
    if not cont: sys.exit("Nao encontrei contact_export*.xlsx em %s"%a.input)
    os.makedirs(a.output,exist_ok=True)
    print("[1/4] Pedidos: %s"%(os.path.basename(ped) if ped else "(nenhum - Fecharam pedido ficara vazio)"))
    orders=load_orders(ped)
    print("[2/4] Contatos: %s"%os.path.basename(cont)); ct_map=load_contacts(cont)
    print("[3/4] Conversas: %s"%os.path.basename(conv)); contacts=classify(conv,ct_map,orders)
    print("[4/4] Gerando HTML + XLSX")
    build_html(contacts,os.path.join(a.output,"funil_comercial_amara.html"))
    build_xlsx(contacts,os.path.join(a.output,"funil_comercial_amara.xlsx"))
    cnt=collections.Counter(c["stage"] for c in contacts)
    print("\nOK - %d contatos"%len(contacts))
    erp=sum(1 for c in contacts if c.get("erp_match"))
    print("   Fecharam pedido (conversa): %d"%cnt.get("FECHADO",0))
    print("   Confirmado no ERP (CNPJ):   %d"%erp)
    print("   Entrar em contato:        %d"%cnt.get("CONTATAR",0))
    print("   Entrou em contato:        %d"%cnt.get("ENTROU",0))
    print("   Pedido suporte/atendim.:  %d"%cnt.get("SUPORTE",0))
    print("   Mencionou orcamento:      %d"%cnt.get("ORCAMENTO",0))
    print("   Perdidos:                 %d"%cnt.get("PERDIDO",0))
    print("\nArquivos em: %s"%os.path.abspath(a.output))
    print("Enviando para o GitHub Pages...")
    upload_to_github(os.path.join(a.output,"funil_comercial_amara.html"),
                     os.environ.get("GITHUB_TOKEN","").strip())

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Funil Comercial · Amara NZero · Weni</title>
<link rel="icon" type="image/x-icon" href="data:image/x-icon;base64,AAABAAEAEBAAAAAAIAB2AwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAAQAAAAEAgGAAAAH/P/YQAAAz1JREFUeJx1k1tslFUUhdc+5/z/zPxzKR1gKNFUQ6ClRUZDC0VtpNJSSAU0MR0eTNQHlUdNfDAkJuPExBcjLyZeiAmK96mNaFJSWqLEGBMupZESQ1tLRYNKp06nZYbOfzln+wAlGHW9fytrr+xFuKl8vldmMn0aAIbGe5oswe0UBGukEB4pmqgG8mRX48AVAMhyVuQoZwCAbocHTnU2xJZbrwpjdtu2cAIW8JnAxsCGLkjGp2NTyD27a6iYZYgcwdASfHx0R7eTkPloRNT8MWv0ZCXk/u5KuAbkCKhkyFhNdYxksDixWPV7Ojd+O5VlCAKDTl7oWi/C8qwiOFMlef3LouOc8wjFgAFtkPS1vicm+a4wmXtXGbtRlsdDSm5pW9dWViCwN0YHa0MifHkW5bcL0dgPM+4YtDkKojIkti0o2fPLXwFvXUZkrtqeUx9vXKMXDhDlDtDw+Z1dy1PWsHYN3rqawOGfqh/2OtYz/Zk+j28WrD7e+6QW4rBxNe9ZqXhDQsq2+LWZluR8s5IWz7nX9QvFMvSp31xeO1J974s3j3r87nMWGsYZE2Xyn/j6iHx/73aK2k+NXvO9+rAURbZXlbzIVsJ/iwDcCJDdpvBKh5Efje7WYesrx/X9fSmLNqeMaoksPC/+x+Bf0sQGYDAAvmHNYCg1eKF7S21cts+U2Lz0c0RHjv/6zrlDI/6tEwopAcoF8oM9jxoi1CqwIyFDrMkGX1JKoMYKyTfqlhm03hnCkQfvSPPqkf1y/yF/qUT65LEMEz3NZU/ft0KJuE1US+5sKq6/JwAYGtsxnKyRD0/PisXX/4zGThfcM9DcD4EKiDqgxOMIDG9OkGmJ2fr+em2vw8LBBxpPvEgA8N3FXWkonJHM9kRJLfbPRSKjVUJJM2AYST/QzVGBtRGhN6w0drNdmZa+1dK+sXVe5fO98qH1feeHfuzqDUfVZ+nVHEnalWCyYntXXAnfgBwhVdI2VlNdIFeY6nSlah7pTg/MMbeKf4zp2Mj2dCxhv0ZsukMhabma4BsCgaGMnrclPr88Ty9nNg0WmLOCKGdu/cHtcz4xuXOTAjpIm7sFkU8CF10W33Q2DE4BwBIMAH8DEBCDYUkETXQAAAAASUVORK5CYII=">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Lato:ital,wght@0,400;0,500;0,700;0,800;1,400&display=swap" rel="stylesheet">
<style>
:root{--paper:#F2F5F0;--card:#FFFFFF;--ink:#182019;--ink-soft:#5C6B61;--ink-faint:#8B978E;--line:#E1E8DE;--line-soft:#EDF2EA;--brand-dk:#00953B;--brand-md:#76BC21;--brand-lt:#C1D116;--call:#C75B12;--over:#B0322E;--orc:#1F6FB2;--orc-bg:#E3EFF8;--won:#00953B;--won-bg:#E1F4E8;--erp:#057A4A;--erp-bg:#DCF1E7;--lost:#5A4A7A;--done:#3D8F1F;--done-bg:#EAF3DD;--i5:#00953B;--i4:#76BC21;--i3:#A8860B;--sup:#6B5B95;--sup-bg:#EDE9F4;--shadow-sm:0 1px 2px rgba(15,30,20,.07);--shadow-md:0 4px 16px rgba(15,30,20,.10);--shadow-lg:0 10px 30px rgba(15,30,20,.16)}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Lato',sans-serif;background:var(--paper);color:var(--ink);font-size:14px;-webkit-font-smoothing:antialiased}
header{padding:22px 28px 20px;border-bottom:1px solid var(--line);background:var(--card)}
.hd-top{display:flex;align-items:flex-start;gap:18px}
.hd-logo{height:38px;width:auto;flex-shrink:0;margin-top:2px}
.hd-titles{flex:1;min-width:0}
.eyebrow{font-family:'Lato',sans-serif;font-size:10.5px;letter-spacing:.14em;text-transform:uppercase;color:var(--brand-dk);font-weight:800}
h1{font-family:'Lato',sans-serif;font-size:23px;font-weight:800;margin:3px 0 4px;color:var(--ink);letter-spacing:-.02em}
.hd-meta{font-family:'Lato',sans-serif;font-size:12px;color:var(--ink-soft);display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.hd-meta .dot-sep{width:3px;height:3px;border-radius:50%;background:var(--ink-faint);display:inline-block}
.hd-meta b{color:var(--brand-dk);font-weight:700}
.kpibar{display:flex;flex-wrap:wrap;gap:10px;margin-top:18px}
.kpi-card{flex:0 0 148px;background:var(--paper);border:1px solid var(--line);border-radius:10px;padding:11px 13px 12px;position:relative;overflow:hidden;transition:box-shadow .15s,transform .15s}
@media(max-width:600px){.kpi-card{flex:0 0 calc(50% - 5px)}}
.kpi-card:hover{box-shadow:var(--shadow-sm);transform:translateY(-1px)}
.kpi-card .kc-label{font-family:'Lato',sans-serif;font-size:10px;color:var(--ink-soft);text-transform:uppercase;letter-spacing:.06em;font-weight:800;margin-bottom:5px}
.kpi-card .kc-value{font-family:'Lato',sans-serif;font-size:24px;font-weight:800;color:var(--ink);line-height:1.05}
.kpi-card .kc-sub{font-family:'Lato',sans-serif;font-size:10.5px;color:var(--ink-soft);margin-top:3px}
.kpi-card .kc-bar{position:absolute;bottom:0;left:0;height:3px;border-radius:0 3px 3px 0;opacity:.85}
.controls{display:flex;gap:9px;padding:13px 28px;align-items:center;flex-wrap:wrap;background:var(--card);border-bottom:1px solid var(--line)}
.controls input[type=search],.controls select{padding:8px 12px;border:1px solid var(--line);border-radius:20px;font:inherit;background:var(--paper);color:var(--ink)}
.controls input[type=search]{flex:1;min-width:190px}
.controls input:focus,.controls select:focus{outline:2px solid var(--brand-dk);outline-offset:1px}
.controls label{font-size:10.5px;color:var(--ink-soft);text-transform:uppercase;letter-spacing:.06em;margin-right:-4px;font-weight:700}
.clear-filters-btn{padding:8px 12px;border:1px dashed var(--ink-faint);border-radius:20px;background:transparent;color:var(--ink-soft);font:inherit;font-size:12.5px;cursor:pointer}
.clear-filters-btn:hover{border-color:var(--ink-soft);color:var(--ink)}
@media(max-width:600px){.controls,header{padding-left:16px;padding-right:16px}.board{padding-left:10px;padding-right:10px}}
.board{display:flex;gap:0;padding:16px 14px 40px;align-items:flex-start;overflow-x:auto;scroll-snap-type:x proximity}
.board>section{flex:0 0 272px;width:272px;padding:0 6px;scroll-snap-align:start}
.col-sep{margin-left:8px;padding-left:8px;border-left:2px dashed var(--erp)}
.col-head{display:flex;justify-content:space-between;align-items:center;padding:8px 10px;position:sticky;top:0;background:var(--paper);z-index:2;border-radius:9px;margin-bottom:2px}
.col-head h2{font-family:'Lato',sans-serif;font-size:11px;font-weight:800;text-transform:uppercase;letter-spacing:.03em;color:var(--tier)}
.col-head .count{font-family:'Lato',sans-serif;font-size:11px;font-weight:800;color:#fff;background:var(--tier);border-radius:999px;min-width:20px;height:20px;display:inline-flex;align-items:center;justify-content:center;padding:0 6px}
.card{background:var(--card);border:1px solid var(--line);border-left:4px solid var(--tier);border-radius:10px;padding:10px 12px;margin-top:8px;cursor:grab;box-shadow:var(--shadow-sm);transition:box-shadow .15s,transform .1s;position:relative}
.card:hover{box-shadow:var(--shadow-md);transform:translateY(-1px)}
.card:focus-visible{outline:2px solid var(--ink);outline-offset:2px}
.card.dragging{opacity:.4;cursor:grabbing}
.card .nm,.card .tel,.card .uuid,.card .reason,.card .evid,.card .extra,.card .why,.card q{cursor:text;user-select:text}
.cardtop{display:flex;justify-content:space-between;gap:8px;align-items:flex-start}
.nm{font-weight:700;font-size:14px;line-height:1.28;color:var(--ink)}
.tel{font-family:'Lato',sans-serif;font-size:10.5px;color:var(--ink-soft);margin-top:2px}
.uuid{display:none}
.card.open .uuid{display:block;font-family:'Lato',sans-serif;font-size:9px;color:var(--ink-faint);margin-top:5px;word-break:break-all}
.iscore{flex-shrink:0;width:28px;height:28px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-family:'Lato',sans-serif;font-weight:800;font-size:13px;color:#fff;box-shadow:var(--shadow-sm)}
.temp-chip{display:inline-flex;align-items:center;gap:4px;font-size:9.5px;font-weight:800;text-transform:uppercase;letter-spacing:.04em;padding:2px 7px;border-radius:999px;margin-top:7px}
.temp-chip.quente{background:#FBE4E3;color:var(--over)}
.temp-chip.morno{background:#FBEADB;color:var(--call)}
.temp-chip.frio{background:#E3EFF8;color:var(--orc)}
.temp-chip::before{content:'●';font-size:7px}
.datebox{margin-top:8px;padding:8px 10px;border-radius:8px;background:var(--paper);border:1px solid var(--line)}
.datebox .when{font-weight:700;font-size:12px;color:var(--tier)}
.datebox .week{font-family:'Lato',sans-serif;font-size:10px;color:var(--ink-soft);margin-top:1px}
.datebox .basis{font-size:9.5px;color:var(--ink-faint);margin-top:3px;font-style:italic}
.orderbox{margin-top:8px;padding:8px 10px;border-radius:8px;background:var(--erp-bg);border:1px solid #B6DECF}
.orderbox .ped{font-weight:700;font-size:11px;color:var(--erp)}
.orderbox .st{font-size:10px;color:var(--ink-soft);margin-top:1px}
.badges{display:flex;gap:5px;flex-wrap:wrap;margin-top:7px}
.b{font-size:10px;padding:2px 7px;border-radius:999px;background:#EDF0F2;color:var(--ink-soft);font-weight:600}
.b.sell{background:#EAE6F4;color:#5A4A7A}.b.orc{background:var(--orc-bg);color:var(--orc)}.b.inb{background:var(--won-bg);color:var(--won)}.b.uf{background:#E8EEF4;color:#33506E;font-weight:700}.b.erp{background:var(--erp-bg);color:var(--erp);font-weight:700}
.reason{font-size:11px;color:var(--ink-soft);margin-top:7px;line-height:1.4;padding-top:7px;border-top:1px solid var(--line-soft)}
.evid{display:none;margin-top:9px;padding-top:9px;border-top:1px dashed var(--line);font-size:12px;line-height:1.5}
.evid .why{font-weight:700;font-size:10px;text-transform:uppercase;letter-spacing:.03em;color:var(--tier);margin-bottom:4px}
.evid q{color:var(--ink-soft);font-style:italic}.evid .extra{margin-top:6px;font-size:10.5px;color:var(--ink-soft)}
.card.open .evid{display:block}
.more{width:100%;padding:9px;border:1px dashed var(--line);background:transparent;border-radius:10px;cursor:pointer;color:var(--ink-soft);font:inherit;font-size:12px;margin-top:9px}
.empty{color:var(--ink-faint);font-size:11.5px;padding:14px 6px;text-align:center;border:1px dashed var(--line);border-radius:8px}
.col-body{min-height:60px}
.col-drop-over{background:rgba(22,36,58,.06);border-radius:8px;outline:2px dashed var(--ink-soft)}
.legend{display:flex;gap:14px;flex-wrap:wrap;margin:10px 0 4px}
.legend span{display:flex;align-items:center;gap:5px;font-size:11px}
.dot{width:13px;height:13px;border-radius:4px;display:inline-block}
footer{padding:0 28px 34px;color:var(--ink-soft);font-size:11.5px;max-width:1020px;line-height:1.6}
footer b{color:var(--ink)}
.msel{position:relative;display:flex;align-items:center;gap:6px}
.msel-btn{padding:8px 12px;border:1px solid var(--line);border-radius:20px;font:inherit;background:var(--paper);color:var(--ink);cursor:pointer;min-width:90px;text-align:left;font-size:13px}
.msel-btn:hover{border-color:var(--ink-soft)}
.msel-panel{display:none;position:absolute;top:calc(100% + 4px);left:0;background:#fff;border:1px solid var(--line);border-radius:10px;box-shadow:var(--shadow-lg);z-index:50;min-width:200px;max-height:280px;overflow-y:auto;padding:6px}
.msel.open .msel-panel{display:block}
.msel-opt{display:flex;align-items:center;gap:8px;padding:6px 8px;border-radius:6px;cursor:pointer;font-size:13px}
.msel-opt:hover{background:var(--paper)}
.msel-opt input{cursor:pointer}
.msel-actions{display:flex;justify-content:space-between;padding:4px 8px 6px;border-bottom:1px solid var(--line);margin-bottom:4px}
.msel-actions button{font-size:11px;color:var(--ink-soft);background:none;border:none;cursor:pointer;text-decoration:underline}
.card-menu-btn{position:absolute;top:8px;right:8px;width:22px;height:22px;border-radius:6px;border:none;background:transparent;color:var(--ink-soft);cursor:pointer;font-size:14px;line-height:1;display:flex;align-items:center;justify-content:center}
.card-menu-btn:hover{background:var(--paper)}
.card-menu{display:none;position:absolute;top:32px;right:8px;background:#fff;border:1px solid var(--line);border-radius:10px;box-shadow:var(--shadow-lg);z-index:60;min-width:180px;padding:4px;font-size:12.5px}
.card-menu.open{display:block}
.card-menu button{display:block;width:100%;text-align:left;padding:7px 10px;border:none;background:none;cursor:pointer;border-radius:6px;color:var(--ink)}
.card-menu button:hover{background:var(--paper)}
.card-menu .sep{border-top:1px solid var(--line);margin:4px 0}
.card-menu .submenu{display:none;padding-left:8px}
.card-menu .submenu.open{display:block}
.toast{position:fixed;bottom:16px;left:50%;transform:translateX(-50%);background:var(--ink);color:#fff;padding:9px 18px;border-radius:8px;font-size:12.5px;z-index:200;opacity:0;transition:opacity .25s;pointer-events:none}
.toast.show{opacity:1}
.saving{position:fixed;bottom:16px;right:16px;background:var(--ink);color:#fff;padding:8px 16px;border-radius:8px;font-size:12px;font-family:'Lato',sans-serif;z-index:99;opacity:0;transition:opacity .3s}
.saving.show{opacity:1}
.kpi-valor{margin-top:12px;padding:14px 18px;background:var(--paper);border:1px solid var(--line);border-radius:10px}
.kpi-valor .kv-total{font-family:'Lato',sans-serif;font-size:22px;font-weight:800;color:var(--won)}
.kpi-valor .kv-label{font-size:10.5px;color:var(--ink-soft);text-transform:uppercase;letter-spacing:.05em;margin-bottom:2px}
.kpi-valor .kv-sub{font-size:11px;color:var(--ink-soft);margin-top:2px}
.kpi-breakdown{display:flex;gap:18px;flex-wrap:wrap;margin-top:10px;padding-top:10px;border-top:1px dashed var(--line)}
.kpi-bk-item{min-width:120px}
.kpi-bk-item .bk-name{font-size:11.5px;font-weight:700;color:var(--ink)}
.kpi-bk-item .bk-val{font-family:'Lato',sans-serif;font-size:12.5px;color:var(--won)}
.kpi-bk-item .bk-count{font-size:10px;color:var(--ink-soft)}
.kpi-toggle{font-size:11px;color:var(--ink-soft);text-decoration:underline;cursor:pointer;margin-top:8px;display:inline-block}
.deal-tag{display:inline-flex;align-items:center;gap:4px;font-size:9.5px;font-weight:700;padding:2px 7px;border-radius:999px;margin-top:5px;margin-right:4px}
.deal-tag.neutro{background:#EDF0F2;color:var(--ink-soft)}
.suporte-tag{display:inline-block;font-size:9.5px;font-weight:700;padding:2px 7px;border-radius:999px;margin-top:5px;cursor:default}
.suporte-tag.refez{background:var(--sup-bg);color:var(--sup)}
.suporte-tag.sozinho{background:#E8EEF4;color:#33506E}
.suporte-tag.apoio{background:#EDF0F2;color:var(--ink-soft)}
.valuebox{margin-top:9px;padding:9px 11px;border-radius:8px;border:1px solid transparent}
.valuebox .vb-amount{font-family:'Lato',sans-serif;font-weight:800;font-size:17px;line-height:1.1}
.valuebox .vb-tag{font-size:10px;font-weight:700;margin-top:3px}
.valuebox.confirmado{background:var(--won-bg);border-color:#B7E0C6}
.valuebox.confirmado .vb-amount{color:var(--won)}
.valuebox.confirmado .vb-tag{color:var(--won)}
.valuebox.a-confirmar{background:#FBEADB;border-color:#F0CFA0}
.valuebox.a-confirmar .vb-amount{color:var(--call)}
.valuebox.a-confirmar .vb-tag{color:var(--call)}
.valuebox.sem-valor{background:#F7F0D4;border-color:#E9D896}
.valuebox.sem-valor .vb-tag{color:#8A6D0B;font-size:11px}
.intent-bar-wrap{margin-top:7px}
.intent-label{display:flex;justify-content:space-between;align-items:center;font-size:10px;margin-bottom:3px}
.intent-label .pri{font-weight:800;padding:1px 7px;border-radius:999px;font-size:9.5px}
.pri.alta{background:#FBE4E3;color:#B0322E}
.pri.media{background:#FBEADB;color:#C75B12}
.pri.baixa{background:#F7F0D4;color:#A8860B}
.intent-bar{height:5px;background:var(--line-soft);border-radius:3px;overflow:hidden}
.intent-bar i{display:block;height:100%;border-radius:3px;transition:width .4s}
</style>
</head>
<body>
<header>
  <div class="hd-top">
    <img class="hd-logo" src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAacAAAC0CAYAAAAw2uMZAACT40lEQVR4nOx9d5wcxbH/t7pnNl1WTiABIklkkZG0J2yCMRgM3sM4YzCY5Jye/ey9tZ9/fg74GRBB2ARjTLglGBswWbfkJINAElkSyvHyxunu+v0xM5f37iTdCQH79Wd1eHdCT81MV1fVt6oIJZRQQgkfJzAoXg+aOTNGY8duItQCaAQ2b05xXR0MAN7hUzBEY2NUdB67dhwvrU9yfT2YaMeP/3EAfdADKKGEEkoYScTjEPX1MWps3ESeAtIDbf/g29ODFblwxJApl0ZEwkEOaoeDBSNCtuXOmQywYZ0XNuWFETkok1HE6fzmdGbevPdzAx2/oQFy7Ngo1daO4/r6JCcSYAyDQvyooaScSiihhI8U4nGI2tqoAIB581Kq9+/MEP9589AJGoVp2qE9WWBPAk0FYTIMxjMwiggVAMIAQpZFIAEIQT0mTGMYxgBKMZiRJ+IsQO0AmoiwiQ2vM6D3paT3LIHltuCVB+2933qiZB/lyByTjY2bqLE2ZRIEM1Ky+TChpJxKKKGEDzuIGdTYGBW1tSnd22328pIZu2thHcTALDbmUCLaF8AU26byQFBACIAZMAbQmmG0q3SMYTDD//RxxzGDyNNXQgAkCEIAUhKkAIQkEAFsgHyBUSiYDAFrALwN4sWWwMsWxOJD9nttRe/rWbgwKjdvTnEsBvNxdQOWlFMJJZTwoYQf1+ltHS1auv9ULe2jYThqgGPA2DccEWHLIhgNOA7DcRhaMwPcNfkziEFE5CoeACB3hhxsnmT21AeR+9/kHoG7NiApJciyCbYtICWgFSOTMXmA3xFEL0CIRiHUc4fvs/S97gdfuDBqDWc87MOCknIqoYQSPjTwXXbdFRIzxH/eO+Awo+XJbMxJhnFYWZmMCAIKrsUCrdkQwbhKhwjosnp2FpjBrt5j3wojIpKBoEAgQGAG0h06TwKvSiEeYcbDK5TzUt0Bywr+MRYujFqNjSmTSHz0XX8l5VRCCSXs8mhogARiqKvrite8+OYhRxD0Zxl0KgEHRsoEHIeRzxlozdq1YkjgA1BEQwWzS4YgYsMMEoJkKCRgBwjZrIE2eIsYDxLR3fffvvg5XykxQySTMaqrS35krald8oaVUEIJJbhWTkwAyU7X28tLZuzOljyLGZ8noiMjZQL5PCOXNSBiBRAxQ+yqymgw9FRWJIMholBIIJc1MIZfJaIGkG7o7vpjjsnuMvqo4EN5A0sooYSPLnyl1J3V9tI7M2tZy/Nh+DNlFbKiUGBkM65CYiZBBPFBjnnEwDDwFFUoLCgYJKTbdVZIekAzbjhy39ce9pXSwoVRqz9CyIcVJeVUQgkl7BLorZTefnt6sN2UxQyZi6QQx9oBQiatYcxHXCEVATMMERsissrKJbRmOA4vIsHXtef1HfMOWNYBfHSUVEk5lVBCCR84mGPSV0oLl8wor7DkVxh0aSgs9teakekwDGID0IfWZTdccF1/bACicFgIO0DIZcxyw7i2qVndePKxy5rc7bpk+mHEx/oml1BCCR8sGhog/VyehQunhsomVn5dEH03HBbTc3mDfM5ol9T28bKShgpmGIA5EBQyHBbIpM1qIlyRM/nrZ+//Vjvgyniwqhi7IkrKqYQSStjp6O3Ce+GNA78gBX4aCsuZ+ZxBPl9SStsIA2ZjBYQViQhkM+Y9Df7t+4tfv7GuDrqhISaXLk3yh4mCXlJOJZRQwk5Fd3fT80sPnGPb9MtAQNQ6jkEuZzSBCCWltF1wc6lYB4LCCoYEslnzkjH4+VH7v/Yw4Maj+ivptCuipJxKKKGEnQK3ACtABJN6eb+JkfJAPREusAOEdIcuWUrDCYZhMIcjUgKAdvh2h+mnx+y/eKVX/YJoF6/hV3oQSiihhBHHwoVRK5GAIYJ5bskB55ZVBhZFysQFhQJzukNrIpIlxTSMIAgiktmMNtmMMaGIOCcgedFLbx10GRGYCGbhwqj1QQ9zIJQspxJKKGHEwAxKIibqKKmfevHAPUOV+FMoLE/LZQ0cxygi2qUnyI8KmFlblpBl5QK5jH4io8S35sxcvJQZAtg1e0yVViollFDCiKChAZIIXEdJ/cKyA78aqaYXQ2F5Wnub0o7D/GFSTDzAZyT3HS4QkVSKua1FKTsojg/b5oUX3jjoW0QuU7KBY3InDmdIKFlOJZRQwrDDD7wvfOXg6vKIuTIUkl/OZgyUYzQJ2uUmwu7orTQEAEEMInfC7FbE3G2nARjDbjFZX+n4q34CGwJICJC7BzqPYBgwXcfoxEhPysyspSRZXiGRzZj70u2Fi6KHv7l+VyNLlJRTCSWUMGzoThF/avGBR4fDdHM4LPZtbVF6V02g7a4YfEUkvX63mokLmpBnwVklkdECGS05pwXltGDFQhQMKcOQDBKGPfVDriISgGNLIyywCFpGh4QRYam5zDIIC00hYSggDSwBIrdWEQwTdLdBjYTAfFZfRZVl5bNmdUGb846ZseTRhoaYjMV2jTp9u9yDUkIJJXw44bHxmAj8/JIDLrSD4goiCuayepeLLfkzLwGQxJDEYBDymjitJLc4Fjc7NloLFrUribwW5LCAcXs+efu6R/EtqoHO5fZ76rK5BMCSGEFpuExqVNoa1baDmoBCpa0QkoYkMRkQtOmi1Q33hM3MKhgUFgAoxT85asbrv3W/h/ig2Xwl5VRCCSXsMPwqBA0NkFMPPOjKsjJxcUe7htFsSNAuEdvuTyEpFmh3bGzOWbwxb3NTwUZaSVLs2j6CXMYAeX+HezwGgGHftUeQYA5ZmqtthbFBB+OCBVTbikKWIQBQTDB+Y8PhGgfDCAIqq6ToSJvbNrxtf+Mzn1mUaeCYrPsAyx+VlFMJJZSwQ/BjFY++sN/o6qrAHWXl4pOtLUoxk9wV3Hi+UpLEsDyF1FwIYH0uiPXZAJoLFgqGALjuPNFpG/U9xkig97kMAM2ue08Sc7mleFzQweRwDmOCDpVZmgwIyngdC4dhDL6br6rGsjId+sVsh47NOXzZqg8yDvWBPzgllFDChxf+5PX4iwfuW1lN94ZDYv+2VrVLuPH8idsSBgRCm2NjdSaENdkgmgoWHONaRlYvZfSBB1vQc2LWXTEoLrc0TwrlsXuZq6gCwpBi0Rmj2tEJnZlVeYVlOQWzOttuzpw9a8nLH5SCKimnEkooYbvgT1pPLT7w6EiY/iEtGp9Jf/DxJZ8tZwsDhwU25kJY3hHG+pyNnBZ9FNKuoIwGgz9W7VlMEsyjgg5PjWSxeyRPFbYizQTlEzJ24FxsWIfCQjLQnk6rurmHLHvog1BQJeVUQgklbDMWctSaRyn19OszTwyH5N1glOfzRhN9cDTx7kopbyRWZcJ4tz2MLQUbzAxLdMWNPgwKqRj8SVsxQTMhIjVPK8vy9PIsqgNKGMYOKylm1rYtpJRQ+Zz+0jEHLr1zwcuz7AsPX+QMy0UMASXlVEIJJWwT/Enq6cUHfiYSwV1ak10oGCM+IOKD774LeEppRTqCt9vDaC5YEMSwiTu3+yiBvH8MA44RCAjD0yJZ3rcyg1EBJRS7LsHtV1AwUoICQUGZtP76cQctuYk5ahHtHAuqpJxKKKGEIcOfnBpfOaCuskL8XTkslWIm2vmKyVc2AWGgWWBFRwRvtJehuSAhycCintt9lOHSOYCCEQiQ4T3Ls7xfRZqqAoocI2CwfZO9MWDLAgdDQnS0q2/MPnjpX3yreXivoC9KyqmEEkoYEvy4g2sx0T2Ow0Jr8AdRsJXh08GBtdkQXm8px+a8DfExU0q90V1JhYTm/SozvG9FmoKSyWUkbvukbwxYWuBQSIh0h2tB7QwFVVJOJZRQwqDoVEyvzDwxUi7vdxy2PgjF5CucoDBoc2y81lqBlekQGDvXfTdcFO6RAsGlpDtGoNp2zCHVHdi9LEeaQdvj6mNXQZlgUMi2DucLcw9edvtIK6hdWb4llFDCLgC/OeDClw88urISjxlDEbdw685XTJIYAoR3OsrwWksZMlogKEzn78NxDsCv6OCiewWI3n+5298++/az384GEaAMwTB4j7IsH1LdQRW2przZDgXlxqBgWYRMlk+dc/Dr/x5JFl9JOZVQQglF0dAQk3V1Sf3c0v33tizrGUE0Npfb+eQHhmsttToBLGquxOpMADYZCNoxpdRZWohcpp8k9yPgfsdMMG6NPRgGNLNb1cHtNAEBhgQgiSA69+2y4AwD2vvbWdlhkHJHww3/XHkjUCY1H1bTxnuUZ4VyyzFt01iYYSyLSErO5vKm9rgDl77UvbPxSIy7hBJKKKEH/Fp5j72436jqqsAzdoD2zaT1TqWLu3qDERDA8nQZFjVVIKuBgODtUkqdFg67isQSgPS+yxtChwZaHaDZYbQqRpsC0oqRM0DBAMpoGBA0BAgMAQNJBFtIBMkgJAXKLaDKIlTbQLUNVFqEMgnY5I5ZM6A8ZbUzFRXBzZPSBjy9PMOH1bRTwItFbctKwxg2obAQbLA+3+Ecc9ysN94fiVp8JeVUQgkl9IHfJDAG4Pllbz5aUSHn7ezKDwzAIoZhwn+aq/BWexiSDOR2WEs+W00SYHuzXkYTNheAdTnG+pzB5oKrkBzNXVqM/OiSb2J0d+j1qivB/UynghCWQLUtMD5ImBwiTAwRamwgKNijgbsKa2coqu5W1KiAY44e3YqxwYLIG7FtFpRhXVYhZT5nXk1rPad25rIMhrlpYUk5lVBCCX3gxxKee33mtVU19jdbmneuYjJw3XjtTgDPbq3Cxpy9zbEl32UnCQgIt7Bqi2KszgLvpQ3WZhmtDncFidxmTf6fztmRt2G67b2fb6V1KTtASMKYAGFqRGDPCGFSiFAmGdpTVIYBMcIzswDgMEGA+chRbTy9IiMKRmwT0YOZVVW1ZbW26nuOPeD1s4Y7/lRSTiWUUEIPdDLzXj3gmzWjrWvbWpUCdq7FFBQG63NhPLOlChlFrpWxDfszXAvJFkCHIqzMMN7oMHg/a5Bzeioj0VuZDDM6SRHU7RwGnSUtqgMC0yMC+1TATAkSBSTIdSGOrDXVxegjPqAyzYfUtJMB0bbEoRjs1FRb9tYmFZ990JJfDqeCKimnEkoooRN+64tnl8w4MhS0nlaKhVLYaU0CfcX0Xkc5XmiqgGGGNUQ3XmdSLgGCCJsKwJI2g2XtGq0FT2URuRM+jZwyGgoI3cbQZVXxhBDxgZWE/csFVdpEjnGtqZFWUjlD2KMsZ44Z1UpC8LbQzVkI1sGQtDpa9KdnH7bkweFqtVFSTiWUUAIAv4st6LFFsyoqIoX/BAJiz1xG77R+TL5iWtpWiUVN5bCEz4kbfD9mICgAIsKaHGNRC+OtDg2lXTPAV0jmg9JGA8BXVMZXUgYot8EHVgk+tFLQqCB1WlIj5e4TBOS0wIRQ3swZ00JBaUgNUUExwwQCRMy8xSFx6DH7LF5XXw9KJHaMILFLNAEroYQSPng0NkYlEUxZIH9dRYXcM5vRemc2CgwIxistVXi5qRz2EBWTYcAiICKBdXngnvUat6xWWNqqoJghpGehYNdUTEDPsREBwgI6NOi5LUbcuErhsc3aZBRzxHOsjsR1GAZCwmBjLiCe2FTDWS2NPURGJBFEIW9MKCzHUkH/lQhcWxsVtIOKtGQ5lVBCCZ2Jts8unvHVyprAze1tOzfOZAvGouYqLG2NICjMkKwlAhASwFaH8FyTweI2DTYMCNcS2FWV0VDQw5rSQMQGH1UjeFa1pKAAZfXIuPoEgDwTKi1lPjGuGWW2Fs4QE3aZWVXXWFbzZv2T4w59/bc7Gn8qKacSSviYg91OEvzssv13DwjrNRCVK8XATvCsMFyLaVsUk/FceBqERS2MZ5sUsg4DsiuO81GCIMC4mb8YFyZTO0Zg73IhnBFy9QkABSZUWMp8YnwzIlKLobj4mMFSwlgWTDatj5x96LJX/Rjm9o6jhBJK+FgjRkRgaOv6UERWOo7xWyONKNwYE+PV5kosGYJi8kMyEQmszwO3rdF4fJODrGYIz8b7qCkmoCtZV1jApjyLhjWa/rVBm7xmDsvhtxANgAAx2pUlGjfWIKcEWzS4i48IpDVDWsKWtrhh4cKoBcT8WOY2o2Q5lVDCxxi+O++5xQd+rbJG3tS2k9x5PvlhWVslXmoqH1QxGbjUcALh+RaDp7ZoaOPGlD7M7rttRWcKsAGqbPCJ4yXvWy5ETg9/MVoCUDCEsYECzxvfDEuA9BBo5r57b+sW9dM5hyz5zfaWNyoppxJK+JgiHneto89++aAxBYeXCUE1jjPy7jwDN/i+PF2OZ7ZUwqKBSV2GgZAEWhXh3xs1lrdr14WHj2dbDKCbq4/BR48WHB0tCQA5w+zmE3Bp5ruFcyY6toU0Bqc5MIMtC0YIKjjKOfjoGW+8C5dIuU3svZJbr4QSPqaYWR+jRAImnTG/KyuXowsFYzDCc4LvytuQC+P5rRWQZAZcIRu4bryVWeCW1QrLO3SXC28kB7qLo7MunwA9v8WIO9YqTivm0DBbku5CgrE6GxYvN1fyUBh8RCClGMGQCBstryQCJ5OxbVaZJcuphBI+hvCrjb/4+sxjrZB8plBgzW6B7RGD3/Iiq208smEUchpF6+T534UF8FIr8OgmBcMMIUbUjdf9yB+auVEQYDRQFYA5Y6KF3cIkMmp4LSiCW4/vyFGtZv/K9JBq8TGzrqiwZLpNffbog5b8Y1uTc0uWUwklfEwRj0Mo0B+lJNdFNMIguC0ont1ShbSiARWTSxMnPLaV8fAGBwwejiRaJrgdMABWYKPArMHGgA2DiEAg1wFl2P2eNZiVuz00dRUe2mVgGBASaHUgblutaFm7MWXW8Cpxl1VpsKi5gtZlgjy0qvBEjsMMgd8tXBENxZDkbSFHlJRTCSV8zLBwYdSqq0vqk2IHfr6yyjoqndaaaOStpoBgvNpSiQ05G8Eik5uvmGwB3L9R4/nNCkL0bOa3LXCVCSuw0YAhtiAQFhJltiUrwpYMB2SkrFyEgxECmzyY82CTj0TKKBSJCBkJSFERtBCxLYSFZAsCbAhsFMDKU1YfOHw3n8Oge9ZqWtRiuMwa3sG5pQiJnt9ahbQSLAdh8BFBZLPaVFZbewfat15CBNPYGB3yc/ahMV1LKKGEHQczqL4e9LWvTQ2sT1cuDQbFHvmcGdGutj4zb2W6DE9troJdhJnn89cFEe7baPBmiwNh0fZYAAZgA8MSAUEIWLAhEVCUD8vgiogdekdq80ZZpGI5FK/Zc/yU9iodzvztjUc3wDIEJfir+58w5b3cxmBLpq0GQu6Wy6X3Tqvc3hmd30cZs2chxHC0AgoKcJghSBOT3I4O6MOK7my+T46X5uhqITJ6+Fx8vntvaiRn5o5tFs7gF2wsi0hr3kodzn5HHvlmEwAMpbVGSTmVUMLHCH7W/rNLDri0qsq6qrVFjWjzwK44k4WHN4xGvkicybeYLCLcu9HgrRbl0sS37WQaMARbCBG0EVEWyu3w0ppgRWOlDD9x4Li9X/nz3B+8bxGZ7ckKtQA4zIEvP1K/z7pc81Gr05ujzfn2uR0mPzUnFJBTgIYmIuIP2CtFAFiDPzlB8jE1QqSHMQblKijCEaPazIwhxJ+MYV1TY8nWZvX/jjloyc+GSi0vKacSSviYwPf3Ny6dURaCeMMOiMlOgUc04dZtXcF4cssorEoHinawda0r4J8bDZY0u4y8IVtMDA02AhGLgsJGBQffmFg+5r7Dxk6/7+ban70kiHSPQ8UgMSNKqAWweRzHvK+TS5Odm8VmuuyyJACM3URoBICUQbdiphKAYo584r5vHbUu23z2pnzLae2cm+Q4BaDABiDGCLtLB4KASzf/9ATJh1ULSuthv9F84vitqA6oAYvEutRyAjO3ZYzer3bmso1DKQxbUk4llPAxQTer6bKqKuvKke7T5Lvz3u0ox7NbKhEo4s4zACICeHgL46Ut22QxGRhDiFhUzkHUhCvv33fUbn9+9OTLHyKiQudWDTEJAPGlMziRSGxv+Mq/KIrXx6kRjSI1M8XwSvMIAI83v1L9o8evO2NjdsulGwuts/KsgKzSJMQHYkl1Tu4G/NkpkmeUi2Fj8RHcZoUTgnlz/Phm0jxY/hOrqirLamnR/3vsga//11Dq7pWUUwklfAzgW03PPTclRJXVSwMBMa0wwlaTIEbOc+flisQ9DANlEni2hfH4Rpf8MJhiIoDZGIOQkBEKYbeKsf84euyMP9x2/M+fcfy941ErXl9rErSDymgwMCiWjIkkkvAVFTPTEXddcPL7mU0/bOPsvHwuCyhWICGxk+dc/2QWwF/YzeLJIRLF7sX2HDtvBI4a1Wr2q8yI/AAFYpnBlk0wmpuJQvscuf9Lg8aeSsqphBI+Bui0ml4/6CuVVeKvba0jH2sKCsZzW2vwdnuo3/JEhoGwBN5OA8m1CoKG0O2WWYOMDJaXYZI16tkDx0z9r3+f9PsnHRggDoGZMUJd8oOhezMTknUCdW48JQCJ2n9+6wuvt74f32Ta9tHtORAJs7OtKJfCD1TaZL66m6SIJFLb0O12IDAAG8wnTdyKiGVoQFcss6qstqy2Vv2TYw4YvGp5STmVUMLHAMygeoBOXHLAS2UReWg2awwwMvEQP860KR/CYxtrYFFfxeS3UW9RwF9XKeS8om3F5jYCwMZoRCw5zq5qPqhmr589eurl1xERIwYZb4hzghK7BK0bgOtKXJpkJGB4M1cc/OT5P3qvefWP0iYXQIEVaOe1IwG6EnWnlpE5Z7JFDg+hDtEQ4NbfE9i7PG2OHtMmCgO31zCBAFE+b1ahvXz/Y455PgcUt55KeU4llPARRwPHJBH4U0sOjkYi8rBMRjNGSDH5YBBeay2HKVImnOC67/65QSOrXFtiAMVkmLWxK0Jyn7IpD/xi5hePfOy0P15LRK4SSELvUooJAOqSGgmYWENM0lhqf/2sG35+8sRjjtstMv41UW5bMFoPsfv8sMBP1H2/g0Vjkx62auYMwBYGy9MR2pSzjT1w7pPI542pqLSmUkXmdCLwQHlPJeVUQgkfdSTdP5r1JZZF2NYCnNsCv5LAqkwYG7J2v+w8t+sq0LiVsS5tIGTxVhfEMExGVJZVilmV03+08vO3n3rprLp3EY9aANh3oe2qSNYlNRjE8ah196cTL6/6YsPR+0emXB2qLJfMGjsziddXUM9vMfR2hzHDpaAIgAZoSWv54NsSYAyzYXMJADQ2popef0k5lVDCRxhxhqirS+qnlszY3ZJ0SiatGTxysSYC4BiJZW1lENR35jNwK4y/mwFebNIDtrwgZs02izGRUVs+PfmYT7945nW/L8S1iMfjAont77C600FgJFIq1hCTRJR9q+6WS4+tnPHN6rIqMLRw87N2DtgdDz20UVNageUAFuu2HDNAjHXZIK3NBs3ApY1IZjoMAgFx7POvzzw4kYBpaOjfii8ppxJK+AijtjEqAMCCPKe8UoaNYY0RqmLgu3jez4SxNW+hvwZ1EkDeAI9s1gAGmMSYNQdJ7l4+cfm5u508+/YTf/GgWTDLRgImkdjFXHhDRLIuqZmZVDxqLfzs/y04fuyhJ44rH9OOACQx7xQFxQCEANoc0BNbNQeHs3Mwgd5oK3NLKQ24HetImRCG6VwAGDs22u/mJeVUQgkfYdTWpnRDA6Rh/mIhz2CmEXvnCYAyAm+3R/q3mjx33tNNjKac6ayZ1+c4zAoRS04IjH7h+7uddPTvP3HxW4hHLVy4yBmpse8sEBEjkVK8YJZ9z6m/fuwb0085eVSwajXbkLSTLCjDroJ6rcXQO2mvzcYOHtMnwWzMBWh9NmgGbq1BIpsxYMLnXl47KzJvXkr1VxC2pJxKKOEjioYGlwix+wGHHREKygNzOWNGqoaebzWtyfZvNbmxKGBtHni5WYOKtL4gZsVBsibKmhf+NvtnJ3177nmbYw0x+aFy4w0FFy5ysDBq/fq4C589IrDn3Kpg5Wq2IXeWi8937z2xVbMy4GF7KAj0dntksFiWcByjy8vlZGdL4ZMA0B8xoqScSijhI4qxYzcRAFgi8tVQ2CVjj+T5DAu82xEG9WM1MVw6c+NWA22K8JiZNQeFNSU8/q36mZ8/6YS9Dm+NNcRkchcnPWw35qXUrAUX2A+fc8XK7+xzxufGl41ph9BiZ5AkfPfe5gyLRW16WJoUMgCLGOtzAdqaH5S5x1ISQ/DZALC5dlyfTUdCORG41+fDnU/1QVzLri2/3vd3Vxzj0NH9Gob3eB+gbJhB8+al1MaNC8uN0Z/NpBWAkXHp+S6dzfkgNuUCfawmt5sq8E4aWN7Rv9VEgIEFOSZYs+EHM8447cLD6z7aisnDoguvd6ILo1bi2G+8eFzFvp8qD1U47EpnxGnmzO4T8dxWQ20O2BqmqhGKid7rCA9YhYKZZDZjyDBOevnlWVV1lNS9XXs7lgjGICRjAks3EWaOY8SSZoDXnNDQbdsPKot7MMQhgKj3EtcaJBJ9+0gzCPVROWzX0f2c9SndrwzjcQE0Cu+cO++F9c8LoJs8+r/eHtv2LJK5S8CvILB0E2HZOEYy6Ublt/el7E82xY6302UTE0BSb+q4dV444oxvb+cRrQghCFiRDkMxI9grgYcAaAaebjJd5cd7ghmGy4MV+dMmHh37zmFffCcaj1rJuuRHy5VXBCnPgrrn9N88c+I/vv/Np9qW3phta1cgyxrZykuu9ZRxQC+3anP8GEk7WnvPde8y1mRD1OakuVjVCLeVu9EVFXJ0JlOoBXCf/8x2brMd53eVDIB+J8kFs2zkZBiRalfxGaORs3L49kP5Ptt2y6LejnH0HNPgKH6XfeWQSOk+2zXEJNa/X4aIsOBYjLDM4txUrs8226owus7Z9wW84uQgKrNhqKBARih866gO9EwydO/BSCmpgcYGAAtmRaDDQWSVK/dgsIBLUuk+Smt4yskM9RktfvyGmMTSGYz+WF6XHx1G2ArBKWh8+8W2Qc8ymGz84w0mG39hN0KLNL80zH/ePuX6YHjd+W2tSo9UkVcBRs5YeGjDaBRMzxtm4LZaX9YB3LvWAfVHgmCtAhXl1rHle3+z8az5C2YtuMBedOH1H3rywzZjwSxbXvgf56A7zr3itfyqb+n2/E6pJEEAggQ+f6qFiOWWcN8RI8pvqXHkqLZBWrqzqqi0ZFurvunYA18/r3crjW0ZQ98J8c9Hj4IJHgbNR8LgIBCmgTAWjAp0ZqCzASgN8GaQWA7gFQjxLNLZl/D957MA3MnDt7p2JvqbtOZHp4NwNASOAPNMMHYDUA3XymQAaRBWgfAygAex4cnHkYAZooLyrMdeCvkv8/aFMsfC4CgwHwBgCoBydOa3oQmCloPoBQh+GBekngfgrsZ3tMpyd8TjAvVAD0V47ewZYDoCJA4DeF8wTwLTaABBsE9K5gKImgBaC8IbAL0IVs/h4qeXdx5nexT4cFxP93t77fGTAX0MwEeCMQPAFBjUICTKkOMXcWnqFMQh+l0sxeMCM5dRj2tYEN0Pio+EwKFg7AdgEgyPBlGop2zQDNA6EJYBeBlCPoMLF77XeZxhlg27DcfBzNaLbxy3zLLbphfyZDACbny/8vhb7RV4YWtFn8rjrsuP8Le1GmvTuo9yIsMaFbacGdjtzje/8LfPqwWH2h8FVt52gZmQJMExpsm3nPXMusLWI5FngxFsBAl4pY0UcNxYYY4fI3e495NfsXxcsGA+Ma6JTLFKSQwTCJLI5c3K1tW5/U455d289+yyf5zB0f3liUdDmGQ+DcizYUwUlhwH2wvhGe769BgtuR3GBLmvR8EAmlcAuBfQN+Cip5f1Oc9Qx3T13IsRkechrTT6c1swa5RZEmnzD1ya+pVfPr+HMpwfnQ4LnwXjDIBnIWgF3foq7Pojul+Pfy2S3N+0WQTNv8TFT/6za0LqZ5VcH+3JOLr2+BkQ5rNg/gwMDkVY2gC8Y3LP5APR7Zx5AwBPQZnf45In/+XekyIT6lDRtYp3ZX/13P1h0dlgfAaGD0LYcmXWeX/RNzlCdLvHDCCnchB4GixuxivtDbh+kYNYTCI5JEuBADCuODkIK3MfLBoLxf1PrgT3cS7oL+PbTy/rVEjdn6UFcz8DpvPAXIuAVek2uoHb7EaxmxWaVs/j0ieP6SPL3rK5avY+sGUdmE8H4xCELKvzWRmKbAAgq/IQ9AyY/4oN1IBEKucuDBI8HAs05rggSpi3V1x2aFvh+UWFQgZEw9Vqrte54AbBF24ajfVZG90pxK7iApZngDvWKFDfALmBZBodrF712Bm/POTQPx3aBsTRr4X7MYEfZ7tg4R/2u2PFE6+0Zdts1/k2cnFLvzBsuQU+f3cbtgQNR+4TM/iECU08KuCIYv2emMGBAJEp6COPPHDpSw0NkHVedfehmYx1SY3fH1SG8lHnQ/BFsKx9QQQUGCgog4L3MjPI03s9x0HMUAz4LBQmCVvsgYD4HvK4BNfW3oR8th51yY2IR60h0UaXukwkGDMNoeBhKHDxmvxBC8gWliAeF1ialEgsc3u9XBP9JCRdCMOfRkiGoRgoaCDjuHGf/q6HmOEwe78LBOUsWOI+XDP3KtCT30YchIQnDV8pUUoBKYUrTg4imDsD4HPB6hMIWhYcBrQG0oOcs+CfkyRCcg4sOQfXzr0JG9KXIrEos0MKisBAUuPaOUdCiB/A4HQEZQCOcRVlj7GBvAqdve6xYRTAndsJCsGWnwTRJzGr/Ee4dm49LkreC2DoylS1CFj2UQja1SBT/PW0CHCkWztl5jJCHK4ymT9nLmz5a1g0GyAgr4Fsr2th0lBsgdH/M+fL5qo5hyEgvg/DZyIkQ3D6e1aKyab7/QNBUBCWOB6Cjsck/WNcOzeBixINSGyDbAZAoxvbMmlnZTRSpskpkMKOxpf7ga+Y2hwbWws2rF65Ley5h15pM645J3rpbTYcCpWJgyv3vOjQmkNbovGolUokPhZxpmJI1iV1dGHUun7eD9485M6v//x1Wfi9bstpCDGiFeSFADoKoDfTxhxeLXY49kQACixobSbIY4MFKC5CRCDW4Yi0Whw+HsBLbkJuCsBgD6wfM9i44QwExK8RFPuiYICs8qwbEiAS8Fe0/vn7joO8L7u2U8bA0QZEQZRZ34QIn4r5cy/EpakHh6ygAAAij5wyULrYC6iQIwvMBW9FVsCC2k8D+AEE1UKS21457ShPMYhOC6z/66FOK5UA5LUBwaAicBmunjMalzz1Ra/uFzqV0uVHj0Ik+FUgdwFs2g8gIKfdcwLCDQ9uwzlznvwrAudiQvl0XHHkqfjWi+2od6fbockNXZPo/GMnwrISIDoPtiRk1cBj6zs+uNt1+43ByCp3kg3Kg2CJe3Dd3JvR3nIpfvhaehsm4Q7kVSWUKeaWYmgiWNKVydJNhAQ05s+NIyDiEESdz2t/99btdETou7BzJXnFUeMQCv0CRBcgICxkFdDhKNfVMlTZ9Lp/DEbOk01AzkBQ3olro6ejkL0I336xrY87chuxOZligKB021xLqxHlCkpibMgFkdOEYC+ryRbA5gLwXocBejP0mDXCltxdjL73yTOv+jfiUSv1Uctl2k6kalMaDTH5SuyGP03665lfWR8sHEgONI9gsV72iCqvtRkcUimGpWS5JMb6XBAzTZqpyFPIDNIakMBcAL+tre2qtVfcl9kQk0jAYOOmT6EycBdA+6LDUXCMAZF0Pzv02Asv2MfoKCgAUxCUD+DK2RcjkVKdE/xgIKYuBUn9fCAgSIC5gCuPOxjX1z4CSfdDUi0KyiDjaDAYIMubaLbtmogEQBbaCwVUBL6A+XP+B4mUQsJTStdG/wtlocUIyT9CYj9klfYmS/ec7hi39Zyu/NsKDiLWHMhQEsmYQDK2bcdySc4MspKoCJwPbRgZZ8fG1v3o/jjz2iCrNML211BZ8zh+f8w4JGBccsGghylyX3t9Clq4Ci+lcPWcG1Bp10MZRs5z927rvY17it4O3oaKwMVQWiDtPSs0jLJxtEHWUQhbX0AwksLlR09GIjFE2fQFM6iuDprZBI3OHZbPaQUSBmA17B9mZRhqfS6gwKyY0fkxBsomqDc7WClllOixHxyATTWV5c858ITvKRgCaj+2rrw+IHAMABGpI0fv/8OgHQKPbIqaS6IUwPos09ocm8AO1txjuMqpuWBRq2OxLJrzRCKfNzDMhz/9xr4VRDA+pby4AvDdZtJMgGaG0g6IAt55DZh7+sf7I9B2uakGepEJRBYcbSCIUR64Glccl8O3UzcOT6CYJLIaIJwJIc+HLQLIOMabZGTnSpa4L1PPv66Bx++fx0a7o2HLn+HauSkw9oAUP0dATEFeAxlHeZamv/oxYO6fBMJDVPyCbLQXHFQGTsTG9T/BZU//ehtk5l55PC5AT4xD2tGegSs9eXgyGuAed/7WzXru90xefk2b46DcOgrl9oOYH63F5toMOLFt1l4xWMK1xK6ccyUqAl9Hm+MAsDrPvW2gblbdBKQLvuU1/LLxf+9wHESsQ8CBh3FTdDZWptq2z8Xn0nHfev+yw8eME1OzOYnwYEPYTghipFUA6bYIKmz3DL5A3IeL8MZWBZSjZwkCA4gyG7up0Zf/6tBzV7rPbOIjnc+0rUjWJTUaYvL+037z8OSbz3p4VbhwErK6/7j6MIEIMBr0RofhqRHpumWHwbW3KRfgMUVce0Qgp8AcCIpxjhOcAeCFZBICgB7cOjFkPKPPuF0oScISBMsL8LpPoaeyuhMH4AaBFbu+fn/foldCAgyDvNII29fhmrmvoC75ihdE35EHl2AYkGICjB8/IQmA4RdblELClhKyl/HJcGMuQxm/Ox1JaAMYPIyQJC9eowCS7mqbDQAFkAVLCNjCt166YODGMpTRrjIbVClaSHtK8co5tyGWXLlNk9rMZYRNnSVTNBgCkgiWIMhu4/MD/d3ViPSGVjCAowcfr4CNDuWg0p6FtsL1SCS+gJkx2T23YbsgCLBpC66afSoqA5eh3XFAsPtsx2Cg14KAoL3rKqYgfeaYBkCQJEZENgQbGeWgwp6J9sKNSOBMxGKys9/FkJE0ANBReHdDCLlvtbUZhjAj4tizAazLG7Q0O7C6KR/j/bZVAZubXE9dD+EaDZulqRg9/hbDIH/MJfREDEASTHMnHPzLe9Y9c2IGHcW8Y8MCZgACeCdtMFdJFsNQIJiIsTFvY78iMSdvGx0OS0sV9OEAXvDjToMrJ0IIQSnACIEIyCoDZdbAwQoIXg7GKjBvBMRWCLSDWcOAAFEJ4hkQqIWgWgSkRLYIo64LApo1AmQjz9ciHj8Ww0WT9ksLEgjMGlJIBC3pMcvakVNvArQMxCth0AIBC4wpEDQLhGMRsiSyygy4EvcnKgJ57iTXKnSVIMGWArYQyGuXraj1EhjzDgjrYaAAGg3C/iAci4g90d3ODHZO1xAus8JQ/H0QLnXz0IY4qTUvF0CkHGFbQhkJZQDHtMLh9+GY5WCsBGMNBG+BQQsILv2fSMLQFAg+CsBJKLOnIDeE8QrY6CgolFnn4Nq5t6Iu+eAOWcgEgjIajN0hxHzk+2mk5y9CLCERsGSPBYFmCyELyBeCfY7NTLg6WoGwLWF1yqYNBfM+YJaDeSUIa8DYAqC5UzYsBMhMBtFRYJyEiL37EO+lJxv7s5g/+3O4NHnXtsrGp+Eevs9D7wG4aqj77WzkATyLRz/oYezScHtBxcXt9Itnx990xqOZcO7EkbSeGK6l1JoHrc8z7xEhyg/AQxrK8SwCmvM2ZbXggOSiLEBmgJmP6P7d0OI6Of0m8uphQDwPi15FR/79zhylwZHA1XOPQEH/EhH7ZGQdDQxoQblKrMw6Cnj8FFz29P3DlgfCbBCUEpYAcmor8urfYPEPGPUcLnt2XdH95kcPgWPqEbZOH1TBunfSdRkyuxT/iC2hGVDmdRhzN9j8G2HxWp9kXh83RauRU18A0a8QlKNc0sUAkxqTRFYzwOfgmtk/R12y2WOPDa7YO2wLQWxB1lkMzU/Awoug4Fu4+NFNg+7r4s9Y8Mkq5Jyvg1CPgKxEYbDxgmDA0IgjjoewdMaOLUAMM4iug6CpKBhGZ16I56j3qfB5vQE5tQjMbwLYABYEmDEwvB9g8uj9HtbXSozDZmTV21D8BJheRMB6E994YuMQR3YDfntsBQR9FYRfIihrBr2XAEExA/QLxKP/QGz7nntmUCOiEo3bs/e2YaBTJIr9WAug1quGUkJxJJeRBuPAsftcuWXLKycWOLNjvrZB4ObHgd5LG95rGFx7AoyMkWh1LJ5gFcjph1LOTEI5DAAHAkBtbWoQKrnPnKkI/hXvPXpdHzdRjzI/Hmb2Kt63dBMBKYNLnnwJwKdwTfR/ELZ+5hICBpzgGSAG0TcA3L/Nno3e8DMBLSnhmNeg9J9hWclekwwh3l/L4JTBpalXAZyBa+ZehYh9KTLO4BMMgxGyBBwDOOYf0LgOY8Y91kPJNsRkZ2zPh1ueqAXANbgy+gQC5t8IyGlwtEGx2mgEgjYaEXsUMoVPALgL9VEJDMh+cu/VbrsVsGLtHPz4qfY+W/QeX7/3F8CFj7UC+D9cE30U4H8iKPcYcBImkshphi2OxLi5s3BJ4qXtXoAwAEEWpNgHjunKNDOsEZLu/XTMo2BcjzA9hnNTLf0exyfg9HjOaw2Cj56A85/dftn8ONUOYD6umfcoiP+BoNxvYAVFEgVtEJQHYiwfDcLT2yMb14LahdlviQ96AB8SeJ10H8L/PjL2xtPe2hLI7UvK7XoxEqfzXXsrs4yCwbC49jQTNeVtnhTuu/7zQIUCg4E9n10yYxTRsibmoZQz+cqjaQBdL69bQ8yrcFC8xW4P+Imvdcn/xjVzgwjbP3DzQ4ooKCaJvCYQzcY1s2tw8TZYAn2OBYZFAGMF8uZn2Ii7OhVv74TcYlTWhphEbAaDEpfhmuhRCFqHozCQgmUDWxK0eQRa/wIXPf1C50/xqNUpw2ITDoNQP8PGt1Jv4srjzkBQPAdBQbjJ90USBvwAPZ0A4K6hCQd+Car2HrX7fHkMdUJkEK6fZeHC1BJccewJCNrPwRajiybOAgCxRlBYcMwpAF7qo6S3BQw3NYEgPMICo9yWKJjXAfwXvpl6oHPbWExiRr8Lgr733qVz75hsAMKCWRYuXPgW/jDnRJSL52HLiQNQ4+E+P4KQd04F8PT2yibOcdHY2HWOcZuXcbFCqsxMtfW1MgUAM1Pue7Y0SrGZ44ru0x3RhfFhyaMaaIy9EY/HRSMaRaoWrvm2LMWYESXUArHN4zg5xKozsYYGuWns0k4Z19bCJKgfKj8zxZJ1Iundj6HKpnOcQJdsAWBslNxxNph+S7n7aIxKmkfOwXeee0dTLhc37fnBFsfbDQYAArbmmZodxqgAQfGORbqIGE0FC1w01wmkFLNti2pdoD0ANCWTQ0uz8rfZMfO7s0hskjFu7iIE5MEoKFN0gmcwbEFQ+jhc8tSzfVaPfi7U/Dn/g3L7Zx7xoJ8XhBXKbAvt6lpc9uTFnSve+m10Kfjnu/q40xAK/NPNUyn2gHjn7HAuw6VPzUc8FsBM6G2uo7ZgllvK5Yo5v0al/VOP8VdkEmCDgCWQV4tw6VOHb5My317FX2y8V845G2XWHQO6QJk1IpZEVj2IS576dK/cHpdmc/nRYYQC78ISkwaezDsP6r60AQkodQXWZ37qJih7pYcGnqx8ak+vQw6zbK4+7jQEA/9EfoDnx8sBQlYtxKVPHT8ciblF4ZXMQV1XHyF/UN1OSJ5/Z9dwwblzgUvS8kBwA419VhjbHxLo+Txsz3H6qR0qux3U9N62mJLiuAAlzJmP/PfeD6x8fkneyQdGkhghXNYenzpR8sGVQmT19ifkEgDFhGrbMSeMbyIqYokxs66otGRHh/P5Y2YuvXPhwqg1lNXO8DyQBEbcT5A0v4aUyQHsAHdlbQsLSu0J4NkdXllLBNAQk3hsucD1i5xtdiu4RWGBTOQJmNwG2GKCR7Lof1wug/FIb9VtUJfa9hdk3SINBuEv9pXIqMsgRAV0kXMyEbQBiHbHFSdXgh5qG/LEOlx+/wsXOd5LfCeumvMThKxDilqYBHdJBprqjjPhUkm293ljNrCEgEAeOfMNXPrU3wBgG2jKRdIwhl02/8JVc55H2Doa+WKyIZfpCezebVIcumyYCUT83WevPHKdbp+u27OGbIisUisePfXy5/LaIyA2xCSItAT0Ncv+PfHOlY98YvnWNYdIEnsSgEAg+O4+Vbs99uBJ//tIwQ1GUK/JkwDwQl4R+utjt5/epjPbHagnFmyFLaqi8i1/jn7vEe7vejtLPCV1ABLfe+qqPZ9oen3W1vatB2QcNX5S1Zhxq1rXr5haNWHLhPIxz/7rhP/3DBEpDLDAFiBc/NQfP7Wx0FJjCswiQJRvz71x/xm/fcX4749XcYSZQyf8+/vHrWvbPEMxq0Or9nj755/65VMHgJwezwkzoZ4IdUltQ+DbL1+715OrXzl2c9uWg2xp71WAIRjuCAUC704tH//yn476yZP7jx3bDlDfepAA/HfjXyf95p3RN562ZAOcw+AM4JUYDjBoXY75kKodPYyfciApqwXKbQ3dT9yJCCwEAKZ9AAC1I1DSZEDUpzQSAOzQo8jkt8ISowec4AUAlmOH5dzsuWLi/ferH9IR4nGBHybSuHrOq7DlyVBO/5afG/AHQJO8hMrtQwIGy2ISyeRGzJ/zOILWGcgUOSfBo/JzDWR2HIA21O/AZL+96MyPoyRsOgSFIudnIo+iU4XrZ4WBRZkBFysDgcEISAFwKxz+LC59ciHiUcu1jnehfkCubAgkGmDR0cgXkw3c9AeiSjQ3lwNo3SbZ1NdKAVL3vPv0TzdGMqejw1VGwpbY7+9f+cPrn7/hh3tecXLw3bpk/qpF/5h088oH4v/1wtWxDlGoKRjlP0cgR2BFZuMPJ94We/hrex5/foJoTTweFwl/8ozHCYkE39p4z5R/bnz+jkw/nJIhwzBQZmFitvxhAXpEN3yuR5Fpt+ZcQtsJgVl3f/Pkjfmm71771r/mZqQTUsRgGKxvbgKEQFP7CgRa3sfEv3126XH/uOyPL59x7Y15KOpvsabZiMk3n3ldUyi3OxXYiwJIHHnXhee/QAtu4PiMgJV4o1B737fOnPT3M/9fi5PeNw8NCGDr+lZ1+OJ7JuMQbOqUi6fwbQg+/l8/OOmd9lXfXfDqvdGcpUMOa0B5azAChBFYmduE2Q+cu3r/u7923bLZv72Sxo/v6LeP1cKoVPNSqiZYfv9Gaj+MC4WRc+15z9rGPEMZcDFrZ6gQABwWyGjJlQFNegDGHhh7AQAad3YnXH8VcuFjrSB6w02OGCT1WexkBTogvH48TG/CrTdQbLXtTrwCNQCAxA4oiBnehCbwmEeBLu6WYjCksKDYPe92a8UdgE8MUHgFivtPzu4BknDKt/85dGcTBvNW5MyJuOTJhVgwy0YipXY5JpgrGwarV+AYFHcLd0KiwNv+/CdSRoKQdnJjck2tKpdO53MdmUKmtYU3Z5u/qJnD7377ofwn//G9z9Qvvek//2l774KmTEtNoS2jkXGUW8XEaM44KtuRVu/nN5z057cffuTR9x6tSiDReU9jM5cRACxvXTepI5fRudaOQq6tQ23XpyOdy3VklCCxWIPRw1Pi9nbSFy+8fPqEW8964JW2d/69Ir3+xNZce8hJ5zV3OA5ldJ5ypkBp5ej2vMpmMmZDZuvMlzveuWHyrbG/MbMbVvCfRy8AcvPKxsr2Qiaca2pT2fa0k2tP5zucNNZ2bPmcYZCVeLMwq+GCXz/X9s7d6zs275vtyGiTdgom7yht9Gs/OPjMzQAo0a3Y8NUv3r7b1DvOufupLYsfWp5Zf1J7riPktGc1Mo6inFGU0wpZpUxHQeU6MnprtnW3d9Jrf73bo5c8c8mjf9gnWZfUcY73eDZim8cxAxgXGvVowEiAecSScQEABDQ7jIwenppJmonaHQlR7JVkIqPZ9fwAaKxNmZ2rnABgZsx/ODZ5Sby71gQyJJhVg47aMMBchh1xVQFdExrxq3CMSxYpBj9RweKwu++ykXNMF4NPCw/QZiiDkcxoB+A+P5IE8vp8fOupF7Hggl235YIvG8JmVzmNwOLQnXxNgXVAGzUZhixiESBAwg5Qa6F9fZBk9rT7f/Lj51qW3bc1vXU80sohCOW77IhIgCABskDCQocurEPL/hc+e8N/UwLGZYICm7w28G2FlinGIgb51XmLfIgAEhIkrD4fkCVIWpawlwNAtNa7noVurPdLD//PaQ2rU8+vzm8+Jd+R1VRgTSTcha1kmwMU5AAFmIzt1lAUgjSZfEtHYQVv+tJBt597vay7S6M+TgAQRz0BwBPvPjdOC1QDZBGRTRAWmDhoBUkS+NA7z//5K4WVP023tirSwkAICWIBS1qjghXvShKMWEzEGmKS6pL6nH/+/JjEG3c8/2523ZmZjg6NHGsiwR5Hm9gYycZ47wRJkJCkiVVrzlmT23RQcv0zj/3mxRv2TFDC9dR4SMbcROUv7D779YCiLbDEiHlFfFJEVoHaFLPfaGBH0aEGnLpIaYCBCcygBGGnKSfXpI5DoHm5cEvmFKkC/aEAbRr4bnXqBAvxOPX+cpuwNOmeSfBqFHQBgoo/lMTstqygvgmlIw+vknqje3+NLl5zYbjPqxiAvQLxuMC6ibuOG68LnmwS3rMfNuhqezQiUvrp83+pcYweC8Ne+BNuRiSwct4/vv+ZJ1pe+990ut2Qlg4LY3NYWqI8aCFiW8w9K0qQIBsdedOc77jgvrUvj0EipZiZUo3u7x1O7iBTblkUCYSoImD1+ZQHLCoPWlQetNB/3zkGMYUdqWaOm/QoAKQaa02sISYxL6W+/O/E6Q9ueP7eLR1NoymrFYSQzEwsIaxIUNYEqjaNE1XPjqaKpyqDlWnYJMDMDAgiYXNrTi3Pbzwv9sgvDkciYeIcF8uS7sJtXcfGSVqyDe5W38YSVFDOez945upjl+RX/7LQ2qGIJDHYrVHIrISQuiIQWWTAmH7uBCtZl9TnPv6buY+3LmnclN4yCWmlIKQEQTAbRlhKqywsyyLlFI5ESJYFLQgmsBvJgyCbcnA2qebdbnj7sb9xb8vI8zpdcvjZrWEruBQ2gTBCJBl4K2oGWhzuLHayQ8cjRsZILsbYA0BaMwg8unHpjDJguGNO3du2A+6q32er+MmpWGSARcDVcz58ysl3WQlq98rVDEMWwCCoByMBIOe0wAq1QGAcuod4dy4I8Thh5jLqcrv0bk+eUkAKWBBN77Qm7QRA6oAb34vvfG+AP4o4qLMNPACvFXz31u0KSABXn9gxUoOI18cpgQS/tuXdCVqYMre9C4FAgnMKVaGaI5/fvPSEtGo1xFIjIuwaLmsfFxx1/+hg5VtN+dYJ6/XWr7cW0jbciYQYIGiYjOVUXrXo9mMA/Ku2vlYCtQZIYVr5xOdsCv0/FSwYEPcrf2aWZSKUaaGO85an1+1OTIbZsxwNK7s6Yu8XmPKLf5zwv8td5ho4SUn93carDvzr+4/c3tTRLIilZgELhg2CQtRYFWv2rpz8s98e/81/nlRzeAuDkVh08/Tr33jg9pW8fhY684EEMlTg/2x6pw7Ay4n6RhH1xtWST++uBQPuRO/WVdcGSmr1t3ceuzrrpAGWzDYsBGzPY28sy7YxoXz8fxYDePeUqwq/eeX2ab9ffPtdTZmmABmpmWB5aQ0ULAvTeLv6X7sHx94+qWz0ux0mbzdlW2rf0mu+1+x0jCblKigm2EgrZ7XdfOyZD//8C0j8zy3RhVErNc9PcYkKjZQJy+CrEDLKeWcHSd7FQa61RC3KVSfDUWcvpyR5a6V+j2Q0g5kqg8quAtCxY8qpeyJup4+/nwA0xwWu/Hc5AuUVMKoShBAM3Fp3O2OCH24YUwBL7JSR+2do6shhbDADIeC1AtsZUvM6924iLxdIe513uyHl/rkpGkIB5ciJKkiEoM1hO/W+mp1Oc+5HNuB+y0bdFA1BWGVI56tgOAzKzxipQS3zXLlbVPtkZQHIwwCQDACasC6/dQqMAYxQosyypwXG33v+3p/8wX8fed7yt+D6GT/z0M/+8+/1L16fT2c1yOsjRMQFKF7Ztn4fwLvrHjHi4dP/cB+A+wYb25n3/+yTD2/5z49g2LA31RGzogrb3pPGNbz++Zt/ZRpiMr50BieWAswcmva3s+9oyreGiaVmgiQDwwESU0Jj3/zOvmed+IMjz1k9DwsAL5/sp7O+8u6XH/mf/964vvWhbFu7gZAAuXlcmzOtY/yxpGoBJIC0zk/V8ImigJcgDhXWF2zJt4agjJHlQbtKB9+psaseKLDzpmXEdKNxxD5jJr75MFzFu8+tX/x7k24dS0ZoJo8xTsaEAiEcXL7nRf+JXf/nXrGAZy9q/L97b33vkWfbTUcVmNx3WgiRz2X5ufWvf8OCuCXV2C2XtBbgBFBmh5bKgoRmZ8RngbZhcpATgJwhKKZ+lRwRyBhASoQMVCWAtduunPwGem6eUM9E3PnRCZC0F9jsC8Le0NgDwBRcs3AsZKgaxikDUwSC3AIzeT3yMYmPCiaFGQ7tDFvEnXQBN0G3O2vIVUDToHkfMO8D0HSAdgfMBKR5DIgrYHEEhAAEAcVoOR9mdCWU95TNFScHEXamQilPNpgO4qkwmIg0j4FwysGiDAIBt+Zy557DOr0kvThQR65jqnb1dddNIACaGczaqghZu4nRV6764p3f/i/cDsRhzTrtAlr0r7f4q/t+MvnE6v9cnhdUQdzZMxAAKCgDUwAAteiq8sBxgcbGvhZTIzD9qLB895SHCuf8O3H0AxtffDCdabfBwq31z6w5SNYkMeqlN7/4t6/RO7cKxBpMYmmtRCKlDt9/zSVr0DwDeaNYCAsAszCosqqav7rHCaf+4MhzVs96+QJ70awFCiBM//ex4l0sQlO+Q/rBvW7giB3UzZ3/NwUCkHGy05h7GyCETdmmEJiVXRay9giNv/ytz9/6c0GU9YUpQbiqwXW9HXXXN2MreeuxyGh/nIAxJlAZkYeUTb/s+c9d82csmGVH15VzbX2tWZZcRsn17da1td9945Dbzl2wWK76MacLfg6jRMGgjTNHnLfw8mkL5n13pU8vj20ex0kA1YHyd0WWocFixLSTd6EditnwjjP2CIAygjQT7CLtM5jBlkVUYFkJbItbz09kpKQGUgoJAPPnTYU0x4IRBfgIgKdDUiVsy6M2o2frdqb+27iXsCvAVUrdJ93Ljw4jEjoM4LkAH4ssHwhgNwSlgPSipIYBI7rdY3a/1zvNutsZ6CubK04OIpw9FAq1AI4FsgfA4d0R8qrbd8oG3Z75nSebnMpPc9uU9TwNMQyHpLWnNe6RVV++69tOzMhYLIZkXVK9XL+A6HDiad+ZhqCwHPTjgREk+r68bh5On4VTPB4XiVMSasHLd0/8+dK/NbTl2mxiYZjcPE+WLMaFxm66bNoJnyOibKwhJpMuKUMveLmh6gev3Ph9rXMGQnjuP2Psqog8sGyP+l/PvvC9kx+8LPjQ8g0Ky90k4ndxVZ6Zy6ff9oWf5nIZJuFxwxgQJKjCjiwG4CrWxriR9CvkjdodxvTx4BBDUVnQ2jcy+Zo3zv7rD+icvxMWRi2/kKBGyiCWNMxsV//l0//tmCx3jpOhEZJyEqqf+M/nFsz3E7BTAFIJ19MQa4ghGY+LSZX5e9/cuPZHOeauDn/M2rERfGXL0oMBrIzNXEZJADNiLqFmcqR646tNAs4Isq39m5w1nlt3B49FxHAYRjEhQCxNPwclgpEWSeGYcmAoysmPI/mJjFdEpyCEM2BwJlgfjaAVdtUiuxz+gmG3RUC3/bu3He/+bwm7BvxETz/Zc8HcWmg6G8CJENgDQelOPY5x73FWdVXX6N2evOv+fjTucU/ZANfNmQsWdeDMyWCxFyKim2z4g5dNo5tYmlPOtH4a1BmWhkZbo5r+6+AvfPVccxvFG+Lsl+rxB3jrsn8HMipv9abSEgg5J7/KP8+AYKZEPYGZ7T1vO+eeTU7zFFJwXXMAMzRXhavNqeOOiP1kzoWrOnN7FkYtzEupBcsfOTsbNBPRxsprSmpgk6x0Amv/POOSm/bHfDx0ylV5wJ2hNXMo+o9vnbD7rXW/XO80HYICMwuSBBgmQ2XKck6aOuv+N3Er4rW1JjEvYRxmGnPDaZOgDYi6sosJMGxDjqeKd16vu/m7dNdfZXxGnBPzulrIRxdGrRSl1PH3fee0bEDPQJvp1krdUNgqw4zR036+EppiNXuaJBb1EM+MpTMYiYSZ+fRfVj+17tVsjhAh+FYqsRYMxdgH6LKGE3Bd6jVq9CZtTDOkqIEe2cBIwYD0AHGibQGDpEs+6P9wzAwpCEZhCISIhph0LaWkxvzoIbBwKcBnIWBVQ7Pbqybj+C+j8Iw/QvcSOz4RguAuq91xudsPnudRwkiCQaiPE+oSGvFoCJPwJRhcBEGHISC8XkSGobTuXGQA1MMVS53Hcu9xjyaUQ6jduKuCQajzrKUFs2yY8i8CuBiCjoC9C8smkdISAgVWu0Eb9CjpwMwyHJJ7hSf84twDTt0QjUetBHVNuHG4ZIq3tyyfwAKVHpmi06UnGAgIayXgUr1TCRRHfa20E0IdvN/Xb1npbDoaOaPYVTJg1jpcWWEdVbnveTee9F9PRhfGraQ/8TemjCTCqvaNdYoL3Fk4h9lQKCAmBUdfe9D+M9qvee3+muc3vbTbG02rD2zVueNH33Tq3DQVpucLOcCBgRSCDGuG5sCoCuvAwLRfXTnvuyvREJMJ1BsggZtf/UeVY9R416rtMjGZDVvBsJhaNu4nRFRAPGolEj2rjKQ2j2MBwrsta84pqDxDCJ+BbTggRJkOLH7g0799nvh3SPaTCJ7wYrcPrHt4S0chvQm2NY2V6Zy1mRlrOjb36knm/nzOnrNyDZueKOSdPIantlYREJAzUIpZStqx2lUuqcItfDKQliMBaGHCwGBt2uuSGpcfPRkL5l4Pm15EUJ4HpmpkHIWc0u7LRm4+RPdmagyvkR8rt0kfCJYQCFoSYUsiYluwZdF8rBJ2Alw6PyORMLh2bh0mYREC8s+wxGEoGIOMo7xiqu5iw510u3cE7mr9zWBIIthCIGxJhG0LEdv60NpOvmySSY2ro2cA5S8jKG+CJY5AwfAuKxvXO8EF1kGl1UQv5kfePwYScpQOr3uhbsGNiEOk6ht7TJo+vXp1unm8Iwxcop4HYrK04IllY9YDwLjN44q/vQujFiVS6qA7v/7jN9TaL3JHwfGsH8AYZVdGrJnBKb977Iw/3ogFs+yUr5jicYEEzI9fum5i3hSOQV4TdbY/IUk5jaZ826dG3fiZZ3/6/Pw37l7+3OJXM6tvfTu7/utNhfbp+XRWkyIFEENrw0GSkfIK62B72u9eiv05wZ2loOoJAB5a9+o4RaYSxmWmdcrJIlmtQ+89/7kF/wKD+ikITahL6rXr15W1FTKzUdBE3lzKzIYCNsJ26F9EZNDYX6eD3kfrnwdni57Tsy/w2pm1FLFCQJ9Y2fCDmbsSl3cQmmG0IVWM9ef3IoMgG+hfOVGnYrpyztkoCy5CwPoGlLGRcZTrL+98IbtO4yskZg1JhLAlUWZbCFkSzArKrENev4ys+jcyzo1wzIsICGAEufolFEFDTCKRMLjqiNG4LnongtadIJqBtKPdYrwQ3sq+1/Ph5XkA7psT8SZaiwiGW1HQbyGrUsg6DUirv8Jw7kOnoHzZxKPVuG7uLQiJeyHoIGQ82XR5BoYmG1u4snHMW8jqFHJOEhnnRhhOb3c1zUHw/16/vdphM8aLpXROmiIcwJhw1S1ElAWioneRUb/adsbkdzfS3cm/OICEpcnZe+weGwAgGWvo972Nem65E+793hlvZNf8r9OWUUQuSYCYFZUHrD3k+HtfP+fmH5t41MIFizonfr/iRGrVK4fkLRMBc1dogkBGGax1mo7bqFuOadHp8el8hp2OXAFZnYdGAYBgCQsRKSPlZWJyeMzT0bEHnbgodt2PnV/oznJIMU8Jbyhs8XKcuvyfvpymlI27i4icfpWLRxj6xkt/3LdAejy0m1Pl/SqkIYwKVzwPoKj7M+7lP351r8+OqQyUjYNSnv/XBRFhfLim0H0fImJPUaQ3ZZvXwrbA/fhuhxckiynPkYLw+BK9XQtdgd+r5/4KYeu/UdBuq3Eiy1VKvfZwufwaUlhuZ1kGcqoFebwENs8CtAiC3kRHcB1+6LXfAICr5/4IQXHkgH2KShh++AuP38+eAVvei5DcBx2Oa90OVEGc0NXJN6cBzW9A6+fAeA6SFyNkVmLV01s7KzAzBK6eeyoEhQasn7grIR53Y6uXHz0dZXQvQtYB6Ci4cbiBZAMQAt26HDvmLTjmeQh6BpCLQfmV2Pjclh6VxefPPQkCZcMZM/BznF5et2yCJlMOA3TlPLKwNJlxoVEPvAG4rTB6H6AWQAJwdH6acckUvunFLECSRNPVx16y+Rpc2m/SS6whJpPzkurCh35z8J0bnrk1k+kwxEJyJzNPWJOtmlff+uTfvkI/v1XEUWsS1NVCwleO6zta9lYus7pXcVNmFNjxJmgbFhEkBWBJt7G1EgjL4PKaUPnj+1VNufPRT//f4/9mBcQgkejiR/rnac5ndtce1aLzPMQiqCyz9/gp97wKr/VGr+uMjt1EKQDrc817OtIQQBqdVX6YglpiZs20VYsBxOtrTcIjQXSH7xF9Ys2rNXlWATB1y09lEgYoKP0uAKCXlUpEBvNnF2D38vp9xNBTOcWjEnVJhflzLke5/T10FJTrriviH/fbnYcsC1ndjpz+NyTuQsB+ski3UMIVJwfQlNVglJXcejsZXnVlXHnMfgjIJ2CJ8ehQDoj6f8r93khhS8IAUOY1aHU3BN+PdeK1fvtfNUACMeDaNaPxYSJFxOFWg/7j7D0REQth0xR0OAPLBswIeV12Hb0Uhu8G8f2w0q/2W0LJl01zczmcwrAvyPwcp43ZtsnaYiAPf9JkEAnLIVVN4TUAONlf9+HNKSYAeaNcMoVPHgMAKWBJa0NQ2pn+1Gk8HheJuoT555sLx1z4wpX3teRay8gI46W1GJYsauzKTV+e9omzaDx1xDku+u2ZBCBkBfc0hV4uK2ZQwCJhy4DNAtIxxpb2prAIrg0EAksmllW/Orls4tN3nVC/RBLl3Fm922K7O2oBJICcLvTIcSLAsCVE0Mj3Gk6qX0RIoL9+TX5ljM3Z1hrj8/PcfxlSkDK6xXHkBqCLxNAb0dpGkUrAZE3uYB0giTx3by8jA47gGZOmvbsEQHzpDE7AzdUiImbmyNgbT5+0pdAMElS09fnwgH0B7fR3uUvpuCtqhSvnXIrywPfQ4Thgsvpd1fmTVpktkVNbkNfXwIgbcenC9zu38RN0/SoR9V5QuCmr3T5Mc0vuvJ0Jl/wALHimCrpwHywxHjmlik++rGFLCUkERz8GLf6EseMe6reT77JxjIakS8SJwYCSjD8frZDfSde2o3DJD4RbWsrQnrsPtpyC7CCysYSELQkFsxAwV0BkHuyhkHzZdDUo7JLNLScojED1v84cJ6d9qmd5dLN8iKSQTefOPm3rffgt+iZTu+OTEFCsd4Mx6DSP3LgZbGGt0UYBdb0ytRiUqEsQM2P6bV+4Z73TNJUcaBauYmRoUxWulp8ac/jn/3fOxctj3CATVFe0zNTG7FbRq9qoQUBQtYy8PLVswp/GWpEtY8rHrvjZCReunyWmdDjswJ94CAmgATKGmKtY+uvB1OjmOOWc/DTPc+mTEBiWRIDsty0SGg2Q3ftc9caq9nUaZcE+0zaBdLPdVuh/LxepRne3Dektcx1Wbl8J+AqSRFgE3r9zn/p3G5BAoj7BSHSdZsuWLZIIYTc1YWRdbkSk4VYc2OHzCIAEsRxMmbJrMXvKKR4XiCUMrp67PyRdjqyjwSiumASAoCWQ1zdCqZ/jsmfXAejdWbZngu5AzJ4SRh51MYFkQmP+nP9DRWAftA9kFXjN7hzzPpT5Hr755D2dvxXr5Ev9/NeHBcmYQDKpMXfO71AZOADthYFlE7IktFmLvPkBLk7d0fnb4F2Od4pssvnCtD5pjgIISrvp9DHHua71/ti8BHZYW6Nu+PSk7mQKt2ajgE1ilQGAGVHqrAwCAPVRaSWfVLMavnHjCmfTHMr5yagEGKXD1RXWERV7X3T7qfGF0YVxK0l1A5Yui1ghsZWz3b5hA0ta0Fi1uO6G2/wrux2/cMcYd+NC8fpak0CCQdBJJBGPx8WymcuoH+vHSBIosNodumdLRQiBANlrDBhY2us6e4/TDoss+iaUMhsK5gcIVTAISGlTz4Gav3z6JGgFcsmQYLBB0KZqWfaw2INyiEctkO+hcG/H/yy5VbY7GQIERjpjLkA7ztQD3MdNCiZLoI+8OrfxuazGvfmuAJctI4/q+jsEZKB4QzswJBiSDHLqfHwzdR4ue3Yd4lHLXX16K5UPZaXxjzAaYhLJpMb82UchIM9FR0GDUGzyNQhbEgWTQk4ehW8+eQ/iceEtPFzmkl9L76OAeNyLsc4+CAHxTaQdPYAb25WN0s/DyKNwceoOxLHryKbRzXHKc2Eqmy63nOdughRijUVkEIfo8456EY/5b9xbpdiMheauAD27AfqQDKxkwHWL+Vgwy6ZESh12xze+syS/6lzTUXCY/CoJWllVEWu/wOQ/NZ5xxXW8MGqluuUKFUOFHWnp/v8JJOBoWFLsb15mOxqPWvF4XHhjZlfuKZWghIFfSjUGmUgkTLKrUWMXEjCO0ZTXzmQ/x6n7dYZtu+91doNfMf3A0XspKYTLmvMGCm3YklbVuNDoiYBLz++zf2NcgsDz7vlWLG2raXCM7iRUsBERBOjIiTP+ynBbwfv7+cdqN4XxFsnRbnPREVRNDAQFkSWww8oJAAisBVgX06ZEBG0Y0pIZAN6LlfReTkt8Glllir6cBIYtBRxzPi556gYsmGW7vvpdsHdOCX1B9D3YLnmr39+ZDYKWQMG8gg0dp+A7T2x0WxYkjGcJfPTucWdbEfFdBKTwGGr9vD1sEJQCBb0ULemTcdETa13ZYNeRTSKlJQh57XTmOAEAEzMkQcCzfPx6mN3gt5B4cuXi8RqmAqZbUR8CJBNs28tx8vaJxqMWLlzknPKvH56yNLf6/wrtaUXe3EGGNZXb1jR7/AP/qfvL91V8rsW1jUVdZEDXpG8Ja4kFARhXYTIg4BjOsjP9R85101IzU7ysfmbvrrwAQNF41AIRyyT0BQv/sN8nHvz+ETaEf0/JV8J/cnOcxnnVajqjaxICETvsegkb+x9nyiMohEPlb0gH6NH0j2EKAZbvdKyagThEY2NPVmesISZT8xKaV6wIvdb6fr2TzzGES9skhkZQislWzTO3ffIXzyMO0d3q851P77WuHesSMUae6RwUMEWqDW0TmAELEJJYFGPAM0NoxRCkOgBAYKxfXVqcg5AkFGv+57p6BLLqLlz61M1+SY4eDKQSdj34Fu0Vs8eC6VPIaRTpCeW22zDGgaPORWJRBnGXFrzTx7yz4Mvm/6LVYP4M8kVlAy+sr6HNefjJotZdTjaeFVFgHXS009MtBwBEqAiG1xebZfwcpxUd68Y6wsgeOU7MQhqB0cHqtYCb4xRriMlUIqUue+yKGc9ufeOOdKadCVKwqwAMh0iOE9XLfr/fWV9yDZOUGcw7lGqsNQCwf9Xez9t55CBIoFPpk0lTIfDg+y/8VNZBJ6lOIx61sDBqIR61XLIJOJVIqZffe7lqnzu/cvWdyx9//dnNS188OHn+ny0QIw7ylXBq3QvjNHrmOAEQUgOVomI10NNq6YE6t7fSn2b9YEnYWBshIMjrfQ0iOKqAFS1rvyQTMKlrEtw5Rq9pIjPTIS/96u/Npm06dbVbZybNYRnmfUdP/S4RcWfvOw8+Bb7DSe+lPdt0IHnuCPwIU7lNJAg7RLpwzVuCLQ1bA2g6IpBSgMOyDQAEalO+Zj7e7YszgM2lDCDwf4hDoGbPHRRMiT6+U+A1hoOkYxGWFdBG9++yZYOQFHD0E/jWM4vdfJ8dmHwDZQT03z5hl0HSK3AbMEciZI3qSqztBWbtWk3maVz69As7LJtCWhRrLbH9cIf9/16/vbpg9JjOenEAwICEAIxYAXRr5NcNPr26w8lONW4DH//9ZhDIMlTYp3raps7t65L6qfefqrl7/ZP3NedbK4iF257CjUnT2OCoddcedulnYoee1RKUtgn80jIBstDvx/MCuvllEHd/6qcrxwarX0BQgLhz0pecccy7HWu+duBd5/2CmctFIqUwz3XpiTpoZg4fd8/Fdac886sX38yuubg10yaz7e3q1Y7l5x/acN7veWYD3X/9egkAmzPNk7SEDe42wRMJqakwtbx6nSuTfhiNvkziUeuwiRPTYyJV/0LIAhvuHCeyymzWbacfdc/FPwwkpSZvjDKRUvGn/zJpZsNX73kju+ZMzinFPkPPsBIVYWuv0MQ/PHjK/74U60wY7naPfMKLyu2vO0l0I4sKOTweAQYQEAw5QNFXIQCluBA0qg0ALBAY18yugeF9oAzA1LcdmNsKW6Cgt8IOLkUCBpzcMcmQmYKhdBouuQuHB4RDIYmLypM8y0nQ0y6zb9P23d96b6GUNmNAVDF4vZJ+Iso7C509qegQWMQuiaefxHS32y4g8MywyCbHowDUoG80ZLvhlx7qlePku6uEMISwDKwGutxSPVALl15tnKmm+3wNgAWRBdl61uTjNt0AtykIM1v73/aV5LrC1unkQDF55CrXQUVc0OGLXryiYdSNpwWJ2HA/VQYIZNiGGGtVvrnk87fUERGiiIsUEmbuxAMuv2fTC9FMts10a9sh8tkcL7VWJcbddPq5u/+1rjEgrbV57bAyaq+aG049KiMLexbyecBhTcISzMYoYbA+0zTnnWMPsRb9/WoNAC0qt5uWgOspIkFeLpcF2fqZMbM33woA9fWMRBEmV32tMYkU5k489A8b3288t53aQJCux4qEKOTy/FLzW78bc+OpZ1YFK54JyUBua75t6p/evPtTrciORlZpdFYwZwfl0p5uTXjy9XNu+ilZaekmOfcS2TUptkDIOPlDWGvP3hr00dgRcI1NxMNAJGcAIWlYEESxZgVCEIzmDoTDrYDP1lNyPCRXeRNJ32EQvBpblEHOynnfbR/qUxr1IMzHUVBDaFWtzYe3PtsuBZo2pDIkLFrhdt3cTkQFOMW42hyGiCXd2ovFElgBMCzo7AdrYTH2GJJsSLQPi2yukYciZAXcIrHD0zLGd8ttyrRM0m4fp26JoSCpWU8pr9q4GOhqF98dHr3a0U6PFhIMt/ySJcTGM/ab04YYJOqS+pP//N5ZK8ymTyDjKBay+ztKYGBLobkGtqzBQP4gA8AS4CxvtIgYMchUIqHQEJN3nJT41963fvGWN6vyX+HWfIGECDAACEFOe1ZvtnPTNgesr/kGKLMBlAZyxhAJQBCYFVsV4cDewUkPLPvEX8+hsZSf0RALLAOQVblpBj2sD4YksoXc9KUjT2l1r2QANyQlTKwhJm/+xI/fOuauS377YuCdnzotmQKRsF3pETnZglkfbDl6vdNxtFAEozVQUICBhhCSGMxsFFUE7N3tMc9efdRXP0NE2uvs14uw4tItnLVrIqMe/MYB0BrE/ZWIHx74Vf6qbUJ/FcS3FcyEiNAkiLtnhneCAJYSpBxuWr7nnh3AIj8rGuWwRPE1LIO8dgiVcLIV3nfbDp9qfn3tkQjIGSiY4tUhfA8jIQSgKHOmhEHg+80ZVZ4yGPg5MzwOO/IszhznFTk1XxriQ2LDDnwwPb182RBXDml7NmMwLLIRXwLRwF4Bl1IbAKkhlQFIxty/mzNNkzQME4Qh1+gxRATbWIUZNXtuBlBEt8aNJDfHibRhr4IzE2BISLaFtVazAmZPtwDg3da1c/Pa0SQs9rbr8RGQRhRYCwfFP4bypKUus0NvdFHUAcQajIoZueyLt3xzL3v8E7I6FGCjDRiaACYhJGkyyCrFHQXFHQVFWa1IQYPIsDHMNmRlWaWYVT79Nyu+ePupNJbamZmWIaldJaymulmtgskNO2kSkm1prTGsAR5k0QzXtWkaIF+IXfez8brieqs6EmCjAWZFgCEiogJrpAvKtOcVZZUidu8H2CgWmsI1FfZ+4Sl3vHzi1SefsNcJrfF4vD+iB2KeC/rsxX+ZWRB6MpweJZNGBEEBVErQcLVlK7cG4MMQWFoEEDbWUVLH4/AqC4rijd29HQmGDQKiCmVmT/gVm7cVYze5lPWC/gUsom61uwY6eakZ4XBgMMqp79QiPggAY9kAhT2LwfeTXzHnWATlychpM6hlQEzIyA/6Hg8uG/cNdWWD2m2Pt/qyuWrOYbDF6cgpM+izTRAIOUN7zxpdV2NbITtdkyJmZTE0GaMFS0NMpuV3x126FeiqiN0DiYRxjKa0yk1lo4hZWwxNbLTF0pAAVhvvdZWua2kvaEcyK5uhqffHsBIGWg740U6AhZEWW2t6ULeJmBvYEFH2nc///bQZod2vLS8rFwgLyWSI2bDbAs9dbQEQzMZiCxJhy4pUlMtp5ROfPmnKkXNeOuu6n+a0Q/CqKyAGFiTQ7uT2ZOMQG2/8RlksDUkhVms2QP0QJ/46GPULLTZ+9R8XHhbe89cVZRUOldkWCxbMhtgY+ONkA8nEggOQgYqwNbF8/Jo5ow44b8XZt50zduzY9ng8LhKJ/qtm+PGm11uWH5O1NEA0IPNxR+BTJKsDxGXW8FhOAszlth5gbcwsJYFBqwGgtjbq+TxZZgbnzHs043ZzOggvIr6Nfvf4jADmpQqYf9zXEbZOGdSl4UuIOQJg8P4xJfQPP67CaPfJtP1vSG5dOMJcXPWJ0diSbHbTBIbICIpDIJY0uPzoMAK0AASB7jVw+oM7Ehskg9tyScOGrphTx8AbkkBBMwQdgwVzJmJdYuM2y2ZpkrFglg1N10OSBYd1Dwpyj9PBm9DIRsaTjR+vKoZGd7fxZTVjwLKdjeN3VtUUFLLcBN8EUABzn5W5XxbngVVPV08Mjw61qVA7LDfrB2AlyoLWFDn6rXUAEKoyio3Y//avVEgj2pm4u/tw28BG26GwnBaa+M67AKKoRcpLevXK9BARZSRw8ekP//zvb7S8f+kmajo+Y/LjtAXZ2aFHaQSNhCWt5RWBSOrQcfve9cjJv3swaQpdiwKffkYwQghMjNRUhkygncn33rCSZUFrDzHunbUAUBsF+qmJ199VIAE4cS1eOuu6//7yY79ueGbr0gubufWUnHamKRuSyeU+kjIIsGirCVcu3q1y4p23z/3V7btXVzchDsH1zEQDdLpuTBkJQlO2+RTDCl18uuEHeZztsQHAFqCsxg7VKGYAFjHKLD1gn1nXmUDLAQC1fswpmF2PvN0KKaqKJ+CSRE4zLPENXHXEH3FpqgnLvBypgUdGuH6WhQsXFTB/9idgy2tRGGKxVwbAVDbodiUMAbx8QDcSgWCMQsSuRkf+20jgF4jPCADLBizDAqBrAkiAcE3w7wjKAwaMNbnng9c1N4gwDW0CHinwEGSjWaPMLkeH830k8APEYwEgOXTZAMA1ZbchZM0aVDYAvAoOFpjD7hdxDFhmJZFSDODq43/4vXFjqn7asb5L35ZPLEcN7FwxKrf//Uu7P9Z6+2d+MUtAELqp6/KJ5dgbE1sJ1wIXLnJwAejW6M/OEhU9t9sulAOHVuzdRLgcvRN03SrcTLqOxD0n/eoZCTzzcvOK6h8/M38vVcjv2YZCsCZQhqyTX3/QmOkbrp793XclkdcNEQDHBSjRZ35yjMINc35y4pjqUbKjo6ec9sbEdsJ8bGOaACMB5oaYvOWTP3uNgEsMc/jUh3+yT8FRe7c5mVDQCJOz9fJPTj70/T8cfvH6VdDYHVd1Ph+UKD77e0ra/PfTf570f2/ePQd5BSIhR/RFYWBySOywCiQAmgllluKINGS4uAXkGub8DgCgEbAQh8A3nm/C1XPehC2OhC7ibiAQlNGI2GPBoetBOAtIGsSjFupTfatC+LX1KKWARQ6unftZSLoVBnaPqlwDLa0ZgEAVgK74QAnbhs64Cr0MzeQSfYtt7NJgEZA/wpWzH8a3nn4GC2bZWLdI97ES/A7JSzcR6pIKvz22AhXW3xCWp3tVFiRQhP3mncylHQsLnC93vxpkAh5udD5T/BIcJgzkwycSyCqDoPw2/jT73/hO8vEhy+byo8MoC96MoKxDxunq7DqYbCwiCHJl05ksPDBmj92/HUD7ULbtjQQlTAKJLYNt5ymzQbcbFnht22MNMZlMJnFozR4tABZ5n048DeAafK9nXb0ihWUB4NjdD2ga9rHWJXU8HheJmcvIbUuCxd6nEy+6fwjxqOT6Rk1DcM9RY60EoB7c8OLpaVtFkO2sWzgiMAyQAE8MketQ20EFpQFU2QoBYeAUUU7MJPM5A2bxDgBsrh3HFmqjAomUAeF+WHTUwCtIctlXEftMXDv3z9iw5RIkUgUk4CqjmTGC137YfWFTBr89tgKV9s8g6cfQDLcfFADpXXIxO49BLmuFxwAAliZLyml7EHMTBhGhp5BRW2GJUdBcbGL0STFBBOR9uCr6OVzo12CGW+pn5jLyaicy4FkEV8+OwrKugi0O7FRMDIYthMfILAaGTYSCGA1gyBPwsMFLpkQ2/DyQXw9bToAyA8vGsETYugtXzzkbFz71SOevvmyWJr3onS+bOcfCElchIA9DxnO1MRg2CS+vsAjYwJISBR69TdfETAygHvU0EzO75JkEYrGlg75Dye77dEMMPfcttl2f4/XpywHEYu73DbGY2x5+CKXb/EoJzEx1yTqxaekmcmNUtRi3eRnPiM3g7nX1BgNz/3F2r5LRds81nTEjBsURp8b6RoFaAI1AauY4RqzBgIiRSKmBrKUeaEwZGwJrO7Z8RTsFkBi5Pq2+Q6M6CB5tEznDxNSrsRUEMfXH1GMGSwuUL5h0QAaXA8DS+iRbaPSKsxrrb8jqn0JQEGYAi6ZLQZ2PCWMPw7XR3yCHx/DdVAu6PxR/jk6Hw6dD4JsIyulI+28iMwJSwNFrABKQNKnffj/kMQQNTeqs8BxPoVSRYhtBYM910IL5c25DSF7m9m8q6lYVKBhGQI6GbR7FNXOvApm/4KKn30D3YO2CT1ZB5Y+CEF8H4WwIoJu7ykCSgNKvg+jA4m8SG1hCIM9TEI9aWN/uW1s7C4x41MIPH03jmjm3ICR/jHY9sGyUYViiGkI8hGvmXg3QAlyUWtpjlf6XYyugxJHQ8uuQ+DyEEOhwNARZndeszOsAHYCi75mXW2V4907ZMMyAi0cA8NJS8EG4R4cIf5YgbLsb11NkruJPAAMVZh3CcUYOBPbaZZiezoBtnOrdd9fUPRo/4p7VzxyBvDbs532NAPx40+4hgbAEZXYw3gS4ZIhRAb8NV79g2xZkjHl/1v6LNgJAIgHj1gZzBfA+rpo9H5XBHw5Yldm9AldBBa3DIJAEqw24eu7bAG8FKAzwFBR4H0SsAAoGXatpr9UAIwfQV8B8NaScBK37KkMmgsMMgT2wZc1oJJ7b1P9gShgUS5NuB83r7N8iq74KS5QP2ABQEMExDIKFiPVdZNVluHrOmwBWgakA4jFQ+ekIyAmQAsgqQBvj3WMH1UEbLflrQPgPItZfkO4MzvcFAYA5Aomn/gZAdf92p8DPu7tW/REZOh+2qIEqalm6S2tfdhHrUmTVxbhm7jJczasB5MAYgzztBVtMQsCTjVIGgiQMXNm0FW4GI4Uy66ZuLtD+QXQ4EqkFABS+PYTrcUfGzc03VYeqx45q3dgKhEaGbxICsCoHtOZD6HOGILCqFdiUd/+7q31KDsFxVdhT17TNm3z4FmyHgvq4QQD84sY3v5WFI0GkMIIUct+C2LOMOrN5dgQGQFAarg4o6CKkcCI2tk0il+VlRDANHJN1lPQqNMWSBvG4QO7ZBET+UwhbB3j9bIonwBJJFJRrY1liAiwxAW6eteuqcww8N4bwJi0FW1oQcJAzZ+OyJxfi6jkVRZP0CAQ2BrYMw7FuwrVz/wKtVuDSZ1/dMXF9DJGAwcyYxEXJtZg/59sot29C2nHgEmKKrdzdSSPtuIWAA/IACHEACO79VQYoGAPW7N5fl9mFioCNNudRXPrUJZg/58KB2XokkdUMQV/B/OgK2PwqtHkVFz/dMvxCKALfsrw4uQlXzb4U5YHboRyFTipYv/u43/uKJSAOgKADvPp7gDaAYwwKnmxABIaDSttGh7MQY8efj02bvtB5/v7AJJFVDKKzMX/O2xBYBB1YjMse3+oroP52W4ionIeUWtPy8NeCuc2/b23Oa1Cfjtc7DAYQIIMV6TK82FyBAJnOATGAgACWZ4D71il3WvB3JBiSEBERXr527dpDJl9/fQ71iYFzvj6m8Bsy/r9Ff93/f19Pno1swRCkNaIuPQOEbfBuIUHODiZSEQDFhLEBBxGpSQ9AhhACIKKXAWCslxbhJ+EykAB++GgaDs6EMusRsizwYC3RSIBIQDEjpzQyjkZWaRS0cWMXZHl0cI2IbQG8CVl1Ci578p9oiEkwbXSjbcXETQIFA9jyFFQE7gHEAgDu6rB2GyX1cUdd0i2UeelTN6Oj8DtU2DbAGgO70fz25IyCNsg67j3OKQ3FLl2FSIBZQRKh3LaQdm7Gxs2nenu3QHer8dbP0eH2Sq1CRPwRYfsJGHEMBiPKDDfqkhoNMYnLnr4DHU4C5QELgjWKFUH20UM2qrtsXMurp2xsZNRt2NBxqktt5hZoLp7s6aofAqECYet3qAg+DqjjAXTVS+wHtXBJHsz0Siaz1SqoLQGlm+RIfAq6WZbLTdJRzbK90CI7HPeTdlpkS75FVopWmTWtcku2WW7NeZ9ss72lbSut5ua9T2z82deQSJiBrufjjERyGUkQ3/HOUz9t4bQNiIGY2DsMfyreMyJQYYEGDhcPDYYJ44IFWAPk0jKTKOQZBvQSAGz2ymt1vRgJGMRiEt958h206U9Cm/dQbtlg1m6+ykBX5U1i/gdEYDZg1ghIgYgt4egHkMkdhW89/RiuODno9X1aApsMuOgE6WbUOzqPnHYAyvbdghjsT7Ks+37gjmOgHIJtheg8Zz/n877nIUxu2wryZDrYeWURf3oi5U7Clz71Y3Q4v0HEtiC9CXTgCnfUaQF3fsBgViAQymwLAk3IqAtx8ZPnAstc95zhV5B3G+ag+AqEYYxBVuVRMA5Y9GIvDXLNxN4172All04F9VQ9Opw4QrYFKbZPNu64FeDJRqIVafUtXJT6IuoXuc+wpRcjpx3PoT+AbNggq3IoaMeTwyBwSR6jRseW5HPhNsuyCLAMYPFwfggWK7a50mIeFQAbtlhCsvA+zJIrpeQDym0mSJZCMpH3kQFwNm/W5bf89/3vP1UDpAyKEBQ+tuC4QF3SnLfwf/Z7N7P2bGQcU9Q1PlyndJ9CnlExPC49AJDEPD5UgClyMGawZZHIZHRb0LJfA4CYR+LquWpLei/oD59ehmzuOOT0vYjYEkHLrTYMVp2KgLv9z1cA/u8EQtiSCNsSml9DTp+DC1On4rsvrERDTOLpCnfyIr4FDAHhTZDwPp3KBgRBBCGCCEgbxBV9L88EEbIkbBlA0JJ9PrYMIGTJrnyRYQDDQsiSCEi733MGpI2QJQGUD9s5XZQP6bymqBuHUZd0Y4wXP/lTZPVXQLQZ5bYFSdR5D8E972/ve+wH9ctsC4KyKOjr4RQOw8VPXo94XKAejDgELnv6bRh+EuW2ALPT8/76VZxBkELAEkEEpQ1w71jnINdsufdCDUMlEV9BXfLkL9HhfB6C1qM80Fc2Az3/vmwitgVBBRTUTcg6s3BJ6irE4+77Fo8LXPjMKjA/OoBsXLtJkoAUIQSlDdlHNn1ABI7HIaZUxbZKGXg1FOpMKKOR+FhC04RQnoy77u78EIEUM+1fIQiCyLCbh+p+WMABt8rsxB8//eefUQIG9bUl66k7ksvIAvHDq179QwdnbeIRK6MHoIulVxMC7x4WVDDD49KrsBRXBxxSpn+XHhFMMEggotcO32/Rlngcgjzd2HcSq0tqxOMC30lsBHAmro1+DgI/QEAchYCwPDq461v3nS9EBEmA9JzLOZVDgRsB3IRXOu7F9YtcqkY9qDMhMR4XuDjxBK6c81uUWT+G8KqhMzy/vRe3MqYDTM3IqxZAPNs5Tr+yMmEFsnoRHKPQb5lz0shqC8BbAHYsXyrmFcw01IScftmtDdhf6wMyYBYArdnuc/XGunLGOH4BeT0ZzkDn1QLQzT3G2xPcOQnXJf+GK6ILAf1DAF9CmT0KAKAYMKbL4UdwSRJSuLUAHAMoXo68vgtG3IiLF7qybYhJ1CW0m1oQJ3CCcJW4CDl1P8oDe3XGF/177BjAcAHatAFoQU5vheCNnSO1qg2QfQ55PdalePdzzX5RYrl9uT190CWbOzE/moJWPwThyyizx3bJhrtSIDpl4z3/jgEcsxIFvhta3YSLnlnaVzaey+5avgxZvSfKA/v1kY0ygGIHbNoANCOnmkC0DsCgz3BtfVQkEiljiaqn7UDbXKQdHqmCApoJE8N5LG2L9DD/CECBgUkhYEpEYHWHdusw+BsIEqYjr1eGN132tcd/87ebPvFfi9FPm4iPI6LxqJWqS6pP3fejukdaFn8aGaVZjBxDD/BYeho4oEIgLIaHpaeZMDFUQEgaypti6ZXMlk1gmKcAoLbWfXaBgRNgXWXiU7evrT0OMCcBdCSY9wJQBYMQCA6ADgDrILEUmp6B5TyJC599r/NYxR46P7B7TfRoSDoKjrEg0AxgCww1wcYWwLQCmVZcuCiznTIqoRi635c/zpmIMnkimI+H4QNgMBHEYa+SRxbETSB6D0T/AXEjOgrP4/vPZzuP05n71A3+/f2/aDXK6NNgnoiCyUKKrdC8FeCtsEQTQmjBNHTsUs37usvmhtlj4ciTwHw8GAeBMRHgiJuqiBwEbQXMClc2ohHU/lzn89oQk1ia5H4TdQmM3x5bgZrgqWAzGY7JAWIryGyFhS3IoxkWt0BmO3DhokHiv90OzTFJlNRvrPhqbd4sWZjJ5AyRGDGGFxHw2IYx2FqQ6N5MzgAIC+C1duBf6xyQ7HQdufsxNAchpwYnLFr55TuPomQd+n2OPkZgZqJ6ovu//lTVV5/43Wtbcy2TYciN744wAgQ+b3cLlTaR2sH8JgLgGOJPjG/iieGCKBSxnJjZRCJSpHPmk8fNfP1xn6nnH2Ng9KdYFkYtvIpy2AghGCzAymdwbirXYxs/KddNdCz+sG1LjbLiLKWBr8O18Ib7gR/qvRvO8w7/Of1qBr3v8e9PKEMgGwEHBMpzGZz3bEcfGcajFpAyA96/bb+/vcf/Qci5azz9yWbBrAhy4TJEAgKZjiyaXuzoc43FlFLv4w/1uexfNv1v6tXKY3664qVlP14B2TJaOcLlHw4zGEBQGLzeWoX/NJchKEyfATIIN65WaM67zr8ev7NRsiJkHRjYLb74nJt+OXdh1ErtSouUnY2FUcua95Ta57YvNbyRWx1DWmkWI1v8WhBgNHBwtTCnTZAio3bMavJdelWWwydOaEKx544ZbNtEjmM2i3Ru+tFHv9vGrh9kG0NefiIsag36q5zLoC7WzSATVtFje/DdFktnMOoTvvvuY7ua2inw79/McVzUteLfJ3ebgRcd/R27O2aOY7evUAKo7+Tz7Zr3uKds+r/uWExihiebbVn9j5BsGhpisq7ubv2ftz9xtx3ceGZHOyv058bfQTDcop5tTgAPbxzVp9GAYSAigRdagUc3OBCyT1EYBhsTiZSLE8YefPx9p/22MRqPW6lE4mOnoKIL41ZqXkJF7/nWN59rf/PaQntGgcSIkiAAzzMN8Nd2szA2uONVIQhA3ggcXNVuDqlpFwO49FR5hSXb29W9xx6w5Cz3me2ae7Z3DNTP2nbXnFhK2D70pn+X7nEXPgSyWchRax6l1Ovvfek8GXjzL60t2RGb6BiALRhPbhqFVZkAAqIvvdGAcMMqhdZCX+uJAMOCaVS4esNl008/PDH7/HWxhphMfpziT56H6oJH/98xt695emF7ts1yG5ePbEqFbzXNrBbms+OlGI5YEwCAwSdN3MpVthKqaD091pVVlmxv0+cec8Drf124MCrndbOat9ePyV5I12/7vUu9mCUMA6h0j4viQyCbWq/n1KjQnEfTHVaOBCyM4DgJjD3Kc+hvGjIMRATj6BqXEdGbm8GAIA3TpNom3rji4XuZOZKsS5q4z278iCPmKaZrX7l92n3rX7y7Pd8eJBY+83FEwew2Iz62Wrhxph08oxdrwsRwnqsHVExgKYVsb1eZgLQeAcCNfik9Dx+Lm19CCR83ECUMM8TkyRetElTxdLhMupT3kTgXAMcITAzlUBNQ6B1MFwTkDHBQpcDYsIDRfWddJpLIabVWNx25+62xW5gZCSSKFmj9yCAeF8m6pG5Y2FD+6yV337cx3zwRDjTvhLlZeAy9Q6oFTwgRFQbuvjZkEIC9yrODKDo24YhgZnpq1v6vrGOGSPQKBZWUUwklfETR2BgVgIFA6K9SWDSkxtPbCQYQEBp7lbv9HHrDAAgQo3a07Kz42gckLNNRcNbp5rMOSZ53O9czUz1RnD+aFlSsISaRSJj169eXfXdNwwNr8psPorxRGGECBNBVqqgsAHNsjaT8MLTG8IkQY4KOmRAqCKcIQ8+HcNtF3QH4z2qv33dsOCWUUMKuitralEvJzeXubWtV6wMBMWJV333raWpZFpW28brad0EAyGpgnzJg30oJo4tMPkS27ig4S7Orzj4oed7twV/ZJkEJ81FTUNF41PItpsMfu+zBdWrrXMoZxQPVMx1GELkuvTmjJSq9UkXDYTVpBvYuz8DuJ+7ow60KIWR7m251CvIBoOtZ7Y6P1A0voYQSukAEXrgwah1yyKNpQCfDEQEMd0mtbjAAIlJhenkWivsytIjcHObjRwuEbK9dW/8Dt1V7Ti3NvH/2uFvOuGvL21sqE5Qw0YXxnTJxjzRmLbjATiVS6tpX7p323XXJxrWFzXM5rXaeYoJLgphaTubQKiGyZsdJEL7VNDrgmN0iOSqW1wQARKwjZYIN8wNzD3t1c0NDTFI/7NaSciqhhI8wNm9OuaVghLw5kzbMPHIuIz8Yvld5GhW27mM9ub8Do22gdozlKqeiM5iwTEderVVNZx3y4kWP/Ozx+VNT8xIKC6NWH7bkhwXMhDisRRde73zlsf858uev3/z42tyWWciagTtAjABsAT5pjKTh9PRqBvapyCAoecBmWcwklMNkCboRcJtP9oeSciqhhI8w6uqg43GII2e+9opT4OfKygTgN+sbARgQIlJh34r+rSdBrnvvsCpgb9+9N4iCWpPddNSCNQ8/e/q/fnQSzUspEDjWEPtQ1eKLNcQkiNhOCPWJ+75zwX2rn23ckm3eExmld6Zi8kgQXDtW8rhhIkH4VtPYoGOmlQ1sNTHDhMJCpNPmjcymUSk36bb/lIGSciqhhI84amvdYDOBr5WSRpQY4VtP08szqAm41lOficpz731qnEBFQMAMNEGSsJA3ekumZdJjmxY/dEjD+b9k5kDSawGzy7P54nGBGGSyLqlveuXe6r1u/+KNz7S+saA10xYmBbMzyA8+/JymfaqIj6wWIjtcOU1w0wVmVKZhi4GtJiI2wSBBWrh+3ryUamws3i5l176xJZRQwg6DPTfYonWzwoWm/FuBoJjsFHjE6rW5zD2D5ekyPLO5CoF+ShoZACGvIeGdaxSIivfNAbxEXWPIqgzRZGvUy0eN3f+yu05MPG8At4xWfUrvSlVk4vG4SMxcRqhLahsCs/9x2Wdea1lxeRO3T+cOR0OMfIJtd/hVx6sDMF+bYlFAEvW7cNiO4xaYMDmUN7XjmovmNQE+EYKgtWm1rcLeh+/39pbu5Yp6o2Q5lVDCRxxE4MbGqDx88qIMgRZEIoIG7dG2I+cDUDAC0yJZTA4X0J+bRwDIaWDvCFA7VsIM0qKBAQEhSLXn1PvpDYc/uPqFp6bf8aU/3fzKvyZTwnX1oSEmP+jE3TjHBRogE4mEEXVJ/c3Hf7vv1Ns+3/BC85v3bc01T+e00nArjO9UxQQAFoFPn2AhbNGwsfMYgAXmA6s6BmWiE7EuKxfETLe4iql/IkTvcZdQQgkfYcTjEPX14MXvHTQ2l+e3haBKpbhoUc4dhV9zr7kQxKMba1CsOIVfufxfmxivNSkIq0/tvf53YyOoLIAxKN86KlT5v9/c76y/fO/Qz7YwAMQgY7EYkrEGAxrQyzQ8YCYk64Rfd1EC+O3Ld+2XXPXYd95sef8rrZwNI6uMW+B/5xsEAoAx4NMmST64UuxwYVcffg29Ayo7zKxRbQPU0AMAMBEgBBWE0jMPP2DpcnabLRVdJJWUUwklfEywcGHUmjcvpZ59/cDLq6rl99pa1IiyxPyK5a+2VGFxS7GK5e4kJAloWGewol0PVUG5XYIlSxkJooJDq/eqnPz3OeMO+fP8ORcv7yzQ1gCJpVGKo9Yk+itYvX0XRvH6OCXQKLoXuQ7CwjmpXx/2wrolF2zOt3xlK6fDnC4AEBqED4TAIQgwCpg7TpjoaCnSw6iYvGaC5qQJTSQFD8L8Y1VRaVltLfpvxx70+ld6F3ktdo4SSijhYwDfenpyycwpISneABDRGiNmPXWB8NjG0dhakLCpb3Imw+1hqYD/396ZxtZxXXf8f+6981Yuj5REeZHceIktiZRkW44tS0okBXWQFHGAAqWTILFRIAlSGEiRpI2btEgkpx9Sd0uKNm1TNEYbp61Boki3pFmaiIody07lWCK1WLbi2ollWyvJx7fP3HP6YWYen2gtpPT4SJH3BwgUyJm5d2Ye7v+d5Z6Dfz7GeKMUQGmajkCBABEWhiFNKYMOTpaWpDq+fV3b1f/09fd9Yehm6h4/awWMxAq9PdIPYPDgGpGdO4XO4ZTasWMHPQygv/cQnTh4gnYDwKHdgsHJbEcDwmf2/N3yHx3b+54TtfEHTpXGtk2oKqEUixK1NLbUSCxMdyxR/O5lWpWb5MoDJvs1be8ZlWsz1fP2a4oQIogxsH6N12/sO/g8LmI1xWM4HI5FQtyI8CfDfX+e6zKfaoX15BHjVC2F/znedd6chdANCFSY8C/HLE6U7bnaa1xoIAGJBcQgZZDSCbRz4vVcsn3X1R1Lvnfr0pv2fG3LQy/6tnZZJTIMAF8k8f7v77z5l+Xj7zxeOPOuk5X83RUTdPu+HwbSiCwRqSbkG1wysTCt61L2vcu1ilppNYXYnbe6vcB3LrmoOw+x1TQ+6n9z8/qD90/HaorHcTgci4TYehp5qa+nVKEjWlO77wvNpvU02ZCwAz8bbTunew+I6+8BJSY8fikCFQ8nYIAJmhSSBp4ySPnKT2hzOKVTI93ZjsNJ5R1emcq9lsi2H3vf8vVl75bKaD8md4MOYYhef97vefTF/zIkamWBKzecKI3fGIi9baJa7C0H1Rv8FFHg+0DNAgwbpZ3N+f6rWJjW5pS9d7lWNQHF7tPLJXbn5Tyf71l+hpTChd15AlEaQkS+ZbV245p9RzENqykey+FwLCLq1tOBvs935cwXx0aDgFqwEVQTsPtkF149T88n4GyBGnzN4vViAGWm5+KbCglESCxECAQNTwFGQykFJQqeBRRTVUOVy7Z2JmOSSGgP5aCKqvWp3UsvLwYVTZ5KBhqwYIi1YZmLQABIGMEhzJnr7qz7xWRpotu6Fb+nR1ONmydMQBQjFMg9V52R7oSv/AukjocnSNCRM2ZszP7l5rUjvx1/9qYz1pw/UIfD0VpEQDt3gvr712QmSB/yPFrhV0WiRXZ2xgSgSVC1Bt8/3o1iQDB07hw+BuAR4Avwb28wXspHFtRlTiGsdSqR2SYEQCG2GbUKZ1lfyQkILOKoVjgtAkBEBJJJLZgXxBMRhmxZqmTrEq2q3HA7TRqjyoS7l4zzze1lVeGLfmDYGAKLjCU9Wv2tx4ZPAcDU1hgXGs/hcCwy4m+we4b7PtyeM4/lxwNLNLvVCuKOuScqSfzoxPnjT/GxGqH/53snLfaN2igTu+kdE4XCa9ZLTcTXJyDOQ5/X66QixPvE5N3LtdweVX9o5qQVgAorrGov8p1L8hdLgAgRCTq6jBkftZ/atHbkK3G26HTHnNcP3eFwzB4DA/36vvcP2qeG+36cbdNvLxZsSwQqqRgvTrRjz+n2c1aPaDyWACQV4ekxxo9OWAgESs04DrVgiUsStXvge6/WuCGjwlbrzRwDQFUUrklWeFvPGE0z6c+m0kpXSvaAV03evmHDsxZR1t5MxnU4HIsVAVRCPhH4EmhN0W9mjzjT663tBfR1FnGhTK/IokGFBXd3ET6wwqDDo3ovqMX8zTq+dw6A67PED6w09JZ0tMG2yePUhNBpfN60dJxAmOau5vgo+cQddzzrDw6ev0zRhcZ2OByLlHpyxL6+R7qWmodalRwhCDvj7jnThRcnUkgpvmAgggVIayAfEH5w0uL5fOi3WoxWVOzGI4Txpc1dmgQgvwl9mRqJM/OSiuVXl59BhxfQhWrnxYhI0JkzZvSM/+iW9Qc/MpMkiKnjOxyORYoIaHAQasXGjQmaKDyXSupbymVmmsXkiEYUAU+e7MIrpeTFBQrhXihDhH15xq6TFuVAZisWNe9QFAmxBa7KEN/To/EraVLlaNlv5mJOAKwQNLG8s2dUlqX8acWZRMCJBJG18rplu/bu3kNjmKE7L8a59RyORUy4aPRj03VPl21AHxOBaA1GC9Z6QbgpafPSMazM+KiwuuCCpBA2tKuy4PYOwm9eZ7CmU0MEEBYQLcxv240p4h4g71im+P4VhlakSJWC5qcNNgrT1mVj6En5qjqdBAgAioQTSUW+jwc39R06MzjYP2N3XuM8HA7HIifOpPrJcN8jXd3eQ2Ojfsvce4oEIgpPnOrGqyXvohYUEFoQngI0EY4WBU+ctng9qs+jKBa+K5uG9HAAkNUdJFu6NS1PEVVsc9PEG8cMhUlk67JRXJ2u0TRSxsMJigS5nDFjZ+yjm9aPfGSm2XnnmovD4VjkhO69fnVD/0sqOFTbk86oDa3I3gMmBQqi8OSpLvyilDhvFYmp5wnCvlCBEEbygmdGLc5UQ5GK6/VcaSJVd99FovSWLMmmbo23ZEixALUmx5Zi4hiTiSymGQkTC6cyWvk1PupT7fbNq44UcYnuvMb5OBwOB0SgiMBDz/aubmtX/yuMlO9DzX5h2EmBIhCePp3DzwupaQkUEK7hCkBShbX5DkwIfjZmcbISeScVhdaUzF+hCrf3NogSQa7PktyRU7gxo0gRqMqTxzYbhTArL62sbO0Zk2VJfzqbbAGEO5u1BnseSaGILe+4dfiZ6dbPuxBOnBwOR5169t5I7/2dnd43JvJBAMy+ew+IBAoCTYSfjXXg4HgGCcX1lPKLwRKWSIpF6mhRsC/PeKXEECuACgWA5olQvUmQACQ15OY2kvUdmlamiYiA2RQlINrHxApdns9vXzZKnQlL1WkKEwBAJOjsMubMafvpLetHvny57rwYJ04Oh+Ms4sXlyeG+r3Z3mwdblV4OTMZREkpwON+GZ0fbQBDo85Q6OhccGktIKoBBeL0CHC4wXigwRms8aWrRpHusFWIVi1F9LA7rUpAmXJtWWJUleWsW6E4QMUL3XXzebM6pwgor0xW+e8k4JTWTL9MXJomEaXzMPr6pb+SDzRKmeG4Oh8NRJ44/LVt2gtLLT/8wmzXvyI8HVqnZjz/V54CwksSr5Qz2nO5AOSAklcyovl7oHQvr9BkFlCzhtYrgaFHwcolxqiaTm6SilLdYQIBQRBrnczFo6v8bhS8OkEXq62nCVSmFGzKEG7KEngTBUwKfw5qCs124jxA+H59JVneU5PZcnkAgO419TDEiYrNZrSsVHk7WaNP69cNlXGacaeocHQ6H4yzi+NOe4b7lOqF+agxdVy5ZVopatv1EACQUI+8n8PTpHN6oGCQV1/82k+tI5PLzVGgtVSzhVE1wrCJ4tSI4XhWM+wJrp1z5TYpzoYEi5an/RF1lkprQnVC4KklYmSZckwJyhmCUIIgESQQtSYeP40uGRN7WlZeb2kuqxmpG2X/RfiYF4HSlZu/asvbgz+PPTLPm6cTJ4XCck4EB6Pvugx3at/q2tox5gi1lgkCkVRt0gbhCuYBFYf94Ow7nM1BgGLq0KuWxUCmKN/TGJZII+QA44wtO18KfeV9QsIKKDV1sATP4TUtmGBHTJPCUQkIJMlqh3QCdhrAkQej2gC6P0GbCqhiCsONG0EJBimcKhPGlZcka39mdp6XJGl28WeDZiECMARtDki/Yd2277eCuAenX911CFYjpzNfhcDjeRBxD+PG+3ve2Zc1/+D6zta3J4IuZjEMxflnKYO9oO/K+viQraup1Y9edotCyMnWhIFgAAQM1CVs41ZgRQMPKpHGkELYC8YjhKY0EMTwVtgMJ0+NDEQ0iDyJH5822224qCoAvBAhkVUdR1uUKZJRQbSaJD6hn5tl0RpuxMf+Bd9x68LFmxpkaceLkcDguSLz4PLGv9yOdOe/vyyUbWAvdSoECJt18ZWuwf6wDRwspECS0rJp0fWBSsKhBRFSDdTP1puvhJGn4/5TrnOu8VhBnOtZYoTvh84auPK5JV2fsxpu8ngTtncaMjfmf2bz24J+KbDVEzRcmXMLcHA7HIiRehH4y0vu7uZz3JxP5IGCmOREoRQKPgFfLaewbbcPpmoGnGFG7wFkbdzrMlwU1nkeNCR6JrOooypqOIiV0aC1d2jyjgq6j9kub1478/mxZTDHz5Vk6HI55zi7ZaraHArWzK+ftyM+RQAGTVpTPGkcmsjicz6Bs1Yz2RS1E4hcRu/BWZCqyLlfAkoSvfFb1DMaZIhC/q8vzRkf9v9jUd+CTu3ZtNdu37bYX7Bh5mThxcjgc0yb+tvz0yNovtuf05+fKggLOjkXlAw+Hx9vwUjEFn8OOu4tJpOKHHwiBBdKTrElvZxHXpqtKot9f8gsi8XM5442P2r/Z2DfyoEi/Bga5WSnj5x12Ni/ucDgWHLRr11a9ffvuYM/I2oc7cvoLE3lrmVubJNGIIExKMAScriXwfD6LX5SSoUsrcvfFxy00Gi0lEcjSpC+r2otYmamQUaAa01nHzQQRiCKxHTlj8nn+i41rhj85MADd349ZFybAiZPD4Zg5dYF6amTtQ+0d6pFigdlakJpZVnLTiFdKQwJFwJlaAkcLGbxSTKJsFTSF6eeNx16pxA+YAQSsQBBZnqrJW9tLuDZdJU8J+UyXVbVcBKIUuL1D64m8/aONvSOfGxjo1/39s28xxThxcjgcl8KkQA33fTyV0X8b+Azfl5Zu1J1Ko0hpEuSDBF4upvFKMYmxmoFAYJRccdZUvX0GwpYWVggpZeWadFVubCujJ1kjrXDZogSEG2yVAmWyiiYK/NnNfSOPtMqV14gTJ4fDccnU08yH+349m1LfBCFTLnNLSx2di3gFDd19giprvFFJ4pViGscrHso2zO1rFKrG8+YD5xIkQyzdCV9WZqpYka5QZ8KSAAiaIEpAWJLISyitCLZclo9uWT/yD7OdlXc+nDg5HI7Loi5Q+9duzKRpwEuolYWJ1hWLvRjxou2pMFdtIjA4XkniWCmJUzUPpUBBouKymuSsRbFVYjV1IbYIBUmE4BFLLhHI1akqrklX0Z0IyFNMQSRY5zr/UhCRIJPVxgZyulThD7593YEfxBmaTbj8jHHi5HA4LptYoHb/tHdlplM/ns2qTeNjQSAyN5l852KqNcUgFAODU9UEjlcSOFXVPOEb8oVIQFBR/EpBznsDzdj/xAA4zLKDIOxCm9ZWuhM+epI+elI1dHoBJRQTCyGQSd9ac0QJAojtzBlTKvFwYYLfv/2OA8/PpTABTpwcDkeTiHtBfec7NyWXXp/5q0xWfbQwYcEsLemoOxMahUpHTQ6rlqRgtYxWPZypGYz5HgqBRsUqCgWBomLjMlk9gnDWVp9zW11UL37OAKRu7QgUiSSVSEZbdHgBupO+dHsBdXgBpTWTIqm79JopSPU5CpgI1NGpqViwA7XTtY9u2XJkYq5ceY04cXI4HE2jsTL1MwfX/Zbx8GWtKVUq2Xnj5ptK46IfilX424BJqqxQtkqKgUbBaikGGhWrqWIV15iMzyqwAi1ExAyRyXq0pAkqEi/rkShPMyeJKWVYsspSm8eS1ZaynkWSmDwlRJFFZ2VS0OK5Nf2+RYJUShsW4cCXz23sHfnj8PfNrS5+qThxcjgcTUUEBPQrokH75HDfHcmk+no2q9aNjVoGgFZWNZ8pjW662DJSkKi2XvhXEYIVCIPIClnLUEJElsEM4uhc0iQ6Oo81CcXCRwSK0xfOdumdPfas3SNDiIQ7ckaXy/xircIf27TuwG4Jc0Oa1o/pcpm3HxKHw3FlQgQhGrS7dm01W9Yd2HvmZWwqTvBXMxmlkklSIjKn7qILQQ3/gLB4ayCEGhOqrFBlhZoQOHLxGSU6ZYTS2qLds6ozEZjORGA6vEBnjEVKMxJalCaQgMgXoslrEXymqf0OZ1eYRAIvSZRt17pc5H8sV+SuTesO7N61a6shas3m2uniLCeHwzFrxD2hAOCpkXX3JpP4SiqtbhgfCxtPEM3dnqhmM3VVn0+Lq0SF2zs7tSpX+DUO8Dt3rhl+HAAGBvr1ffc1txdTM5hPz8/hcCxAGt18331qTfeSJeYPlaIHtQLKJRsI5k9G30IjzsRLpbVRBPg1fqxU9H9v6x3Pv97qig8zxX0gHA5HS2jslvr0yPrtJiGPpDPqbcUiI/A5gBOp5iISKEOmrU2jXLIHAqs+u3HN/m+HfwozK+d6ihdiwZjUDodjfnMfDVoRkEi/3rh2/67/fK1rU7FgP6kIxztzxigFEhGL+VWo4YojeoboyBmjNY0WS/YPXjqdvXPjmv3fHhjo1yKg+S5MgLOcHA7HHNAYi9q9d9XVbe2Jh1jk4+mMTk/kLUTERvEot0ZNDxERBki1d2iqlG0A0KNBmb509237Xwbmb2zpfLgX73A45op68VgA2DPSu9ok9UPC8qFMRnuFCQsRmVdVJuYbUUyJiUi3tWtUyiwgDLJPX7qrb/8+IKzesW3bbjtfY0vnw71wh8MxpzQmTADAnuG+dV5CfVoEH8hkVbJYZFjLgQip+bxHqqUIGBBWmky2TaNUtJYU/StAf/a2W/b/FAjjSjt3DsrDD8/9htpLwYmTw+GYF+wQqF70Uz1p4sCta0zCflyYPpTJqiXViqBSYQ53H9GcNTecK2IrCQAlk0ql0gqlgs2LogE/wF9v7h1+LjpOAcB8qPJwOSyql+twOOY/IlCDDSK199CGq5lqHyaFBzxP9WlNKIXWlCUCIotqYa5lAgkLSYgoRSaT0WARVKv8ogI9VrX6G5t7n3sFCON4/f2QK12UYhbmC3U4HFc8OwRqJ/rrmWVf27vBu729eg8J3W8Z78m26U5rBeUSR7Ep0EKwqEQgBGEQhIhMOq1gDKFQtEVF9H0mfqyUmPjv7de/UgmP79fA4IIRpZgr+iU6HI6FjwhoaGgycQIAnjuy/tqA5V4BfkNENmfbdIpZUCkLgkCYSFgkzPab72IVuuvCencQkDZKp1IEbQiFCVtTRHsE8i0P6t9vWx1m3gFhosPQ0G6+UmNKF2NevzSHw+GoI6AB9CsMDiJOQweAvS/03mh9/S5S8mvMuDudVkuMR/BrgmqVYa0wERgCAhGF+3zmbO0TAAIRAYVOO6VJJxIKiSTBBoJSicdJ4RkCvlMDvrtl9ciR+smRy7O/xS3T5wInTg6H44ojtqamWg7PH9uwtFD0N1rL7wRos7D0pjM6awzBWsCvMXxfwAwmEo6vBRCF8avLF67IEgIAIUQihDA2phSU8QiJBEFrgrWCUpErRDgMkqe00j+0vr/nrr5Db1zsXhc6TpwcDscVzY4dUNu2bVXbtu3mqXGXp19Yu8KAbg0CvguMDSC6BYQVqZRKeF64/FkLWCuwgcBagFnCuA+BRQQXs1AiQUPYYANKKUDrUHy0IeiozWLgC8plGxDoVQGOKIXnFMkzvuh9dze466JrqqGhrWpo225+eIHFkqaLEyeHw7FgiPdMDQ2doO3bd7+pFNKBA2sSQZu51q/xTWyxCkw3A7jeClaQoAcknUpRxvMUjAcoRVFzDNR/Tg4W+egEYAaCQODXBMxSFkGeSE4Q0TEQ/k8RjoDwgrX66JI1hV/cTEerU+cebpbtESwCl910cOLkcDgWLDt2QO3c2U9DQyfoXJZVI3v3bsggg84Sl7tSQjkrlBONNhJkWSitgUT9ZAIJUNMkZYEUyVJBFI8r8katwVimLGO33jpcPN9YsWW0bVuPXMkbZWcTJ04Oh2PREFpWIKCfhnCCMAScPLlbGhMsmsnAAPSyZVsJ24Bt6BFgMMrMc5bRxfh/2SL6o2u3mEwAAAAASUVORK5CYII=" alt="Amara NZero">
    <div class="hd-titles">
      <h1>Funil Comercial</h1>
      <div class="hd-meta" id="hdMeta" data-updated="__UPDATED__"></div>
    </div>
  </div>
  <div class="kpibar" id="kpibar"></div>
</header>
<div class="controls">
  <input type="search" id="q" placeholder="Buscar nome, telefone, empresa, CNPJ ou cidade…">
  <div class="msel" id="msel-time"><label>Time</label><button class="msel-btn" type="button">Todos</button><div class="msel-panel"></div></div>
  <div class="msel" id="msel-vend"><label>Vendedor</label><button class="msel-btn" type="button">Todos</button><div class="msel-panel"></div></div>
  <div class="msel" id="msel-uf"><label>UF</label><button class="msel-btn" type="button">Todas</button><div class="msel-panel"></div></div>
  <div class="msel" id="msel-mes"><label>Mês</label><button class="msel-btn" type="button">Todos</button><div class="msel-panel"></div></div>
  <div class="msel" id="msel-orig"><label>Origem</label><button class="msel-btn" type="button">Todas</button><div class="msel-panel"></div></div>
  <button type="button" class="clear-filters-btn" id="clear-filters">Limpar filtros</button>
</div>
<div class="board" id="board"></div>
<footer>
  <div class="legend">
    <span><i class="dot" style="background:var(--i5)"></i> Intenção 5</span>
    <span><i class="dot" style="background:var(--i4)"></i> Intenção 4</span>
    <span><i class="dot" style="background:var(--i3)"></i> Intenção 3</span>
    <span><i class="dot" style="background:var(--erp)"></i> Confirmado no ERP</span>
  </div>
  <strong>Colunas:</strong>
  <b>Entrar em contato</b> · <b>Atrasados</b> · <b>Mencionou orçamento</b> · <b>Fecharam pedido</b> (conversa; fluig/repasse excluído) · <b>Perdidos</b> · <b>Entrou em contato</b> (atendente humano respondeu) · <b>Confirmado no ERP</b> (CNPJ cruzado com relatório de pedidos — independente da conversa).
  <br>Filtro de mês pré-selecionado no mês vigente. Para ver histórico, mude o select de Mês.
</footer>
<script>
const DATA=__DATA__;
// ---- normaliza UF que vieram por extenso em vez da sigla (ex.: "Bahia"/"bahia" -> "BA") ----
const UF_FULLNAME_TO_SIGLA={'bahia':'BA'};
DATA.forEach(c=>{
  if(c.uf){
    const sigla=UF_FULLNAME_TO_SIGLA[c.uf.trim().toLowerCase()];
    if(sigla) c.uf=sigla;
  }
});
function startOfWeek(d){const x=new Date(d);const wd=(x.getDay()+6)%7;x.setHours(0,0,0,0);x.setDate(x.getDate()-wd);return x;}
function fmt(d){return String(d.getDate()).padStart(2,'0')+'/'+String(d.getMonth()+1).padStart(2,'0');}
function parseISO(s){const[y,m,dd]=s.split('-').map(Number);return new Date(y,m-1,dd);}
const NOW=new Date();
const WEEK_START=startOfWeek(NOW);const WEEK_END=new Date(WEEK_START);WEEK_END.setDate(WEEK_START.getDate()+6);
const MONTH_START=new Date(NOW.getFullYear(),NOW.getMonth(),1);
const TODAY_START=new Date(NOW.getFullYear(),NOW.getMonth(),NOW.getDate());
const MES_NOMES=['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'];
function renderHdMeta(){
  const el=document.getElementById('hdMeta');
  if(!el)return;
  const updated=el.dataset.updated||'';
  const divisao=f.time.length===1?f.time[0]:(f.time.length>1?f.time.length+' times':'Todos os times');
  el.innerHTML='Semana: '+fmt(WEEK_START)+' – '+fmt(WEEK_END)
    +'<span class="dot-sep"></span><b>'+esc(divisao)+'</b>'
    +'<span class="dot-sep"></span>Atualizado em: '+esc(updated);
}
// atrasado = prazo de contato ja passou (comparado a HOJE, nao ao inicio da semana)
function isOverdue(c){if(!c.contact_date_iso)return false;const d=parseISO(c.contact_date_iso);return d<TODAY_START&&d>=MONTH_START;}
function isStale(c){if(!c.contact_date_iso)return false;return parseISO(c.contact_date_iso)<MONTH_START;}
// ---- reclassificacao por sinais de conversa, por cima do 'stage' bruto vindo dos dados ----
// intent_score>=3 = cliente sinalizou intencao real de fechar ("Como proceder", "Desconto/negociacao",
// "Quer fechar negocio", "Compra imediata") -> precisa de contato ativo, mesmo que tenha sumido depois.
// mencionou_orcamento = cliente pediu orcamento/cotacao explicitamente (ex.: "Quero uma cotacao agora").
function baseStage(c){
  if(c.stage==='FECHADO'||c.stage==='PERDIDO'||c.stage==='SUPORTE')return c.stage;
  if(c.intent_score>=3)return 'CONTATAR';
  if(c.mencionou_orcamento)return 'ORCAMENTO';
  return c.stage;
}
function funnelStage(c){
  const s=baseStage(c);
  if(s==='SEM_SINAL')return 'HIDE';
  if(s==='CONTATAR'){if(isStale(c))return 'HIDE';if(isOverdue(c))return 'ATRASADO';return 'CONTATAR';}
  return s;
}
function inColumn(c,s){
  if(s==='ERP') return c.erp_match;
  if(kanbanState[c.uuid]) return kanbanState[c.uuid]===s;
  return funnelStage(c)===s;
}
const STAGES=[
  {id:'CONTATAR',nome:'Entrar em contato',v:'call'},
  {id:'ATRASADO',nome:'Atrasados',v:'over'},
  {id:'ORCAMENTO',nome:'Mencionou orçamento',v:'orc'},
  {id:'FECHADO',nome:'Fecharam pedido',v:'won'},
  {id:'PERDIDO',nome:'Perdidos',v:'lost'},
  {id:'ENTROU',nome:'Entrou em contato',v:'done'},
  {id:'SUPORTE',nome:'Pedido suporte/atendimento',v:'sup'},
  {id:'ERP',nome:'Confirmado no ERP',v:'erp',sep:true},
];
const ICOLOR={5:'var(--i5)',4:'var(--i4)',3:'var(--i3)'};
const shown=Object.fromEntries(STAGES.map(s=>[s.id,15]));
// ---- populate selects ----
const MES_PT={'01':'Jan','02':'Fev','03':'Mar','04':'Abr','05':'Mai','06':'Jun','07':'Jul','08':'Ago','09':'Set','10':'Out','11':'Nov','12':'Dez'};
const CUR_MONTH=(()=>{const d=new Date();return d.getFullYear()+'-'+String(d.getMonth()+1).padStart(2,'0');})();
// ---- filter state: vend/uf/mes/orig sao arrays (multi-selecao); time/q sao unicos ----
let f={q:'',vend:[],uf:[],mes:[CUR_MONTH],orig:[],time:[]};

function mesLabel(m){const[y,mm]=m.split('-');return MES_PT[mm]+'/'+y;}

// ---- componente de multi-selecao (dropdown com checkboxes) ----
function buildMsel(id,getOptions,filterKey,fallbackFmt){
  const root=document.getElementById('msel-'+id);
  const btn=root.querySelector('.msel-btn');
  const panel=root.querySelector('.msel-panel');
  function getLabelFor(v){const opts=getOptions();const o=opts.find(o=>o.v===v);return o?o.label:(fallbackFmt?fallbackFmt(v):v);}
  function renderPanel(){
    const opts=getOptions();
    panel.innerHTML='<div class="msel-actions"><button type="button" data-act="all">Marcar todos</button><button type="button" data-act="none">Limpar</button></div>'
      +opts.map(o=>'<label class="msel-opt"><input type="checkbox" value="'+o.v+'" '+(f[filterKey].includes(o.v)?'checked':'')+'> '+o.label+'</label>').join('');
    panel.querySelector('[data-act="all"]').onclick=()=>{
      f[filterKey]=opts.map(o=>o.v);renderPanel();updateBtn();
      if(id==='time'){const pool=new Set(DATA.filter(c=>f.time.includes(c.team)).map(c=>c.vendedor));f.vend=f.vend.filter(v=>pool.has(v));mselVend.updateBtn();}
      render();
    };
    panel.querySelector('[data-act="none"]').onclick=()=>{
      f[filterKey]=[];renderPanel();updateBtn();
      if(id==='time'){mselVend.updateBtn();}
      render();
    };
    panel.querySelectorAll('input[type=checkbox]').forEach(cb=>{
      cb.addEventListener('change',()=>{
        if(cb.checked){if(!f[filterKey].includes(cb.value))f[filterKey].push(cb.value);}
        else{f[filterKey]=f[filterKey].filter(v=>v!==cb.value);}
        updateBtn();
        if(id==='time'){
          // ao trocar Time, remove da selecao de vendedor quem nao pertence mais aos times escolhidos
          const pool=f.time.length?new Set(DATA.filter(c=>f.time.includes(c.team)).map(c=>c.vendedor)):null;
          if(pool) f.vend=f.vend.filter(v=>pool.has(v));
          mselVend.updateBtn();
        }
        render();
      });
    });
  }
  function updateBtn(){
    const n=f[filterKey].length;
    const allLabel=(id==='vend'||id==='mes'||id==='time')?'Todos':'Todas';
    btn.textContent = n===0 ? allLabel : (n===1 ? getLabelFor(f[filterKey][0]) : n+' selecionados');
  }
  btn.addEventListener('click',e=>{
    e.stopPropagation();
    const wasOpen=root.classList.contains('open');
    document.querySelectorAll('.msel.open').forEach(m=>m.classList.remove('open'));
    if(!wasOpen){renderPanel();root.classList.add('open');}
  });
  document.addEventListener('click',e=>{if(!root.contains(e.target))root.classList.remove('open');});
  updateBtn();
  return {renderPanel,updateBtn};
}

const mselTime=buildMsel('time',()=>[
  {v:'Digital',label:'Digital'},{v:'Consultivo',label:'Consultivo'},
  {v:'Grandes Usinas/BESS',label:'Grandes Usinas/BESS'},{v:'Outros',label:'Outros'},
  {v:'S/ Vendedor',label:'S/ Vendedor'}
],'time');
const mselVend=buildMsel('vend',()=>{
  const pool=f.time.length?DATA.filter(c=>f.time.includes(c.team)):DATA;
  return [...new Set(pool.map(c=>c.vendedor))].sort().map(v=>({v,label:v}));
},'vend');
const mselUf=buildMsel('uf',()=>[...new Set(DATA.map(c=>c.uf).filter(Boolean))].sort().map(v=>({v,label:v})),'uf');
function ddmmyyyyToMonth(s){
  if(!s) return null;
  const p=s.split('/');
  return p.length===3 ? p[2]+'-'+p[1].padStart(2,'0') : null;
}
function filterMonthOf(c,colId){
  if(colId==='ERP') return ddmmyyyyToMonth(c.order_data) || c.last_month;
  if(c.stage==='FECHADO') return ddmmyyyyToMonth(c.ev_date) || c.last_month;
  return c.last_month;
}
const mselMes=buildMsel('mes',()=>{
  const months=new Set([CUR_MONTH]);
  DATA.forEach(c=>{
    if(c.last_month)months.add(c.last_month);
    const fm=filterMonthOf(c,null); if(fm)months.add(fm);
    const erpm=ddmmyyyyToMonth(c.order_data); if(erpm)months.add(erpm);
  });
  return [...months].sort().reverse().map(v=>({v,label:mesLabel(v)}));
},'mes',mesLabel);
const mselOrig=buildMsel('orig',()=>[{v:'Inbound',label:'Inbound'},{v:'Disparo',label:'Disparo'}],'orig');

function filtraBase(){
  const q=f.q.toLowerCase();
  return DATA.filter(c=>
    (f.vend.length===0||f.vend.includes(c.vendedor))&&(f.time.length===0||f.time.includes(c.team))&&
    (f.uf.length===0||f.uf.includes(c.uf))&&
    (f.orig.length===0||f.orig.includes(c.origem))&&
    (!q||c.name.toLowerCase().includes(q)||c.urn.includes(q)||(c.empresa||'').toLowerCase().includes(q)||(c.cnpj||'').includes(q)||(c.cidade||'').toLowerCase().includes(q)));
}
// usado pelos KPIs do topo (visao geral, sem ser especifica de uma coluna)
function filtra(){
  return filtraBase().filter(c=>f.mes.length===0||f.mes.includes(filterMonthOf(c,null)));
}
function esc(s){return String(s).replace(/[&<>"]/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[m]));}
function debounce(fn,ms){let t;return(...args)=>{clearTimeout(t);t=setTimeout(()=>fn(...args),ms);};}
function weekOf(iso){const d=parseISO(iso);const s=startOfWeek(d);const e=new Date(s);e.setDate(s.getDate()+4);return fmt(s)+'–'+fmt(e);}
function intentBar(score,priority){
  const p=(priority||'').toLowerCase();
  const priClass=p.includes('alta')?'alta':p.includes('media')||p.includes('média')?'media':'baixa';
  const priLabel=p.includes('alta')?'Alta':p.includes('media')||p.includes('média')?'Média':p?'Baixa':'';
  const barColor=priClass==='alta'?'var(--over)':priClass==='media'?'var(--call)':'var(--warm,#A8860B)';
  const pct=Math.min(score,100);
  return '<div class="intent-bar-wrap">'
    +'<div class="intent-label">'
    +'<span style="color:var(--ink-soft)">Score Weni: <b>'+score+'</b></span>'
    +(priLabel?'<span class="pri '+priClass+'">'+priLabel+'</span>':'')
    +'</div>'
    +'<div class="intent-bar"><i style="width:'+pct+'%;background:'+barColor+'"></i></div>'
    +'</div>';
}
function tempInfo(basis){
  if(!basis) return null;
  const low=basis.toLowerCase();
  if(low.includes('quente')) return {label:'Quente',cls:'quente'};
  if(low.includes('morno')) return {label:'Morno',cls:'morno'};
  if(low.includes('frio')) return {label:'Frio',cls:'frio'};
  return null;
}
function card(c,v,isErpCol){
  const d=document.createElement('article');
  d.className='card';d.tabIndex=0;d.style.setProperty('--tier','var(--'+v+')');
  const fs=funnelStage(c);
  const movidoManual=!isErpCol && kanbanState[c.uuid] && kanbanState[c.uuid]!==fs;
  let dateHtml='';
  const _temp=tempInfo(c.date_basis);
  const tempHtml=_temp?'<div class="temp-chip '+_temp.cls+'" title="Cadência de follow-up sugerida com base no engajamento da conversa (não é a intenção de compra)">'+_temp.label+'</div>':'';
  if(!isErpCol&&(fs==='CONTATAR'||fs==='ATRASADO')&&c.contact_date){
    dateHtml='<div class="datebox"><div class="when">Contatar até '+esc(c.contact_date)+'</div>'
      +'<div class="week">Semana de '+esc(weekOf(c.contact_date_iso))+'</div>'
      +'<div class="basis">'+esc(c.date_basis)+'</div></div>';
  }
  let orderHtml='';
  if(c.erp_match&&(isErpCol||c.stage==='FECHADO')){
    orderHtml='<div class="orderbox"><div class="ped">Pedido '+esc(c.order_pedido)+'</div>'
      +'<div class="st">'+esc(c.order_status)+' · '+esc(c.order_data)+'</div></div>';
  }
  // ---- evidencia de fechamento: comprovante enviado (alta confianca) vs. cliente so falou em texto
  // (valor pode estar errado/inferido, ex.: erro de casas decimais ou alucinacao) vs. sem valor nenhum ----
  let valorHtml='';
  if(c.stage==='FECHADO'){
    const descLow=(c.ev_desc||'').toLowerCase();
    const hasReceipt=descLow.includes('comprovante');
    const byClient=(c.close_kind||'').toLowerCase().includes('cliente');
    if(c.valor_negocio!=null){
      const vcls=hasReceipt?'confirmado':'a-confirmar';
      valorHtml='<div class="valuebox '+vcls+'">'
        +'<div class="vb-amount">'+fmtBRL(c.valor_negocio)+'</div>'
        +'<div class="vb-tag">'+(hasReceipt?'Comprovante enviado':'Valor citado em texto - Confirmar na conversa')+'</div>'
        +'</div>';
    } else {
      valorHtml='<div class="valuebox sem-valor"><div class="vb-tag">Valor não identificado - Confirmar na conversa</div></div>';
    }
    if(!byClient){
      valorHtml+='<div class="deal-tag neutro">Confirmado pelo atendente</div>';
    }
    if(c.comercial_suporte){
      const sup=c.comercial_suporte;
      const supCls=sup.includes('refez')?'refez':(sup.includes('sozinho')?'sozinho':'apoio');
      valorHtml+='<div class="suporte-tag '+supCls+'"'+(c.comercial_evid?' title="'+esc(c.comercial_evid)+'"':'')+'>'+esc(sup)+'</div>';
    }
  }
  const stageLabel={CONTATAR:'Entrar em contato',ENTROU:'Entrou',ORCAMENTO:'Orçamento',FECHADO:'Fechou',PERDIDO:'Perdido',SUPORTE:'Suporte/atendimento',SEM_SINAL:'Sem sinal'};
  let badges=[];
  if(c.uf)badges.push('<span class="b uf">'+esc(c.uf)+'</span>');
  if(c.vendedor)badges.push('<span class="b sell">'+esc(c.vendedor)+'</span>');
  if(c.erp_match&&!isErpCol&&c.stage!=='FECHADO')badges.push('<span class="b erp">ERP</span>');
  if(isErpCol)badges.push('<span class="b">Conversa: '+esc(stageLabel[c.stage]||c.stage)+'</span>');
  if(c.origem==='Inbound')badges.push('<span class="b inb">Inbound</span>');
  if(c.mencionou_orcamento&&fs!=='ORCAMENTO'&&!isErpCol)badges.push('<span class="b orc">Tem orçamento</span>');
  if(movidoManual)badges.push('<span class="b" style="background:var(--brand-md);color:#fff;font-weight:700">Movido manualmente</span>');
  const iscore=c.intent_score?'<div class="iscore" style="background:'+ICOLOR[c.intent_score]+'" title="Intenção '+c.intent_score+'/5">'+c.intent_score+'</div>':'';
  d.innerHTML='<button type="button" class="card-menu-btn" data-act="menu">⋮</button>'
    +'<div class="card-menu" data-menu>'
    +'<button type="button" data-act="copy">Copiar dados</button>'
    +'<button type="button" data-act="move">Mover para...</button>'
    +'<div class="submenu" data-submenu>'
    +STAGES.filter(s=>s.id!=="ERP").map(s=>'<button type="button" data-move-to="'+s.id+'" style="padding-left:18px;font-size:12px">'+s.nome+'</button>').join('')
    +'</div>'
    +'</div>'
    +'<div class="cardtop"><div>'
    +'<div class="nm">'+esc(c.name)+'</div>'
    +(c.empresa?'<div class="tel">'+esc(c.empresa)+'</div>':'')
    +'<div class="tel">'+esc(c.urn)+(c.cidade?' · '+esc(c.cidade):'')+(c.uf?'/'+esc(c.uf):'')+'</div>'
    +(c.cnpj?'<div class="tel">'+esc(c.cnpj)+'</div>':'')
    +'<div class="uuid">'+esc(c.uuid)+'</div></div>'+iscore+'</div>'
    +tempHtml+dateHtml+orderHtml+valorHtml
    +'<div class="badges">'+badges.join('')+'</div>'
    +(c.weni_intent_score>0?intentBar(c.weni_intent_score,c.weni_intent_priority):'')
    +'<div class="reason">últ. msg '+esc(c.last)+(c.client_waiting?' · cliente aguarda':'')+'</div>'
    +'<div class="evid">'+(c.ev_desc?'<div class="why">'+esc(c.ev_desc)+' · '+esc(c.ev_date)+'</div><q>"'+esc(c.ev_text)+'"</q>':'<q>Sem evidência de fechamento na conversa.</q>')
    +(c.intent_score?'<div class="extra">Intenção: <b>'+c.intent_score+'/5</b> — '+esc(c.intent_desc)+'</div>':'')
    +'<div class="extra">Vendedor: '+esc(c.vendedor)+(c.cnpj?' · CNPJ: '+esc(c.cnpj):'')+(c.erp_match?' · Pedido ERP: '+esc(c.order_pedido)+' ('+esc(c.order_status)+')':'')+'</div></div>';
  const t=()=>d.classList.toggle('open');
  let _dragged=false;
  d.dataset.uuid=c.uuid;
  // draggable comeca DESLIGADO. So ligamos no mousedown se o clique NAO
  // comecou em um campo de texto selecionavel -- assim duplo-clique e
  // selecao com o mouse (nome, telefone, CNPJ, etc) funcionam normalmente,
  // e arrastar pelas bordas/areas sem texto ainda move o card.
  d.setAttribute('draggable','false');
  const TEXT_SEL='.nm,.tel,.reason,.evid,.uuid,.extra,.why,q,.badges,.b';
  d.addEventListener('mousedown',e=>{
    d.setAttribute('draggable', e.target.closest(TEXT_SEL) ? 'false' : 'true');
  });
  d.addEventListener('dblclick',e=>{
    // duplo clique sempre seleciona a palavra/texto, nunca abre o card
    e.stopPropagation();
  });
  d.addEventListener('dragstart',e=>{e.dataTransfer.setData('uuid',c.uuid);d.classList.add('dragging');_dragged=true;});
  d.addEventListener('dragend',()=>{d.classList.remove('dragging');setTimeout(()=>{_dragged=false;},50);});
  d.addEventListener('click',e=>{
    if(e.target.closest('[data-act],[data-move-to]'))return; // clique no menu nao abre/fecha o card
    if(window.getSelection&&window.getSelection().toString().length>0)return; // nao fecha se acabou de selecionar texto
    if(!_dragged)d.classList.toggle('open');
  });
  d.addEventListener('keydown',e=>{if(e.key==='Enter'||e.key===' '){e.preventDefault();d.classList.toggle('open');}});

  // ---- menu de copiar/mover ----
  const menuBtn=d.querySelector('[data-act="menu"]');
  const menu=d.querySelector('[data-menu]');
  const submenu=d.querySelector('[data-submenu]');
  menuBtn.addEventListener('click',e=>{
    e.stopPropagation();
    document.querySelectorAll('.card-menu.open').forEach(m=>{if(m!==menu)m.classList.remove('open');});
    submenu.classList.remove('open');
    menu.classList.toggle('open');
  });
  menu.querySelector('[data-act="copy"]').addEventListener('click',e=>{
    e.stopPropagation();
    const linhas=[c.name,c.urn,c.empresa,c.cnpj,c.vendedor].filter(Boolean);
    const txt=linhas.join(' | ');
    navigator.clipboard.writeText(txt).then(()=>showToast('Dados copiados: '+txt)).catch(()=>showToast('Não foi possível copiar'));
    menu.classList.remove('open');
  });
  menu.querySelector('[data-act="move"]').addEventListener('click',e=>{
    e.stopPropagation();
    submenu.classList.toggle('open');
  });
  submenu.querySelectorAll('[data-move-to]').forEach(btn=>{
    btn.addEventListener('click',async e=>{
      e.stopPropagation();
      const novoStage=btn.dataset.moveTo;
      kanbanState[c.uuid]=novoStage;
      menu.classList.remove('open');
      render();
      await pushKanban(c.uuid,novoStage);
      showToast('Card movido para "'+(STAGES.find(s=>s.id===novoStage)||{}).nome+'"');
    });
  });
  document.addEventListener('click',e=>{if(!d.contains(e.target))menu.classList.remove('open');});

  return d;
}

function showToast(msg){
  const t=document.getElementById('toast');
  if(!t)return;
  t.textContent=msg;
  t.classList.add('show');
  clearTimeout(t._h);
  t._h=setTimeout(()=>t.classList.remove('show'),2200);
}
function sortFn(s){
  if(s==='ATRASADO')return (a,b)=>(b.intent_score-a.intent_score)||((a.contact_date_iso||'9')>(b.contact_date_iso||'9')?1:-1);
  if(s==='CONTATAR')return (a,b)=>((a.contact_date_iso||'9')>(b.contact_date_iso||'9')?1:-1)||(b.intent_score-a.intent_score);
  if(s==='ERP')return (a,b)=>(b.order_data||'').split('/').reverse().join('').localeCompare((a.order_data||'').split('/').reverse().join(''));
  return (a,b)=>(b.intent_score-a.intent_score)||b.last_iso.localeCompare(a.last_iso);
}
function render(){
  const base=filtraBase();
  const board=document.getElementById('board');board.innerHTML='';
  STAGES.forEach(s=>{
    const items=base.filter(c=>(f.mes.length===0||f.mes.includes(filterMonthOf(c,s.id)))&&inColumn(c,s.id)).sort(sortFn(s.id));
    const col=document.createElement('section');
    if(s.sep)col.className='col-sep';
    col.innerHTML='<div class="col-head" style="--tier:var(--'+s.v+')"><h2>'+s.nome+'</h2><span class="count">'+items.length+'</span></div>';
    const body=document.createElement('div');body.className='col-body';
    if(!items.length)body.insertAdjacentHTML('beforeend','<div class="empty">Nenhum contato.</div>');
    items.slice(0,shown[s.id]).forEach(c=>body.appendChild(card(c,s.v,s.id==='ERP')));
    if(s.id!=='ERP'){
      body.addEventListener('dragover',e=>{e.preventDefault();body.classList.add('col-drop-over');});
      body.addEventListener('dragleave',e=>{if(!body.contains(e.relatedTarget))body.classList.remove('col-drop-over');});
      body.addEventListener('drop',async e=>{
        e.preventDefault();body.classList.remove('col-drop-over');
        const uuid=e.dataTransfer.getData('uuid');
        if(!uuid)return;
        kanbanState[uuid]=s.id;
        render();
        await pushKanban(uuid,s.id);
      });
    }
    col.appendChild(body);
    if(items.length>shown[s.id]){
      const b=document.createElement('button');b.className='more';
      b.textContent='Mostrar mais ('+( items.length-shown[s.id])+' restantes)';
      b.onclick=()=>{shown[s.id]+=15;render();};
      col.appendChild(b);
    }
    board.appendChild(col);
  });
  renderHdMeta();
  renderKPI(filtra());
}
function fmtBRL(v){return 'R$ '+v.toLocaleString('pt-BR',{minimumFractionDigits:2,maximumFractionDigits:2});}
function renderKPI(data){
  const el=document.getElementById('kpibar');
  if(!el)return;
  const total=data.length;
  const fechados=data.filter(c=>funnelStage(c)==='FECHADO').length;
  const perdidos=data.filter(c=>funnelStage(c)==='PERDIDO').length;
  const entrouContato=data.filter(c=>funnelStage(c)==='ENTROU').length;
  const aContatar=data.filter(c=>funnelStage(c)==='CONTATAR').length;
  const atrasados=data.filter(c=>funnelStage(c)==='ATRASADO').length;
  const erpConfirmados=data.filter(c=>c.erp_match).length;
  const pctFechados=total?((fechados/total)*100):0;
  const pctPerdidos=total?((perdidos/total)*100):0;
  const cards=[
    {label:'Total de contatos',value:total.toLocaleString('pt-BR'),sub:'',color:'var(--brand-dk)',pct:100},
    {label:'% Fechados',value:pctFechados.toFixed(1)+'%',sub:fechados+' de '+total,color:'var(--won)',pct:Math.min(pctFechados,100)},
    {label:'% Perdidos',value:pctPerdidos.toFixed(1)+'%',sub:perdidos+' de '+total,color:'var(--over)',pct:Math.min(pctPerdidos,100)},
    {label:'A contatar',value:aContatar.toLocaleString('pt-BR'),sub:'',color:'var(--call)',pct:total?Math.min((aContatar/total)*100,100):0},
    {label:'Atrasados',value:atrasados.toLocaleString('pt-BR'),sub:'',color:'var(--over)',pct:total?Math.min((atrasados/total)*100,100):0},
    {label:'Confirmado no ERP',value:erpConfirmados.toLocaleString('pt-BR'),sub:'',color:'var(--erp)',pct:total?Math.min((erpConfirmados/total)*100,100):0},
  ];
  el.innerHTML=cards.map(k=>
    '<div class="kpi-card"><div class="kc-label">'+esc(k.label)+'</div><div class="kc-value" style="color:'+k.color+'">'+esc(k.value)+'</div>'
    +(k.sub?'<div class="kc-sub">'+esc(k.sub)+'</div>':'')+'<div class="kc-bar" style="width:'+k.pct+'%;background:'+k.color+'"></div></div>'
  ).join('');
}
// ---- KANBAN drag-and-drop via Supabase ----
const SB_URL='https://uuhxsafnhiacgmiiwkhz.supabase.co';
const SB_KEY='eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InV1aHhzYWZuaGlhY2dtaWl3a2h6Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODE1NDg5MDAsImV4cCI6MjA5NzEyNDkwMH0.Y7i4f60vK67DlSHJB2CrTiW31VtNWHCQmEGTte0_GT4';
let kanbanState={};

async function loadKanban(){
  try{
    const r=await fetch(SB_URL+'/rest/v1/kanban_posicoes?select=lead_id,coluna',
      {headers:{'apikey':SB_KEY,'Authorization':'Bearer '+SB_KEY}});
    if(r.ok){const rows=await r.json();kanbanState={};rows.forEach(row=>{kanbanState[row.lead_id]=row.coluna;});}
  }catch(e){console.error('Erro ao carregar kanban:',e);}
}

async function pushKanban(uuid,coluna){
  const sav=document.getElementById('saving');if(sav)sav.classList.add('show');
  try{
    const r=await fetch(SB_URL+'/rest/v1/kanban_posicoes?on_conflict=lead_id',{method:'POST',
      headers:{'apikey':SB_KEY,'Authorization':'Bearer '+SB_KEY,'Content-Type':'application/json','Prefer':'resolution=merge-duplicates,return=minimal'},
      body:JSON.stringify({lead_id:uuid,coluna:coluna})});
    if(!r.ok){
      const txt=await r.text();
      console.error('Erro ao salvar (status '+r.status+'):',txt);
      showToast('Não foi possível salvar a posição (status '+r.status+')');
    }
  }catch(e){
    console.error('Erro ao salvar:',e);
    showToast('Sem conexão com o servidor de posições');
  }
  if(sav)sav.classList.remove('show');
}

document.getElementById('q').addEventListener('input',debounce(e=>{f.q=e.target.value;render();},200));
document.getElementById('clear-filters').addEventListener('click',()=>{
  f={q:'',vend:[],uf:[],mes:[CUR_MONTH],orig:[],time:[]};
  document.getElementById('q').value='';
  mselTime.updateBtn();mselVend.updateBtn();mselUf.updateBtn();mselMes.updateBtn();mselOrig.updateBtn();
  render();
});
render();
// carrega as posicoes manuais salvas no Supabase e re-renderiza para refletir o kanban salvo
loadKanban().then(render);
</script>
<div class="saving" id="saving">Salvando...</div>
<div class="toast" id="toast"></div>
</body>
</html>"""



def upload_to_github(html_path, token, repo="Marketing-Amara/Weni", filename="index.html"):
    """Sobe o HTML gerado para o GitHub Pages via API — sem precisar do Git instalado."""
    import base64, urllib.request, urllib.error, json as _json
    if not token:
        print("  GITHUB_TOKEN nao configurado — pulando envio ao GitHub.")
        print("  Configure com:  setx GITHUB_TOKEN \"seu_token\"")
        return
    if not os.path.exists(html_path):
        print("  HTML nao encontrado em %s — pulando." % html_path)
        return
    with open(html_path, "rb") as f:
        content_b64 = base64.b64encode(f.read()).decode("utf-8")
    api_url = "https://api.github.com/repos/%s/contents/%s" % (repo, filename)
    headers = {
        "Authorization": "token %s" % token,
        "Accept": "application/vnd.github+json",
        "Content-Type": "application/json",
        "User-Agent": "funil-weni-script",
    }
    sha = None
    try:
        req = urllib.request.Request(api_url, headers=headers)
        with urllib.request.urlopen(req) as r:
            sha = _json.loads(r.read().decode())["sha"]
    except urllib.error.HTTPError as e:
        if e.code != 404:
            print("  Aviso ao buscar SHA: %d" % e.code)
    body = {"message": "Atualiza funil comercial", "content": content_b64}
    if sha:
        body["sha"] = sha
    try:
        data = _json.dumps(body).encode("utf-8")
        req = urllib.request.Request(api_url, data=data, headers=headers, method="PUT")
        with urllib.request.urlopen(req) as r:
            print("  Publicado: https://marketing-amara.github.io/Weni/")
    except urllib.error.HTTPError as e:
        msg = e.read().decode()
        print("  Erro ao subir para o GitHub (%d): %s" % (e.code, msg[:200]))

if __name__ == "__main__":
    main()
